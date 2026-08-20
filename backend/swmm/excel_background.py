from importlib import import_module

from django.db import transaction
from openpyxl import load_workbook


TRUE_VALUES = {"1", "true", "yes"}


def _import_value(field, value):
    if value is None or value == "":
        return None
    many_related_field_name = "ManyRelatedField"
    if field.__class__.__name__ == many_related_field_name and isinstance(value, str):
        return [item.strip() for item in value.split(",") if item.strip()]
    return value


def import_excel_with_serializer(
    serializer_path,
    file_path,
    dry_run=False,
    max_rows=5000,
):
    module_path, class_name = serializer_path.rsplit(".", 1)
    serializer_class = getattr(import_module(module_path), class_name)

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        workbook.close()
        return {
            "success": False,
            "status_code": 400,
            "result": {"file": ["The Excel file does not contain a header row."]},
        }

    headers = [str(header).strip() if header is not None else "" for header in headers]
    serializer_instance = serializer_class(context={"request": None})
    serializer_fields = serializer_instance.fields
    allowed_fields = [
        field_name
        for field_name, field in serializer_fields.items()
        if not field.read_only
        and field.__class__.__name__ not in {"FileField", "ImageField", "HiddenField"}
    ]
    unknown_fields = [
        header for header in headers if header and header not in allowed_fields
    ]
    if unknown_fields:
        workbook.close()
        return {
            "success": False,
            "status_code": 400,
            "result": {
                "headers": ["Unsupported columns: " + ", ".join(unknown_fields)],
                "allowed_columns": allowed_fields,
            },
        }

    raw_rows = []
    for row_number, values in enumerate(rows, start=2):
        if row_number > max_rows + 1:
            workbook.close()
            return {
                "success": False,
                "status_code": 400,
                "result": {
                    "file": [f"Excel import is limited to {max_rows} data rows."]
                },
            }
        if not any(value not in (None, "") for value in values):
            continue
        raw_rows.append((row_number, dict(zip(headers, values, strict=False))))
    workbook.close()

    errors = []
    valid_rows = []
    for row_number, row in raw_rows:
        payload = {
            field_name: _import_value(serializer_fields[field_name], value)
            for field_name, value in row.items()
            if field_name and value not in (None, "")
        }
        serializer = serializer_class(data=payload, context={"request": None})
        if serializer.is_valid():
            valid_rows.append(serializer)
        else:
            errors.append({"row": row_number, "errors": serializer.errors})

    if errors:
        return {
            "success": False,
            "status_code": 400,
            "result": {
                "imported": 0,
                "valid_rows": len(valid_rows),
                "invalid_rows": len(errors),
                "errors": errors,
            },
        }

    if dry_run:
        return {
            "success": True,
            "status_code": 200,
            "result": {
                "dry_run": True,
                "valid_rows": len(valid_rows),
                "invalid_rows": 0,
                "errors": [],
            },
        }

    try:
        with transaction.atomic():
            for serializer in valid_rows:
                serializer.save()
    except Exception:
        return {
            "success": False,
            "status_code": 400,
            "result": {
                "imported": 0,
                "detail": "Excel import was rolled back.",
            },
        }

    return {
        "success": True,
        "status_code": 201,
        "result": {"imported": len(valid_rows), "invalid_rows": 0, "errors": []},
    }
