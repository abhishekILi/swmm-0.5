import json
from io import BytesIO

from django.db import transaction
from django.http import HttpResponse
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import extend_schema, inline_serializer
from openpyxl import Workbook, load_workbook
from rest_framework import serializers, status
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

MAX_EXCEL_ROWS = 5000
EXCLUDED_EXCEL_FIELDS = (
    serializers.FileField,
    serializers.ImageField,
    serializers.HiddenField,
)


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, default=str)
    return value


def _import_value(field, value):
    if value is None or value == "":
        return None
    if isinstance(field, serializers.ManyRelatedField) and isinstance(value, str):
        return [item.strip() for item in value.split(",") if item.strip()]
    return value


def _header_key(value):
    return "".join(char for char in str(value).upper() if char.isalnum())


class ExcelImportExportMixin:
    excel_max_rows = MAX_EXCEL_ROWS

    def get_excel_export_fields(self):
        serializer = self.get_serializer()
        return [
            field_name
            for field_name, field in serializer.fields.items()
            if not isinstance(field, EXCLUDED_EXCEL_FIELDS)
        ]

    def get_excel_import_fields(self):
        serializer = self.get_serializer()
        return [
            field_name
            for field_name, field in serializer.fields.items()
            if not field.read_only and not isinstance(field, EXCLUDED_EXCEL_FIELDS)
        ]

    def get_excel_import_header_map(self):
        return {}

    def normalize_excel_import_headers(self, headers):
        aliases = {
            _header_key(source): target
            for source, target in self.get_excel_import_header_map().items()
        }
        return [
            aliases.get(_header_key(header), header) if header else ""
            for header in headers
        ]

    def build_excel_import_payload(self, row, serializer_fields):
        return {
            field_name: _import_value(serializer_fields[field_name], value)
            for field_name, value in row.items()
            if field_name
            and field_name in serializer_fields
            and value not in (None, "")
        }

    @extend_schema(
        responses={
            (
                200,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ): OpenApiTypes.BINARY
        }
    )
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        fields = self.get_excel_export_fields()
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.get_queryset().model._meta.model_name[:31]
        worksheet.append(fields)
        for record in serializer.data:
            worksheet.append([_excel_value(record.get(field)) for field in fields])

        output = BytesIO()
        workbook.save(output)
        model_name = self.get_queryset().model._meta.model_name
        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{model_name}_export.xlsx"'
        )
        return response

    @extend_schema(
        responses={
            (
                200,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ): OpenApiTypes.BINARY
        }
    )
    @action(detail=False, methods=["get"], url_path="download-template")
    def download_template(self, request):
        fields = self.get_excel_import_fields()
        workbook = Workbook()
        worksheet = workbook.active
        model_name = self.get_queryset().model._meta.model_name
        worksheet.title = model_name[:31]
        worksheet.append(fields)

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{model_name}_template.xlsx"'
        )
        return response

    @extend_schema(
        request=inline_serializer(
            name="OnboardSparesExcelImportRequest",
            fields={
                "file": serializers.FileField(),
                "dry_run": serializers.BooleanField(required=False),
            },
        )
    )
    @action(
        detail=False,
        methods=["post"],
        url_path="import-excel",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        uploaded_file = request.FILES.get("file") or request.FILES.get("excel")
        if uploaded_file is None:
            return Response(
                {"file": ["An Excel file is required."]},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not uploaded_file.name.lower().endswith(".xlsx"):
            return Response(
                {"file": ["Only .xlsx files are supported."]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"file": ["The uploaded Excel file is invalid or corrupted."]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return Response(
                {"file": ["The Excel file does not contain a header row."]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        original_headers = [
            str(header).strip() if header is not None else "" for header in headers
        ]
        headers = self.normalize_excel_import_headers(original_headers)
        allowed_fields = self.get_excel_import_fields()
        mapped_fields = [
            field_name
            for field_name in self.get_excel_import_header_map().values()
            if field_name
        ]
        allowed_fields = [*allowed_fields, *mapped_fields]
        unknown_fields = [
            original_header
            for original_header, header in zip(original_headers, headers, strict=False)
            if header and header not in allowed_fields
        ]
        if unknown_fields:
            return Response(
                {
                    "headers": ["Unsupported columns: " + ", ".join(unknown_fields)],
                    "allowed_columns": allowed_fields,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        raw_rows = []
        data_rows_seen = 0
        empty_tail_rows = 0
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                empty_tail_rows += 1
                if empty_tail_rows >= 50:
                    break
                continue

            empty_tail_rows = 0
            data_rows_seen += 1
            if data_rows_seen > self.excel_max_rows:
                return Response(
                    {
                        "file": [
                            (
                                "Excel import is limited to "
                                f"{self.excel_max_rows} data rows."
                            )
                        ]
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            raw_rows.append((row_number, dict(zip(headers, values, strict=False))))
        workbook.close()

        dry_run = str(request.data.get("dry_run", "")).lower() in {
            "1",
            "true",
            "yes",
        }
        try:
            with transaction.atomic():
                errors = []
                valid_rows = []
                serializer_fields = self.get_serializer().fields
                for row_number, row in raw_rows:
                    payload = self.build_excel_import_payload(row, serializer_fields)
                    serializer = self.get_serializer(data=payload)
                    if serializer.is_valid():
                        valid_rows.append((row_number, serializer))
                    else:
                        errors.append({"row": row_number, "errors": serializer.errors})

                if errors:
                    transaction.set_rollback(True)
                    return Response(
                        {
                            "imported": 0,
                            "valid_rows": len(valid_rows),
                            "invalid_rows": len(errors),
                            "errors": errors,
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if dry_run:
                    transaction.set_rollback(True)
                    return Response(
                        {
                            "dry_run": True,
                            "valid_rows": len(valid_rows),
                            "invalid_rows": 0,
                            "errors": [],
                        }
                    )

                for _, serializer in valid_rows:
                    self.perform_create(serializer)
        except Exception:
            return Response(
                {
                    "imported": 0,
                    "detail": "Excel import was rolled back.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {"imported": len(valid_rows), "invalid_rows": 0, "errors": []},
            status=status.HTTP_201_CREATED,
        )
