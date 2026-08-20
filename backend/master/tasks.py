from io import BytesIO

from celery import shared_task
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ems.models import AddRoutineDetails, RoutineDescription
from swmm.async_jobs import (
    build_export_path,
    build_media_url,
)

HEADERS = [
    "ShipName",
    "MaintopNo",
    "Nomenclature",
    "Frequency",
    "RoutineName",
    "RoutineCategory",
    "Sub Depatment",
    "ClassName",
    "EquipmentName",
    "EquipmentCode",
    "RoutineNo",
    "RoutineDescription",
    "ByWhom",
    "frequency in months",
    "frequency in hours",
    "RHSI",
    "RHSIUpdatedUpto",
    "LastRoutineCompletionDate",
    "LastRoutineCompletedAtRH",
]


def _get_routine_queryset(department_id=None):
    queryset = AddRoutineDetails.objects.select_related(
        "ship",
        "equipment_name__section__department",
        "routine_name",
    )

    if department_id:
        queryset = queryset.filter(equipment_name__section__department_id=department_id)

    return queryset


def _get_routine_description(item):
    return RoutineDescription.objects.filter(
        add_routine_details=item,
    ).first()


def _build_routine_row(item, routine_description):
    equipment = item.equipment_name
    section = equipment.section

    return [
        "INS KOLKATA",
        item.maintop_no or "",
        item.nomenclature or "",
        item.frequency or "",
        item.routine_name.name if item.routine_name else "",
        item.routine_category or "",
        section.name,
        item.class_name or "",
        equipment.name or "",
        equipment.equipment_code or "",
        (routine_description.routine_no if routine_description else ""),
        (routine_description.routine_description if routine_description else ""),
        (routine_description.by_whom if routine_description else ""),
        item.frequency_in_months or "",
        item.frequency_in_hours or "",
        item.rhs_i or "",
        item.rhs_i_updated_upto or "",
        "",
        "",
    ]


def _append_routine_rows(worksheet, queryset):
    seen_pairs = set()

    for item in queryset:
        equipment = item.equipment_name
        section = equipment.section if equipment else None

        if not section or not item.nomenclature:
            continue

        pair_key = (
            section.id,
            item.nomenclature.strip().lower(),
        )

        if pair_key in seen_pairs:
            continue

        seen_pairs.add(pair_key)

        routine_description = _get_routine_description(item)
        worksheet.append(
            _build_routine_row(
                item,
                routine_description,
            )
        )


def _format_routine_worksheet(worksheet):
    for column_number in range(1, len(HEADERS) + 1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = 25

    bold_font = Font(bold=True)
    red_fill = PatternFill(
        start_color="FF9999",
        end_color="FF9999",
        fill_type="solid",
    )

    for column_number, cell in enumerate(
        worksheet[1],
        start=1,
    ):
        cell.font = bold_font

        if column_number > len(HEADERS) - 2:
            cell.fill = red_fill


def _build_routine_workbook(department_id=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Routine Details"

    worksheet.append(HEADERS)

    queryset = _get_routine_queryset(department_id)
    _append_routine_rows(worksheet, queryset)
    _format_routine_worksheet(worksheet)

    return workbook


def _save_routine_workbook(workbook):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    file_path = build_export_path(
        "master-exports",
        "routine_details.xlsx",
    )

    with open(file_path, "wb") as file_obj:
        file_obj.write(output.getvalue())

    return file_path


@shared_task(name="master.tasks.export_routine_excel")
def export_routine_excel_task(user_department_id=None):
    workbook = _build_routine_workbook(user_department_id)
    file_path = _save_routine_workbook(workbook)

    return {
        "success": True,
        "status_code": 200,
        "result": {
            "file_path": file_path,
            "download_url": build_media_url(file_path),
            "filename": "routine_details.xlsx",
        },
    }
