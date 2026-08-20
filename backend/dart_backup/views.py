import csv
import json
import logging
import os
import re
from datetime import date, datetime, timedelta
from io import BytesIO

from celery.result import AsyncResult
from dateutil.relativedelta import relativedelta
from django.apps import apps
from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count, Max, Min, Q
from django.forms.models import model_to_dict
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import now
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
)
from openpyxl import Workbook
from rest_framework import status, viewsets
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from dart.serializers import (
    CompletedRoutineCreateSerializer,
    CreateRADLDefectDataSerializer,
    DefectRectifyRequestSerializer,
    DefectResponseSerializer,
    ExportPendingDefectsACCDBResponseSerializer,
    ExportPendingDefectsACCDBSerializer,
    GenericSuccessResponseSerializer,
    OpdefAnalysisRequestSerializer,
    OpdefInitiateRequestSerializer,
    OpdefInitiateResponseSerializer,
    OpdefPhotoResponseSerializer,
    OpdefPriorParamRequestSerializer,
    OpdefSpareRequestSerializer,
    OpdefSyncPayloadResponseSerializer,
    OpdefTrialRequestSerializer,
)
from dart.tasks import export_all_ra_data_task
from ems.models import EquipmentName, SectionName
from ilms.models import DartMOSpare, Item, ItemExtra, MoMappingTable, Vendor
from master.models import (
    ChMasterShipRemarksBy,
    ChMasterSymptoms,
    Department,
    MDelay,
    MDiagnostic,
    MRepair,
    MRepairAgency,
    MRequiredAssistance,
    MSeverity,
    OpsMaintenancePeriod,
    RefitMaintenancePeriod,
    SubDepartment,
)
from master.utils import get_ship_status
from obs.models import (
    Authority,
    Demand,
    EquipmentClass,
    Issue,
    SpareClass,
    Spares,
    SparesMapping,
    Survey,
)
from sfd.models import Equipment, ShipEquipment, TrialUnit
from swmm.async_jobs import accepted_task_response, dispatch_task, sync_task_response
from swmm.cache_utils import make_cache_key
from swmm.celery import app as celery_app
from wlms.models import DartWedSpare, SpareDataMap, WLMSSpare

from .models import (
    CompleteDefectDart,
    CompletedRoutine,
    DartSpare,
    DartSpareUsed,
    InitiateDart,
    InitiateRADL,
    RADLMaster,
    TCandef,
    TempDartSpare,
)
from .serializers import (
    AddDartMetadataSerializer,
    AllDataOfRASerializer,
    AllPeriodsMaintenanceOverviewSerializer,
    CompleteDartListSerializer,
    CompleteDefectGETDetailSerializer,
    CompleteDefectGETResponseSerializer,
    CompleteDefectPOSTResponseSerializer,
    CompleteDefectSerializer,
    CreateDLDefectDataSerializer,
    CreateDLFunResponseSerializer,
    CreateDLFunSerializer,
    CreateDLRefitItemSerializer,
    CreateRAAjaxSerializer,
    CreateRADLFunResponseSerializer,
    CreateRADLFunSerializer,
    DartDashboardSerializer,
    DartHistoryDefaultDataResponseSerializer,
    DartHistoryDetailResponseSerializer,
    DartHistoryFilteredDataResponseSerializer,
    DartHistoryFilteredDataSerializer,
    DeleteDLRowSerializer,
    ExportPendingDefectsDL2ResponseSerializer,
    ExportPendingDefectsDL2Serializer,
    GeneratedDLIIReportRowSerializer,
    GeneratedDLIIReportSerializer,
    GetDartDetailsResponseSerializer,
    GetDartSparesDataResponseSerializer,
    GetEquipmentDetailsResponseSerializer,
    GetEquipmentObjectsResponseSerializer,
    GetMaintenancePeriodNameResponseSerializer,
    GetNomenclatureDetailsResponseSerializer,
    GetRefitPeriodsResponseSerializer,
    GetSubDeptDefectsResponseSerializer,
    InitiateDartSerializer,
    IssuedSpareSerializer,
    MaintenanceOverviewDetailsSerializer,
    MaintenanceOverviewKPIResponseSerializer,
    MaintenancePeriodItemSerializer,
    MergeDefectsSerializer,
    MoveToDraftSerializer,
    PendingDefectSerializer,
    PendingDefectsResponseSerializer,
    RADLMasterSerializer,
    RefitOccasionCreateSerializer,
    RefitOccasionEditSerializer,
    RefitOperationalOccasionResponseSerializer,
    SaveDLRowsResponseSerializer,
    SaveDLRowsSerializer,
    SaveOEMSpareSerializer,
    SeveritySerializer,
    ShipRemarkSerializer,
    SymptomSerializer,
)
from .services import (
    build_dart_history_row,
    build_dl_ii_export_rows,
    build_pending_defect_export_rows,
    build_simple_pdf,
    get_closed_dart_history_queryset,
    get_export_dart_ids,
    get_generated_dl_ii_rows,
)
from .tasks import complete_candef_sync_task, initiate_candef_sync_task
from .utils import (
    calculate_percentage_change,
    generate_dart_number,
    get_health_status,
    get_readiness_status,
)

DATE_FORMAT = "%d %b %Y"

logger = logging.getLogger(__name__)


def resolve_user_department(request, explicit_department_id=None):
    """Resolve department using the same priority as the legacy DART flow."""
    user = getattr(request, "user", None)
    user_profile = getattr(user, "user_profile", None)
    department = getattr(user_profile, "department", None)

    if department:
        return department

    department = getattr(user, "department", None)
    if department:
        return department

    if explicit_department_id and str(explicit_department_id).lower() != "all":
        return Department.objects.filter(id=explicit_department_id).first()

    return None


@extend_schema(tags=["DART"])
class RADLMasterViewSet(viewsets.ModelViewSet):
    queryset = RADLMaster.objects.select_related("refit_type_f_key").order_by("-id")
    serializer_class = RADLMasterSerializer


@extend_schema(tags=["DART"])
class SymptomViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ChMasterSymptoms.objects.filter(active=1).order_by("id")
    serializer_class = SymptomSerializer


@extend_schema(tags=["DART"])
class SeverityViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = MSeverity.objects.all().order_by("id")
    serializer_class = SeveritySerializer


@extend_schema(tags=["DART"])
class RemarkViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ChMasterShipRemarksBy.objects.filter(active=1).order_by("id")
    serializer_class = ShipRemarkSerializer


@extend_schema(tags=["DART"])
class CompleteDartListAPIView(APIView):
    """
    GET  → List all closed DARTs (is_closed=True, defective_discriptions is not null)
    """

    serializer_class = CompleteDartListSerializer

    def get(self, request, *args, **kwargs):
        defect_obj = (
            InitiateDart.objects.filter(is_closed=True)
            .exclude(defective_discriptions=None)
            .select_related(
                "symptom_code",
                "severity_code",
                "remark_code",
                "department_id",
                "equipment_ship",
                "equipment_ems",
            )
            .prefetch_related(
                "complete_defect_dart_set",
            )
            .order_by("-id")
        )

        serializer = CompleteDartListSerializer(defect_obj, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class PendingDefectAPIView(APIView):
    """
    Returns the pending (open) defects list for a department, along with filter options and DL III defects.
    """

    serializer_class = PendingDefectSerializer

    @extend_schema(
        parameters=[
            OpenApiParameter(
                name="department_id",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="ID of the department to filter pending defects, or 'All'",
            ),
            OpenApiParameter(
                name="sub_department",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="Sub-department name to filter pending defects, or 'All'",
            ),
            OpenApiParameter(
                name="maintenance_period",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="Maintenance period to filter pending defects, or 'All'",
            ),
            OpenApiParameter(
                name="dart_occasion",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="DART occasion to filter pending defects, or 'All'",
            ),
            OpenApiParameter(
                name="equipment_nomenclature",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="Equipment nomenclature to filter pending defects, or 'All'",
            ),
            OpenApiParameter(
                name="date_range",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description=(
                    "Defect date range preset: Today, Yesterday, Last 7 Days, "
                    "Last 30 Days, This Month, Last Month, Custom, or All Data"
                ),
            ),
            OpenApiParameter(
                name="defect_date_from",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="Custom defect date range start in YYYY-MM-DD format",
            ),
            OpenApiParameter(
                name="defect_date_to",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="Custom defect date range end in YYYY-MM-DD format",
            ),
        ],
        responses={200: PendingDefectsResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        user = getattr(request, "user", None)
        user_profile = getattr(user, "user_profile", None)
        user_dept = resolve_user_department(request)
        designation = (
            getattr(user_profile, "designation", None)
            or getattr(user, "designation", "")
            or ""
        ).upper()
        privileged_roles = ["EO", "CO", "EO ENGG HOD"]
        is_privileged = any(role in designation for role in privileged_roles)

        selected_sub_department = request.query_params.get(
            "sub_department"
        ) or request.query_params.get("sub_dept")
        selected_maintenance_period = request.query_params.get(
            "maintenance_period"
        ) or request.query_params.get("maintenancePeriod")
        selected_dart_occasion = request.query_params.get(
            "dart_occasion"
        ) or request.query_params.get("dartOccasion")
        selected_equipment_nomenclature = request.query_params.get(
            "equipment_nomenclature"
        ) or request.query_params.get("equipmentNomenclature")
        selected_date_range = (
            request.query_params.get("date_range")
            or request.query_params.get("defect_date_range")
            or request.query_params.get("defectDateRange")
        )
        defect_date_from = (
            request.query_params.get("defect_date_from")
            or request.query_params.get("defectDateFrom")
            or request.query_params.get("from_date")
            or request.query_params.get("start_date")
        )
        defect_date_to = (
            request.query_params.get("defect_date_to")
            or request.query_params.get("defectDateTo")
            or request.query_params.get("to_date")
            or request.query_params.get("end_date")
        )

        def has_filter_value(value):
            return value and str(value).lower() != "all"

        def parse_filter_date(value):
            if not value:
                return None
            try:
                return datetime.strptime(str(value), "%Y-%m-%d").date()
            except ValueError:
                return None

        def get_date_bounds(date_range, date_from, date_to):
            today = timezone.now().date()
            normalized_range = str(date_range or "").strip().lower().replace("_", " ")
            start_date = parse_filter_date(date_from)
            end_date = parse_filter_date(date_to)

            if normalized_range in {"", "all", "all data"} and not (
                start_date or end_date
            ):
                return None, None
            if normalized_range == "today":
                return today, today
            if normalized_range == "yesterday":
                yesterday = today - timedelta(days=1)
                return yesterday, yesterday
            if normalized_range in {"last 7 days", "last 7 day"}:
                return today - timedelta(days=7), today
            if normalized_range in {"last 30 days", "last 30 day"}:
                return today - timedelta(days=30), today
            if normalized_range == "this month":
                return today.replace(day=1), today
            if normalized_range == "last month":
                first_this_month = today.replace(day=1)
                last_prev_month = first_this_month - timedelta(days=1)
                return last_prev_month.replace(day=1), last_prev_month
            return start_date, end_date

        selected_dept_id = request.query_params.get(
            "department_id"
        ) or request.query_params.get("department_filter")

        filter_dept = None
        if is_privileged:
            if selected_dept_id and str(selected_dept_id).lower() != "all":
                filter_dept = selected_dept_id
            elif not selected_dept_id:
                if (
                    user_dept
                    and InitiateDart.objects.open()
                    .filter(department_id=user_dept.id)
                    .exists()
                ):
                    filter_dept = user_dept.id
                else:
                    filter_dept = None
        elif user_dept:
            filter_dept = user_dept.id
        # Fetch all departments for selection if privileged
        all_departments_data = []
        if is_privileged:
            all_depts = Department.objects.all().order_by("name")
            all_departments_data = [
                {"id": dept.id, "name": dept.name} for dept in all_depts
            ]

        # Fetch open defects for open_defects (filtered by department if applicable)
        open_defect_obj = InitiateDart.objects.open().select_related(
            "equipment_ship",
            "equipment_ship__sub_department_f_key",
            "equipment_ship__equipment",
            "remark_code",
        )
        if filter_dept:
            open_defect_obj = open_defect_obj.filter(department_id=filter_dept)

        if has_filter_value(selected_sub_department):
            open_defect_obj = open_defect_obj.filter(
                equipment_ship__sub_department_f_key__name__iexact=selected_sub_department
            )
        if has_filter_value(selected_maintenance_period):
            open_defect_obj = open_defect_obj.filter(
                maintenance_period__iexact=selected_maintenance_period
            )
        if has_filter_value(selected_dart_occasion):
            open_defect_obj = open_defect_obj.filter(
                dart_occasion__iexact=selected_dart_occasion
            )
        if has_filter_value(selected_equipment_nomenclature):
            open_defect_obj = open_defect_obj.filter(
                equipment_ship__nomenclature__iexact=selected_equipment_nomenclature
            )
        start_date, end_date = get_date_bounds(
            selected_date_range, defect_date_from, defect_date_to
        )
        if start_date and end_date:
            open_defect_obj = open_defect_obj.filter(
                dart_date__range=(start_date, end_date)
            )
        elif start_date:
            open_defect_obj = open_defect_obj.filter(dart_date__gte=start_date)
        elif end_date:
            open_defect_obj = open_defect_obj.filter(dart_date__lte=end_date)

        metadata_queryset = InitiateDart.objects.open().select_related(
            "equipment_ship",
            "equipment_ship__sub_department_f_key",
            "equipment_ship__equipment",
        )
        dept_metadata_qs = (
            metadata_queryset.filter(department_id=filter_dept)
            if filter_dept
            else metadata_queryset
        )
        if not dept_metadata_qs.exists():
            dept_metadata_qs = metadata_queryset

        def get_distinct_values(queryset, field_name, normalizer=None):
            unique_values = {}
            for value in queryset.values_list(field_name, flat=True):
                if not value:
                    continue
                display_value = normalizer(value) if normalizer else value
                unique_values.setdefault(display_value.lower(), display_value)
            return sorted(unique_values.values())

        def normalize_maintenance_period(value):
            return str(value).strip().title()

        sub_departments = get_distinct_values(
            dept_metadata_qs,
            "equipment_ship__sub_department_f_key__name",
        )
        maintenance_periods = get_distinct_values(
            dept_metadata_qs,
            "maintenance_period",
            normalize_maintenance_period,
        )

        occasion_queryset = dept_metadata_qs
        if has_filter_value(selected_maintenance_period):
            occasion_queryset = occasion_queryset.filter(
                maintenance_period__iexact=selected_maintenance_period
            )
        dart_occasions = get_distinct_values(occasion_queryset, "dart_occasion")

        # Equipment nomenclature and names stay data-driven from the available metadata records.
        equipment_nomenclatures = get_distinct_values(
            dept_metadata_qs,
            "equipment_ship__nomenclature",
        )
        equipment_names = get_distinct_values(
            dept_metadata_qs,
            "equipment_ship__equipment__equipment_class",
        )

        min_date_val = dept_metadata_qs.aggregate(Min("dart_date"))["dart_date__min"]
        max_date_val = dept_metadata_qs.aggregate(Max("dart_date"))["dart_date__max"]
        min_date = min_date_val.strftime("%Y-%m-%d") if min_date_val else None
        max_date = max_date_val.strftime("%Y-%m-%d") if max_date_val else None

        open_defect_obj = open_defect_obj.order_by("-id")

        # Fetch DL III open defects for Tab 2 (filtered by department if applicable)
        dl_3_defect_obj = (
            InitiateDart.objects.open()
            .filter(
                maintenance_period__iexact="REFIT",
                dart_occasion__icontains="DL III",
            )
            .select_related(
                "equipment_ship",
                "equipment_ship__sub_department_f_key",
                "equipment_ship__equipment",
                "remark_code",
            )
        )
        if filter_dept:
            dl_3_defect_obj = dl_3_defect_obj.filter(department_id=filter_dept)

        if has_filter_value(selected_sub_department):
            dl_3_defect_obj = dl_3_defect_obj.filter(
                equipment_ship__sub_department_f_key__name__iexact=selected_sub_department
            )
        if has_filter_value(selected_equipment_nomenclature):
            dl_3_defect_obj = dl_3_defect_obj.filter(
                equipment_ship__nomenclature__iexact=selected_equipment_nomenclature
            )
        if start_date and end_date:
            dl_3_defect_obj = dl_3_defect_obj.filter(
                dart_date__range=(start_date, end_date)
            )
        elif start_date:
            dl_3_defect_obj = dl_3_defect_obj.filter(dart_date__gte=start_date)
        elif end_date:
            dl_3_defect_obj = dl_3_defect_obj.filter(dart_date__lte=end_date)

        dl_3_defect_obj = dl_3_defect_obj.order_by("-id")

        # Serialize datasets
        open_defects_data = PendingDefectSerializer(open_defect_obj, many=True).data
        dl_3_defects_data = PendingDefectSerializer(dl_3_defect_obj, many=True).data

        response_data = {
            "is_privileged": is_privileged,
            "all_departments": all_departments_data,
            "selected_dept_id": filter_dept,
            "filter_options": {
                "sub_departments": sub_departments,
                "maintenance_periods": maintenance_periods,
                "dart_occasions": dart_occasions,
                "equipment_nomenclatures": equipment_nomenclatures,
                "equipment_names": equipment_names,
                "min_date": min_date,
                "max_date": max_date,
            },
            "open_defects": open_defects_data,
            "dl_3_defects": dl_3_defects_data,
        }
        return Response(response_data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class DartDashboardAPIView(APIView):
    """
    Returns the complete DART Dashboard data as JSON.
    Mirrors the legacy dart_dashboard template view.
    """

    serializer_class = DartDashboardSerializer

    @extend_schema(
        parameters=[
            OpenApiParameter(
                name="department_id",
                type=int,
                location=OpenApiParameter.QUERY,
                required=True,
                description="ID of the department to fetch dashboard for",
            ),
        ],
        responses={200: DartDashboardSerializer},
    )
    def get(self, request, *args, **kwargs):
        # ── Get department from query param ───────────────
        department_id = request.query_params.get("department_id")
        if not department_id:
            return Response(
                {"error": "department_id query parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        cache_key = make_cache_key(
            "dart:dashboard",
            request,
            extra={"department_id": department_id},
        )
        cached_payload = cache.get(cache_key)
        if cached_payload is not None:
            serializer = DartDashboardSerializer(data=cached_payload)
            serializer.is_valid(raise_exception=True)
            return Response(serializer.validated_data, status=status.HTTP_200_OK)

        try:
            department = Department.objects.get(pk=department_id)
        except Department.DoesNotExist:
            return Response(
                {"error": f"Department with id {department_id} not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        today = timezone.now().date()

        # ── Sub-departments for department ────────────────
        sub_depts = SubDepartment.objects.filter(department_name=department)

        # ── 1. Ship Status ────────────────────────────────
        ship_status_result = get_ship_status()
        ship_status = ship_status_result[0] if ship_status_result else None

        # ── 2. Open DARTs — Operational ───────────────────
        open_darts_ops_count = (
            InitiateDart.objects.open().for_department(department).operational().count()
        )

        # ── 3. Open DARTs — Refit ─────────────────────────
        open_darts_refit_count = (
            InitiateDart.objects.open().for_department(department).refit().count()
        )

        # ── 4. DARTs Due for Closing ──────────────────────
        due_for_closing_count = (
            InitiateDart.objects.open()
            .for_department(department)
            .filter(rectification_date__lt=today)
            .count()
        )

        # ── 5. Sub-Department Status & Equipment Data ─────
        sub_dept_status_data = []
        sub_dept_equipment_data = []

        for sd in sub_depts:
            # Status based on major defects (Non-Ops severity)
            major_defects_count = (
                InitiateDart.objects.open()
                .non_ops()
                .filter(equipment_ship__sub_department_f_key=sd)
                .count()
            )
            sub_dept_status_data.append(
                {
                    "name": sd.name,
                    "status": "Non-Ops" if major_defects_count > 0 else "Ops",
                    "count": major_defects_count,
                }
            )

            # Equipment counts for gauge charts
            total_eq = EquipmentName.objects.filter(sub_department=sd).count()
            non_ops_eq = (
                InitiateDart.objects.open()
                .non_ops()
                .filter(equipment_ship__sub_department_f_key=sd)
                .values("equipment_ship")
                .distinct()
                .count()
            )
            ops_eq = total_eq - non_ops_eq

            sub_dept_equipment_data.append(
                {
                    "sub_dept": sd.name,
                    "operational": ops_eq,
                    "non_operational": non_ops_eq,
                    "total": total_eq,
                }
            )

        # ── 6. Last 6 months Chart Data ──────────────────
        open_chart_data = []
        closed_chart_data = []

        for i in range(5, -1, -1):
            month_date = today - relativedelta(months=i)
            m_start = month_date.replace(day=1)
            m_end = (month_date + relativedelta(months=1)).replace(day=1) - timedelta(
                days=1
            )

            month_label = month_date.strftime("%b %Y")

            open_entry = {"month": month_label}
            closed_entry = {"month": month_label}
            total_open_month = 0
            total_closed_month = 0

            for sd in sub_depts:
                # Opened count
                op_cnt = (
                    InitiateDart.objects.for_department(department)
                    .filter(
                        equipment_ship__sub_department_f_key=sd,
                        dart_date__range=(m_start, m_end),
                    )
                    .count()
                )
                open_entry[sd.name] = op_cnt
                total_open_month += op_cnt

                # Closed count
                cl_cnt = CompleteDefectDart.objects.filter(
                    dart_details__department_id=department,
                    dart_details__equipment_ship__sub_department_f_key=sd,
                    rectified_date__range=(m_start, m_end),
                ).count()
                closed_entry[sd.name] = cl_cnt
                total_closed_month += cl_cnt

            open_entry["total"] = total_open_month
            closed_entry["total"] = total_closed_month
            open_chart_data.append(open_entry)
            closed_chart_data.append(closed_entry)

        # ── 7. Maintenance Periods (last 1 year) ─────────
        one_year_ago = today - timedelta(days=365)
        maintenance_periods = []

        # Operational maintenance periods
        ops_periods = OpsMaintenancePeriod.objects.filter(
            start_date__gte=one_year_ago
        ).order_by("-start_date")
        for op in ops_periods:
            is_current = (
                op.start_date and op.end_date and op.start_date <= today <= op.end_date
            )
            maintenance_periods.append(
                {
                    "maintenance_period": op.maintenance_period or "Operational",
                    "occasion": op.occasion or op.name or "-",
                    "start_date": (
                        op.start_date.strftime(DATE_FORMAT) if op.start_date else "-"
                    ),
                    "end_date": (
                        op.end_date.strftime(DATE_FORMAT) if op.end_date else "-"
                    ),
                    "is_current": is_current,
                }
            )

        # Refit maintenance periods
        refit_periods = RefitMaintenancePeriod.objects.filter(
            Q(actual_start_date__gte=one_year_ago)
            | Q(plan_start_date__gte=one_year_ago)
        ).order_by("-actual_start_date", "-plan_start_date")
        for rp in refit_periods:
            s_date = rp.actual_start_date or rp.plan_start_date
            e_date = rp.actual_end_date or rp.plan_end_date
            is_current = s_date and e_date and s_date <= today <= e_date
            maintenance_periods.append(
                {
                    "maintenance_period": rp.maintenance_period or "Refit",
                    "occasion": rp.occasion or rp.name or "-",
                    "start_date": s_date.strftime(DATE_FORMAT) if s_date else "-",
                    "end_date": e_date.strftime(DATE_FORMAT) if e_date else "-",
                    "is_current": is_current,
                }
            )

        # Sort: current first, then by start_date
        maintenance_periods.sort(
            key=lambda x: (not x["is_current"], x["start_date"]), reverse=False
        )

        # ── Build & serialize response ────────────────────
        dashboard_data = {
            "ship_status": ship_status,
            "open_darts_ops_count": open_darts_ops_count,
            "open_darts_refit_count": open_darts_refit_count,
            "due_for_closing_count": due_for_closing_count,
            "sub_dept_status_data": sub_dept_status_data,
            "sub_dept_equipment_data": sub_dept_equipment_data,
            "open_chart_data": open_chart_data,
            "closed_chart_data": closed_chart_data,
            "sub_depts": [sd.name for sd in sub_depts],
            "maintenance_periods": maintenance_periods,
        }

        cache.set(cache_key, dashboard_data, timeout=300)
        serializer = DartDashboardSerializer(data=dashboard_data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.validated_data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class AddDartMetadataAPIView(APIView):
    """
    GET  → Returns all the form options, dropdown choices, and next generated serial number
           needed to render the "Add DART" form.
    """

    serializer_class = AddDartMetadataSerializer

    @extend_schema(
        responses={200: AddDartMetadataSerializer},
    )
    def get(self, request, *args, **kwargs):
        department = getattr(request.user, "department", None)
        if not department and hasattr(request.user, "CustomUser_profile"):
            department = getattr(request.user.CustomUser_profile, "department", None)
        if not department:
            department = Department.objects.first()

        # ── 1. ec_code ──
        ec_code = Equipment.objects.filter().distinct()

        # ── 2. Dynamic Models (Spares, SpareClass, EquipmentClass) ──
        spare_data = []
        oem_pil_spares_data = []
        try:
            spare_classes_qs = (
                SpareClass.objects.filter(department_id=department)
                if department
                else SpareClass.objects.all()
            )
            spare_classes = [e.id for e in spare_classes_qs]
            equipment_classes = [
                e.id
                for e in EquipmentClass.objects.filter(spare_class_id__in=spare_classes)
            ]
            spare_obj = Spares.objects.filter(
                equipment_class_id__in=equipment_classes
            ).exclude(
                Q(authority__name="MO ITEM")
                | Q(authority__name="B & D")
                | Q(authority__name="MO ALLOWANCE")
            )
            spare_data = [
                {
                    "id": s.id,
                    "name": s.description or "",
                    "part_number": s.pattern_number or "",
                }
                for s in spare_obj
            ]

            oem_pil_spares = Spares.objects.filter(
                authority__name="PIL", equipment_class__name="SPARES PIL"
            )
            oem_pil_spares_data = [
                {"id": s.id, "name": s.description or ""} for s in oem_pil_spares
            ]
        except Exception as e:
            print("Error loading spares metadata:", str(e))

        # ── 3. ilms_objs ──
        ilms_data = [
            {
                "item_code": i.item_code,
                "item_desc": i.item_desc or "",
                "pk": i.pk,
            }
            for i in ItemExtra.objects.all()
        ]

        # ── 4. denominations ──
        denominations_data = []
        try:
            Denomination = apps.get_model("master", "Denomination")
            denominations_data = [
                {"id": d.id, "name": getattr(d, "name", str(d))}
                for d in Denomination.objects.filter(department_id=department)
            ]
        except Exception:
            pass

        # ── 5. Serial Generation ──
        dept_name = department.name or "X"
        new_serial = generate_dart_number(dept_name, "Defect")

        # Determine last_serial (highest existing number) for response
        prefix = new_serial.rsplit("-", 1)[0] + "-"
        dart_list = InitiateDart.objects.filter(dart_number__startswith=prefix)
        last_entry = None
        if dart_list.exists():
            max_val = -1
            for entry in dart_list:
                dn = entry.dart_number
                if dn:
                    match = re.search(r"\d+$", dn)
                    if match:
                        val = int(match.group())
                        if val > max_val:
                            max_val = val
                            last_entry = entry
            if last_entry is None:
                last_entry = dart_list.order_by("-id").first()

        last_serial = last_entry.dart_number if last_entry else None

        # ── 6. Master Lists ──
        s_ids = (
            ChMasterSymptoms.objects.values("symptom_code")
            .annotate(min_id=Min("id"))
            .values_list("min_id", flat=True)
        )
        symptom_list = ChMasterSymptoms.objects.filter(id__in=s_ids).order_by(
            "symptom_code"
        )

        severity_list = MSeverity.objects.filter(active=1)
        remark_list = ChMasterShipRemarksBy.objects.all().distinct()

        a_ids = (
            MRequiredAssistance.objects.values("required_assistance_for")
            .annotate(min_id=Min("required_assistance_id"))
            .values_list("min_id", flat=True)
        )
        assistance_list = MRequiredAssistance.objects.filter(
            required_assistance_id__in=a_ids
        )

        # ── 7. Trial Agencies ──

        trial_agency = [
            {"id": str(tu.id), "name": tu.name}
            for tu in TrialUnit.objects.filter(status=1)
        ]

        db_occ = RefitMaintenancePeriod.objects.exclude(maintenance_period__isnull=True)
        m_types = list(db_occ.values_list("maintenance_period", flat=True).distinct())
        m_occasions = {
            t: list(
                db_occ.filter(maintenance_period=t)
                .values_list("occasion", flat=True)
                .exclude(occasion__isnull=True)
                .distinct()
            )
            for t in m_types
        }

        # ── Serialize with DRF Serializers ──
        context_data = {
            "ec_code": ec_code,
            "spare_obj": spare_data,
            "last_serial": last_serial,
            "new_serial": new_serial,
            "ilms_objs": ilms_data,
            "refit_period": RefitMaintenancePeriod.objects.filter(
                maintenance_period="REFIT"
            ),
            "ops_period": RefitMaintenancePeriod.objects.filter(
                maintenance_period="OPERATIONAL"
            ),
            "symptom_list": symptom_list,
            "severity_list": severity_list,
            "remark_list": remark_list,
            "assistance_list": assistance_list,
            "oem_pil_spares": oem_pil_spares_data,
            "denominations": denominations_data,
            "trial_agency": trial_agency,
            "maitainance_period_types": m_types,
            "maitianance_period_occasions": m_occasions,
        }

        serializer = AddDartMetadataSerializer(context_data)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class InitiateDartAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @extend_schema(
        request=InitiateDartSerializer,
        responses={
            201: OpenApiResponse(
                description="Dart initiated successfully",
                examples=[
                    OpenApiExample(
                        "Success Response",
                        value={
                            "message": "Dart initiated successfully",
                            "dart_id": 1,
                            "dart_number": "DART-ENGINEERING-0001",
                        },
                    )
                ],
            ),
            400: OpenApiResponse(description="Bad Request"),
        },
        description=("Initiate a new DART."),
    )
    def post(self, request, *args, **kwargs):
        serializer = InitiateDartSerializer(data=request.data)
        if not serializer.is_valid():
            errors = serializer.errors
            msg = "Validation failed"
            if errors:
                first_field = next(iter(errors))
                first_err = errors[first_field]
                if isinstance(first_err, list):
                    msg = f"{first_field}: {first_err[0]}"
                else:
                    msg = f"{first_field}: {first_err}"
            return Response(
                {"status": "error", "message": msg, "errors": errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        val_data = serializer.validated_data

        # Date Validation: closing/resolution date cannot be earlier than creation/defect date
        dart_type = val_data.get("dart_type")
        defect_date = None
        resolution_date = None

        if dart_type == "Guarantee Defect":
            defect_date = val_data.get("g_defectDate")
            resolution_date = val_data.get("g_completionDate")
        else:
            defect_date = val_data.get("defectDate")
            resolution_date = val_data.get("scheduledDate")

        if defect_date and resolution_date and resolution_date < defect_date:
            return Response(
                {
                    "status": "error",
                    "message": "DART Closing Date cannot be earlier than the DART Creation Date",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            with transaction.atomic():
                # 1. BASIC EQUIPMENT
                dart_type = val_data.get("dart_type")

                nom_id = None
                if dart_type == "Guarantee Defect":
                    nom_id = val_data.get("g_nomenclature")
                elif dart_type in ["ASR", "ABER"]:
                    nom_id = val_data.get("a_nomenclature")
                else:
                    nom_id = val_data.get("nomenclature")

                equipment_obj = None
                if nom_id and str(nom_id).isdigit():
                    equipment_obj = ShipEquipment.objects.filter(id=nom_id).first()

                photograph = val_data.get("attachPhotograph")

                # 2. COMMON DATA (dropdown IDs)
                symptom_id = val_data.get("symptoms")
                severity_id = val_data.get("severity")
                remark_id = val_data.get("ssRemarks")
                require_assistance_id = val_data.get("requiredAssistance")
                maintenance_period = val_data.get("maintenance_period")
                dart_occasion = val_data.get("dart_occasion")
                universal_id_t_dart = val_data.get("universal_id_t_dart")

                # 3. DEPARTMENT RESOLUTION
                department = resolve_user_department(
                    request,
                    explicit_department_id=val_data.get("department_id"),
                )
                if not department:
                    department = Department.objects.first()
                # 4. DART NUMBER GENERATION
                dept_name = department.name if department else "X"
                new_serial = generate_dart_number(dept_name, "Defect")

                # EquipmentName (EMS) lookup
                equipment_name_obj = None
                if equipment_obj:
                    equipment_name_obj = EquipmentName.objects.filter(
                        sfd_equipment=equipment_obj
                    ).first()

                # Guarantee-specific fields
                g_repairs = val_data.get("g_repairs")
                g_repair_date = val_data.get("g_repairDate")
                g_place = val_data.get("g_place")
                if g_repairs == "0":
                    g_repair_date = None
                    g_place = None

                dart_sr_no = new_serial.split("-")[-1] if new_serial else None

                defect_desc = val_data.get("defect_description") or val_data.get("dig")
                g_defect_desc = (
                    val_data.get("g_defect_description")
                    or val_data.get("g_dig")
                    or defect_desc
                )
                trial_agency_val = val_data.get("trial_agency") or val_data.get(
                    "UniversalIDMOSTList"
                )

                # 5. CREATE DART (conditional on dart_type)
                if dart_type == "Guarantee Defect":
                    dart = InitiateDart.objects.create(
                        department_id=department,
                        equipment_ship=equipment_obj,
                        equipment_ems=equipment_name_obj,
                        dart_number=new_serial,
                        dart_sr_number=dart_sr_no,
                        universal_id_t_dart=universal_id_t_dart,
                        dart_date=val_data.get("g_defectDate"),
                        defective_discriptions=g_defect_desc,
                        maintenance_period=maintenance_period,
                        dart_occasion=dart_occasion,
                        is_ra_initiate=False,
                        is_guarantee_defect=True,
                        guarantee_cause=val_data.get("g_cause"),
                        guarantee_op_availability=(
                            val_data.get("opAvailability") == "YES"
                        ),
                        guarantee_hot_work=(val_data.get("hotWork") == "YES"),
                        guarantee_repairs=g_repairs or None,
                        guarantee_completion_date=val_data.get("g_completionDate"),
                        guarantee_repair_date=g_repair_date,
                        guarantee_place=g_place,
                    )

                elif dart_type in ["ABER", "ASR"]:
                    spares_req = (
                        str(
                            val_data.get("spares_required")
                            or val_data.get("spares_required2")
                            or "NO"
                        ).upper()
                        == "YES"
                    )
                    dart = InitiateDart.objects.create(
                        department_id=department,
                        equipment_ship=equipment_obj,
                        equipment_ems=equipment_name_obj,
                        dart_number=new_serial,
                        dart_sr_number=dart_sr_no,
                        universal_id_t_dart=universal_id_t_dart,
                        maintenance_period=maintenance_period or "REFIT",
                        dart_occasion=dart_occasion or "DL III (As and As / ABER)",
                        is_ra_initiate=False,
                        is_guarantee_defect=False,
                        sapres_required=spares_req,
                        defective_discriptions=(
                            val_data.get("description")
                            or val_data.get("asr_desc")
                            or defect_desc
                        ),
                    )

                else:
                    # Normal Defect (default)
                    trial_req = str(val_data.get("trial") or "NO").upper() == "YES"
                    spares_req = (
                        str(val_data.get("spares_required") or "NO").upper() == "YES"
                    )

                    dart = InitiateDart.objects.create(
                        department_id=department,
                        equipment_ship=equipment_obj,
                        equipment_ems=equipment_name_obj,
                        dart_number=new_serial,
                        dart_sr_number=dart_sr_no,
                        universal_id_t_dart=universal_id_t_dart,
                        dart_date=val_data.get("defectDate"),
                        rectification_date=val_data.get("scheduledDate"),
                        defective_discriptions=defect_desc,
                        defective_component=val_data.get("defectiveComponent"),
                        trial_required=trial_req,
                        sapres_required=spares_req,
                        photograph=photograph,
                        maintenance_period=maintenance_period,
                        dart_occasion=dart_occasion,
                        is_ra_initiate=False,
                        is_guarantee_defect=False,
                        symptom_code_id=symptom_id if symptom_id else None,
                        severity_code_id=severity_id if severity_id else None,
                        remark_code_id=remark_id if remark_id else None,
                        require_assistance_for_code_id=(
                            require_assistance_id if require_assistance_id else None
                        ),
                        universal_id_trial_required=(
                            trial_agency_val if trial_req else None
                        ),
                    )

                # Generate universal_id_t_dart if missing (legacy behavior)
                if not dart.universal_id_t_dart:
                    try:
                        ship_obj = get_this_ship()
                        ship_no = (
                            ship_obj.code
                            if (ship_obj and ship_obj.code)
                            else (
                                ship_obj.universal_id_m_ship if ship_obj else "UNKNOWN"
                            )
                        )
                        now_dt = timezone.now()
                        year_str = now_dt.strftime("%Y")
                        month_str = now_dt.strftime("%m")
                        dart.universal_id_t_dart = (
                            f"SWMM/{ship_no}/{year_str}/{month_str}/{dart.id}"
                        )
                        dart.save(update_fields=["universal_id_t_dart"])
                    except Exception as ex_uid:
                        print("Error generating universal_id_t_dart:", str(ex_uid))

                # 6. REFIT / OPS PERIOD LINK
                ops_period_id = val_data.get("ops_period_id")
                if ops_period_id:
                    refit_obj = RefitMaintenancePeriod.objects.filter(
                        id=ops_period_id
                    ).first()
                    if refit_obj:
                        dart.refit_maintenance_period_f_key = refit_obj
                        dart.save()

                # 7. SPARES (split by inventory type — exactly like legacy)
                obs_spares = json.loads(val_data.get("obs_spares") or "[]")
                pil_spares = json.loads(val_data.get("pil_spares") or "[]")
                ilms_spares = json.loads(val_data.get("ilms_spares") or "[]")

                # OBS and PIL insertion
                for spare in obs_spares:
                    inv_type = spare.get("inventory_type", "OBS")
                    if inv_type in ["OBS", "PIL"]:
                        pattern = spare.get("pattern", "")
                        description = spare.get("description", "")
                        qty = int(spare.get("qty", 1) or spare.get("quantity", 1))
                        spare_obj = Spares.objects.filter(
                            pattern_number=pattern
                        ).first()
                        DartSpare.objects.update_or_create(
                            dart=dart,
                            spare_id=spare_obj.id if spare_obj else None,
                            equipment_id=equipment_obj,
                            pattern=pattern,
                            description=description,
                            quantity=qty,
                            inventory_type=inv_type,
                        )

                for spare in pil_spares:
                    pattern = spare.get("pattern", "")
                    description = spare.get("description", "")
                    qty = int(spare.get("qty", 1) or spare.get("quantity", 1))
                    spare_obj = Spares.objects.filter(pattern_number=pattern).first()
                    DartSpare.objects.update_or_create(
                        dart=dart,
                        spare_id=spare_obj.id if spare_obj else None,
                        equipment_id=equipment_obj,
                        pattern=pattern,
                        description=description,
                        quantity=qty,
                        inventory_type="PIL",
                    )

                # ILMS (WED and MO)
                for spare in ilms_spares:
                    inv_type = spare.get("inventory_type", "")
                    pattern = spare.get("pattern", "")
                    desc = spare.get("description", "")
                    qty = int(spare.get("qty", 1) or spare.get("quantity", 1))

                    if inv_type == "WED":
                        wed_spare_obj = WLMSSpare.objects.filter(
                            item_code=pattern
                        ).first()
                        DartWedSpare.objects.update_or_create(
                            dart_id=dart,
                            equipment_id=equipment_obj,
                            wed_spare=wed_spare_obj,
                            pattern=pattern,
                            description=desc,
                            quantity=qty,
                        )
                    elif inv_type == "MO":
                        mo_spare_obj = Item.objects.filter(item_code=pattern).first()
                        DartMOSpare.objects.update_or_create(
                            dart_id=dart,
                            equipment_id=equipment_obj,
                            mo_spare=mo_spare_obj,
                            pattern=pattern,
                            description=desc,
                            quantity=qty,
                        )

                # 8. SPARES MAPPING CHECK & CREATE
                obs_pil_patterns = [s.get("pattern") for s in obs_spares] + [
                    s.get("pattern") for s in pil_spares
                ]
                if (
                    obs_pil_patterns
                    and equipment_obj
                    and hasattr(equipment_obj, "sub_department_f_key")
                ):
                    section_name_obj = getattr(
                        equipment_obj.sub_department_f_key, "name", None
                    )
                    if section_name_obj:
                        section_obj = SectionName.objects.filter(
                            name=section_name_obj
                        ).first()
                        if section_obj:
                            spare_objs = Spares.objects.filter(
                                pattern_number__in=obs_pil_patterns
                            )
                            eq_classes = EquipmentClass.objects.filter(
                                id__in=spare_objs.values_list(
                                    "equipment_class_id", flat=True
                                )
                            ).distinct()

                            for eq_class in eq_classes:
                                SparesMapping.objects.get_or_create(
                                    equipment_class=eq_class,
                                    equipment=equipment_obj,
                                    section_name_id=section_obj.id,
                                )

                # WED Mapping
                wed_patterns = [
                    s.get("pattern")
                    for s in ilms_spares
                    if s.get("inventory_type") == "WED"
                ]
                if wed_patterns and equipment_obj:
                    for pattern in wed_patterns:
                        wed_spare_obj = WLMSSpare.objects.filter(
                            item_code=pattern
                        ).first()
                        if wed_spare_obj and wed_spare_obj.eqpt:
                            SpareDataMap.objects.get_or_create(
                                equipment=equipment_obj,
                                wed_equipment=wed_spare_obj.eqpt,
                                wed_spares=wed_spare_obj,
                            )

                # MO Mapping
                mo_patterns = [
                    s.get("pattern")
                    for s in ilms_spares
                    if s.get("inventory_type") == "MO"
                ]
                if mo_patterns and equipment_obj:
                    for pattern in mo_patterns:
                        item_obj = Item.objects.filter(item_code=pattern).first()
                        if item_obj:
                            vendor = None
                            if pattern and "-" in pattern:
                                vendor_prefix = pattern.split("-", 1)[0]
                                vendor = Vendor.objects.filter(
                                    vendor_code=vendor_prefix[1:]
                                ).first()

                            if not MoMappingTable.objects.filter(
                                ilms_spare_id=item_obj,
                                equipment=equipment_obj,
                            ).exists():
                                MoMappingTable.objects.create(
                                    ilms_spare_id=item_obj,
                                    equipment=equipment_obj,
                                    vendor_id=vendor,
                                )

                # 9. UPDATE DART NUMBER ON OBS ISSUE, DEMAND & SURVEY ENTRIES (Legacy Step 9)
                if obs_pil_patterns:
                    matched_issue_pks = list(
                        Issue.objects.filter(
                            spare__pattern_number__in=obs_pil_patterns,
                            dart_number__in=["", "pending", "Pending", "None"],
                        ).values_list("pk", flat=True)
                    )

                    if matched_issue_pks:
                        Issue.objects.filter(pk__in=matched_issue_pks).update(
                            dart_number=new_serial
                        )
                        Demand.objects.filter(
                            issue_entry_id__in=matched_issue_pks
                        ).update(dart_number=new_serial)
                        Survey.objects.filter(
                            issue_entry_id__in=matched_issue_pks
                        ).update(dart_number=new_serial)

                return Response(
                    {
                        "status": "success",
                        "message": "Dart initiated successfully",
                        "dart_id": dart.id,
                        "dart_number": dart.dart_number,
                    },
                    status=status.HTTP_201_CREATED,
                )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


@extend_schema(tags=["DART"])
class GetEquipmentDetailsAPIView(APIView):
    """
    Get all nomenclatures for a specific equipment code.
    Replaces: get_equipment_details in old views.
    """

    @extend_schema(responses={200: GetEquipmentDetailsResponseSerializer})
    def get(self, request, equipment_code, *args, **kwargs):
        try:
            equipment = Equipment.objects.get(equipment_code=equipment_code)
            nomenclatures = (
                ShipEquipment.objects.filter(equipment__equipment_code=equipment_code)
                .exclude(nomenclature__isnull=True)
                .exclude(nomenclature__exact="")
                .values_list("nomenclature", flat=True)
                .distinct()
            )
            data = {
                "equipment_code": equipment.equipment_code,
                "nomenclatures": list(nomenclatures),
            }
            return Response(data, status=status.HTTP_200_OK)
        except Exception:
            return Response(
                {"error": "Equipment not found"}, status=status.HTTP_404_NOT_FOUND
            )


@extend_schema(tags=["DART"])
class GetEquipmentObjectsAPIView(APIView):
    """
    Get all ship equipments matching an equipment code.
    Replaces: get_equipment_objects in old views.
    """

    @extend_schema(
        parameters=[OpenApiParameter("code", type=str, description="Equipment Code")],
        responses={200: GetEquipmentObjectsResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        code = request.query_params.get("code")
        try:
            equipment = Equipment.objects.get(equipment_code=code)
            ship_eq_obj = ShipEquipment.objects.filter(
                equipment__equipment_code=equipment.equipment_code
            )

            list1 = []
            for obj in ship_eq_obj:
                list1.append(
                    {
                        "id": obj.id,
                        "nomenclature": obj.nomenclature,
                    }
                )
            return Response({"ship_equipments": list1}, status=status.HTTP_200_OK)
        except Equipment.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class GetNomenclatureDetailsAPIView(APIView):
    """
    Get detailed information about a specific ShipEquipment (Nomenclature).
    Replaces: get_nomenclature_details in old views.
    """

    @extend_schema(
        parameters=[
            OpenApiParameter("nomenclature", type=str, description="ShipEquipment ID")
        ],
        responses={200: GetNomenclatureDetailsResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        nomenclature = request.query_params.get("nomenclature")
        try:
            Sequipment = ShipEquipment.objects.get(id=nomenclature)

            # Get previous DART No
            prev_dart = (
                InitiateDart.objects.filter(equipment_ship=Sequipment)
                .order_by("-id")
                .first()
            )
            prev_dart_no = prev_dart.dart_number if prev_dart else "None"

            department_name = ""
            if getattr(Sequipment, "sub_department_f_key", None) and getattr(
                Sequipment.sub_department_f_key, "department_name", None
            ):
                department_name = Sequipment.sub_department_f_key.department_name.name
            elif getattr(Sequipment, "department", None):
                department_name = Sequipment.department.name

            data = {
                "equipment_serial_no": getattr(Sequipment, "equipment_serial_no", ""),
                "equipment_id": Sequipment.id,
                "location_on_board": getattr(Sequipment, "location_on_board", ""),
                "department": department_name,
                "sub_department": (
                    Sequipment.sub_department_f_key.name
                    if getattr(Sequipment, "sub_department_f_key", None)
                    else ""
                ),
                "prev_dart_no": prev_dart_no,
            }
            return Response(data, status=status.HTTP_200_OK)
        except Exception:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)


@extend_schema(tags=["DART"])
class GetRefitPeriodsAPIView(APIView):
    """
    Get refit periods for DL II.
    Replaces: get_refit_periods in old views.
    """

    @extend_schema(responses={200: GetRefitPeriodsResponseSerializer(many=True)})
    def get(self, request, *args, **kwargs):
        qs = RefitMaintenancePeriod.objects.filter(
            maintenance_period="REFIT", occasion="DL II"
        )
        data = []
        for r in qs:
            s_date = r.actual_start_date or r.plan_start_date
            e_date = r.actual_end_date or r.plan_end_date
            s_year = str(s_date.year)[-2:] if s_date else "00"
            e_year = str(e_date.year)[-2:] if e_date else "00"
            year_label = f"{getattr(r, 'name', '')}_{s_year}_{e_year}"
            data.append({"value": year_label, "label": year_label})
        return Response(data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class MaintenancePeriodsListAPIView(APIView):
    """
    Get maintenance periods from the last 1 year.
    Replaces: maintenance_periods_list in old views.
    """

    @extend_schema(responses={200: MaintenancePeriodItemSerializer(many=True)})
    def get(self, request, *args, **kwargs):
        today = timezone.now().date()
        one_year_ago = today - timedelta(days=365)
        maintenance_periods = []

        ops_periods = OpsMaintenancePeriod.objects.filter(
            start_date__gte=one_year_ago
        ).order_by("-start_date")

        for op in ops_periods:
            is_current = bool(
                op.start_date and op.end_date and op.start_date <= today <= op.end_date
            )
            maintenance_periods.append(
                {
                    "maintenance_period": getattr(
                        op, "maintenance_period", "Operational"
                    )
                    or "Operational",
                    "occasion": getattr(op, "occasion", getattr(op, "name", "-"))
                    or "-",
                    "start_date": (
                        op.start_date.strftime(DATE_FORMAT) if op.start_date else "-"
                    ),
                    "end_date": (
                        op.end_date.strftime(DATE_FORMAT) if op.end_date else "-"
                    ),
                    "is_current": is_current,
                }
            )

        refit_periods = RefitMaintenancePeriod.objects.filter(
            Q(actual_start_date__gte=one_year_ago)
            | Q(plan_start_date__gte=one_year_ago)
        ).order_by("-actual_start_date", "-plan_start_date")

        for rp in refit_periods:
            s_date = rp.actual_start_date or rp.plan_start_date
            e_date = rp.actual_end_date or rp.plan_end_date
            is_current = bool(s_date and e_date and s_date <= today <= e_date)
            maintenance_periods.append(
                {
                    "maintenance_period": getattr(rp, "maintenance_period", "Refit")
                    or "Refit",
                    "occasion": getattr(rp, "occasion", getattr(rp, "name", "-"))
                    or "-",
                    "start_date": s_date.strftime(DATE_FORMAT) if s_date else "-",
                    "end_date": e_date.strftime(DATE_FORMAT) if e_date else "-",
                    "is_current": is_current,
                }
            )

        maintenance_periods.sort(
            key=lambda x: (
                not x["is_current"],
                (
                    datetime.strptime(x["start_date"], DATE_FORMAT)
                    if x["start_date"] != "-"
                    else datetime.min
                ),
            ),
            reverse=True,
        )
        return Response(maintenance_periods, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class RADefectListAPIView(APIView):
    """
    Get list of RA initiated defects.
    Replaces: ra_defect_list in old views.
    """

    @extend_schema(responses={200: OpenApiResponse(description="Successful Response")})
    def get(self, request, *args, **kwargs):
        defects = (
            InitiateDart.objects.open()
            .ra_initiated()
            .with_non_empty_description()
            .values("id", "dart_number", "defective_discriptions")
        )
        return Response({"defects": list(defects)}, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class MergeInitiateDefectsAPIView(APIView):
    """
    Merge multiple defects and set them as RA initiated.
    Replaces: merge_initiate_defects in old views.
    """

    @extend_schema(
        request=MergeDefectsSerializer,
        responses={200: OpenApiResponse(description="Successful operation")},
    )
    def post(self, request, *args, **kwargs):
        serializer = MergeDefectsSerializer(data=request.data)
        if serializer.is_valid():
            defect_ids = serializer.validated_data.get("defect_ids")
            # Update all at once
            InitiateDart.objects.for_ids(defect_ids).mark_ra_initiated()
            return Response(
                {"success": True, "message": "Defects merged successfully"},
                status=status.HTTP_200_OK,
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["DART"])
class MoveToDraftAPIView(APIView):
    """
    Move defects to RA or DL draft status.
    Replaces: move_to_draft in old views.
    """

    @extend_schema(
        request=MoveToDraftSerializer,
        responses={200: OpenApiResponse(description="Successful operation")},
    )
    def patch(self, request, *args, **kwargs):
        serializer = MoveToDraftSerializer(data=request.data)
        if serializer.is_valid():
            ids = serializer.validated_data.get("ids")
            draft_type = serializer.validated_data.get("type")

            if draft_type == "RA":
                InitiateDart.objects.for_ids(ids).mark_ra_draft()
            elif draft_type == "DL":
                InitiateDart.objects.for_ids(ids).mark_dl_draft()

            return Response(
                {"success": True, "message": f"Moved to {draft_type} draft"},
                status=status.HTTP_200_OK,
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["DART"])
class CheckExistingSparesAPIView(APIView):
    """
    Check if existing spares are assigned to a specific equipment.
    Replaces: check_existing_spares in old views.
    """

    @extend_schema(
        parameters=[
            OpenApiParameter("equipment_id", type=int, description="ShipEquipment ID")
        ],
        responses={200: OpenApiResponse(description="Successful Response")},
    )
    def get(self, request, *args, **kwargs):
        equipment_id = request.query_params.get("equipment_id")
        if not equipment_id:
            return Response(
                {"exists": False, "count": 0, "spares": []}, status=status.HTTP_200_OK
            )

        spares_qs = TempDartSpare.objects.for_equipment(equipment_id).active_items()
        count = spares_qs.count()

        spares_list = list(
            spares_qs.values(
                "id",
                "pattern",
                "description",
                "quantity",
                "issue_obj_id",
                "spare_id",
                "created_date",
                "modified_date",
            )
        )

        return Response(
            {"exists": count > 0, "count": count, "spares": spares_list},
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class CompleteDefectAPIView(APIView):
    """
    Mark a defect as closed (completed) and save its complete details (GET and POST).
    """

    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @extend_schema(
        responses={200: CompleteDefectGETResponseSerializer},
    )
    def get(self, request, defect_id, *args, **kwargs):
        try:
            defect_obj = InitiateDart.objects.get(id=defect_id)
        except InitiateDart.DoesNotExist:
            return Response(
                {
                    "status": "error",
                    "message": "Defect not found.",
                    "data": None,
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        issued_spares = Issue.objects.filter(
            dart_number=defect_obj.dart_number
        ).select_related("spare")

        repair_agency_list = MRepairAgency.objects.all()

        dept = getattr(
            getattr(request.user, "CustomUser_profile", None), "department", None
        )
        if dept:
            univ_id = getattr(dept, "universal_id_m_department", None)
            if univ_id:
                diagnostic_qs = MDiagnostic.objects.filter(
                    universal_id_m_department=univ_id, active=1
                )
            else:
                diagnostic_qs = MDiagnostic.objects.filter(department=dept, active=1)
        else:
            diagnostic_qs = MDiagnostic.objects.filter(active=1)

        if not diagnostic_qs.exists():
            diagnostic_qs = MDiagnostic.objects.filter(active=1)

        repair_list = MRepair.objects.all()
        delay_list = MDelay.objects.all()

        # Format dropdown lists cleanly without duplicate keys
        repair_agencies_data = [
            {
                "id": r.id,
                "repair_agency_code": r.repair_agency_code or "",
                "repair_agency_name": r.repair_agency_name or "",
            }
            for r in repair_agency_list
        ]
        diagnostics_data = [
            {
                "id": d.id,
                "diagnostic_code": d.diagnostic_code or "",
                "diagnostic_name": d.diagnostic_name or "",
            }
            for d in diagnostic_qs
        ]
        repairs_data = [
            {
                "id": r.id,
                "repair_code": r.repair_code or "",
                "repair_name": r.repair_name or "",
            }
            for r in repair_list
        ]
        delays_data = [
            {
                "id": d.id,
                "delay_code": d.delay_code or "",
                "delay_name": d.delay_name or "",
            }
            for d in delay_list
        ]

        defect_serializer = CompleteDefectGETDetailSerializer(defect_obj)
        issued_spares_serializer = IssuedSpareSerializer(issued_spares, many=True)

        payload = {
            "defect": defect_serializer.data,
            "issued_spare_obj": issued_spares_serializer.data,
            "repair_agency_list": repair_agencies_data,
            "diagnostic_list": diagnostics_data,
            "repair_list": repairs_data,
            "delay_list": delays_data,
        }

        return Response(
            {
                "status": "success",
                "message": "Defect completion details retrieved.",
                "data": payload,
            },
            status=status.HTTP_200_OK,
        )

    @extend_schema(
        request=CompleteDefectSerializer,
        responses={200: CompleteDefectPOSTResponseSerializer},
    )
    def post(self, request, defect_id, *args, **kwargs):
        try:
            defect_obj = InitiateDart.objects.get(id=defect_id)
        except InitiateDart.DoesNotExist:
            return Response(
                {
                    "status": "error",
                    "message": "Defect not found.",
                    "data": None,
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = CompleteDefectSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {
                    "status": "error",
                    "message": "Validation failed.",
                    "data": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        val = serializer.validated_data
        repair_agency = MRepairAgency.objects.filter(
            id=val.get("repair_agency_code")
        ).first()
        diagnostic = MDiagnostic.objects.filter(id=val.get("diagnostic_code")).first()

        try:
            complete_obj = CompleteDefectDart.objects.create(
                serial_no=defect_obj.dart_number or "",
                dart_details=defect_obj,
                rectified_date=val.get("rectified_date"),
                repair_agency_code=repair_agency,
                diagnostic_code=diagnostic,
                days_delay=val.get("days_delay", 0),
                other_reasons=val.get("delay_reason", ""),
                lesson_learnt=val.get("lesson_learnt", ""),
                defect_report=val.get("defect_report"),
            )

            spares_list = []
            spares_input = val.get("spares_used")
            if isinstance(spares_input, list):
                spares_list = spares_input
            elif isinstance(spares_input, str) and spares_input.strip():
                try:
                    spares_list = json.loads(spares_input)
                except Exception:
                    spares_list = []

            if not spares_list:
                spare_patterns = request.POST.getlist("spare_pattern[]") or (
                    request.data.getlist("spare_pattern[]")
                    if hasattr(request.data, "getlist")
                    else []
                )
                spare_descs = request.POST.getlist("spare_desc[]") or (
                    request.data.getlist("spare_desc[]")
                    if hasattr(request.data, "getlist")
                    else []
                )
                spare_qtys = request.POST.getlist("spare_qty[]") or (
                    request.data.getlist("spare_qty[]")
                    if hasattr(request.data, "getlist")
                    else []
                )
                for i in range(len(spare_patterns)):
                    spares_list.append(
                        {
                            "pattern": spare_patterns[i],
                            "desc": spare_descs[i] if i < len(spare_descs) else "",
                            "qty": spare_qtys[i] if i < len(spare_qtys) else 0,
                        }
                    )

            for spare in spares_list:
                pattern_no = spare.get("pattern") or spare.get("pattern_no") or ""
                description = spare.get("desc") or spare.get("description") or ""
                qty_raw = spare.get("qty") or spare.get("quantity") or 0
                if pattern_no or description:
                    DartSpareUsed.objects.create(
                        complete_dart=complete_obj,
                        pattern_no=pattern_no,
                        description=description,
                        quantity=int(qty_raw) if qty_raw else 0,
                    )

            defect_obj.is_closed = True
            defect_obj.save()
        except Exception as e:
            return Response(
                {
                    "status": "error",
                    "message": f"Failed to complete defect: {e!s}",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                "status": "success",
                "message": "Defect closed successfully and moved to history.",
                "data": None,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class GetDartDetailsAPIView(APIView):
    """
    Get all details for a specific DART defect.
    Replaces: get_dart_details in old views.
    """

    @extend_schema(
        parameters=[OpenApiParameter("dart_id", type=int, description="DART ID")],
        responses={200: GetDartDetailsResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        dart_id = request.query_params.get("dart_id")
        try:
            defect = InitiateDart.objects.get(id=dart_id)
            spares = []
            for s in defect.spares.all():
                spares.append(
                    {
                        "pattern": s.pattern or "",
                        "description": s.description or "",
                        "inventory_type": s.inventory_type or "",
                        "quantity": s.quantity or 0,
                    }
                )
            for s in defect.wed_spares.all():
                spares.append(
                    {
                        "pattern": s.pattern or "",
                        "description": s.description or "",
                        "inventory_type": "WED",
                        "quantity": s.quantity or 0,
                    }
                )
            for s in defect.mo_spares.all():
                spares.append(
                    {
                        "pattern": s.pattern or "",
                        "description": s.description or "",
                        "inventory_type": "MO",
                        "quantity": s.quantity or 0,
                    }
                )
            data = {
                "id": defect.id,
                "dart_number": defect.dart_number or "",
                "dart_date": (
                    defect.dart_date.strftime(DATE_FORMAT) if defect.dart_date else ""
                ),
                "rectification_date": (
                    defect.rectification_date.strftime(DATE_FORMAT)
                    if defect.rectification_date
                    else ""
                ),
                "nomenclature": (
                    defect.equipment_ship.nomenclature if defect.equipment_ship else ""
                ),
                "equipment_code": (
                    defect.equipment_ship.equipment.equipment_code
                    if defect.equipment_ship and defect.equipment_ship.equipment
                    else ""
                ),
                "sub_dept": (
                    defect.equipment_ship.sub_department_f_key.name
                    if defect.equipment_ship
                    and defect.equipment_ship.sub_department_f_key
                    else ""
                ),
                "department": defect.department_id.name if defect.department_id else "",
                "serial_no": (
                    defect.equipment_ship.equipment_serial_no
                    if defect.equipment_ship
                    else ""
                ),
                "location": (
                    defect.equipment_ship.location_on_board
                    if defect.equipment_ship
                    else ""
                ),
                "previous_dart_no": "",
                "symptoms": (
                    defect.symptom_code.symptom_code if defect.symptom_code else ""
                ),
                "severity": (
                    defect.severity_code.severity_name if defect.severity_code else ""
                ),
                "assistance": (
                    defect.require_assistance_for_code.required_assistance_for
                    if defect.require_assistance_for_code
                    else ""
                ),
                "defective_component": defect.defective_component or "",
                "resolved_by": (
                    defect.remark_code.description if defect.remark_code else ""
                ),
                "trial_required": "YES" if defect.trial_required else "NO",
                "trial_agency": defect.universal_id_trial_required or "",
                "description": defect.defective_discriptions or "",
                "photograph": defect.photograph.url if defect.photograph else "",
                "occasion": defect.dart_occasion or "",
                "maintenance_period": defect.maintenance_period or "",
                "rha_defect": defect.rha_defect or "",
                "created_date": (
                    defect.created_date.strftime(DATE_FORMAT)
                    if defect.created_date
                    else ""
                ),
                "is_gd": defect.is_guarantee_defect,
                "sapres_required": "YES" if defect.sapres_required else "NO",
                "guarantee_cause": defect.guarantee_cause or "",
                "op_availability": "YES" if defect.guarantee_op_availability else "NO",
                "hot_work": "YES" if defect.guarantee_hot_work else "NO",
                "repairs": defect.guarantee_repairs or "",
                "repair_date": (
                    defect.guarantee_repair_date.strftime(DATE_FORMAT)
                    if defect.guarantee_repair_date
                    else ""
                ),
                "completion_date": (
                    defect.guarantee_completion_date.strftime(DATE_FORMAT)
                    if defect.guarantee_completion_date
                    else ""
                ),
                "place": defect.guarantee_place or "",
                "spares": spares,
            }
            return Response(
                {"status": "success", "data": data}, status=status.HTTP_200_OK
            )
        except InitiateDart.DoesNotExist:
            return Response(
                {"status": "error", "message": "DART not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["DART"])
class RefitOperationalOccasionAPIView(APIView):
    """
    GET: List all refit/operational periods.
    POST: Create or edit a refit/operational period with overlap check.
    """

    @extend_schema(
        parameters=[
            OpenApiParameter(
                "search",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="Search ",
            )
        ],
        responses={200: RefitOperationalOccasionResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        all_periods = RefitMaintenancePeriod.objects.all().order_by(
            "-actual_start_date"
        )

        # Search
        search = request.query_params.get("search")
        if search:
            all_periods = all_periods.filter(
                Q(refit_category_f_key__refit_category_name__icontains=search)
                | Q(maintenance_period__icontains=search)
            )

        refit_periods = [
            period
            for period in all_periods
            if period.maintenance_period != "OPERATIONAL"
        ]

        maint_periods = [
            period
            for period in all_periods
            if period.maintenance_period == "OPERATIONAL"
        ]

        payload = {
            "title": "DART | Refit Nomenclature Create / Amend",
            "pqr": "DART-Copyright",
            "ref": "Kolkata@Indian Navy",
            "refit_periods": refit_periods,
            "maint_periods": maint_periods,
        }
        serializer = RefitOperationalOccasionResponseSerializer(payload)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        request=RefitOccasionCreateSerializer,
        responses={200: OpenApiResponse(description="Period created")},
    )
    def post(self, request, *args, **kwargs):
        """Create a new refit/operational period."""
        import re

        try:
            start_date = request.data.get("start_date")
            completion_date = request.data.get("completion_date")
            refit_type = request.data.get("refit_type")
            if not start_date or not completion_date or not refit_type:
                return Response(
                    {"status": "error", "message": "All fields are required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            s = datetime.strptime(start_date, "%Y-%m-%d").date()
            e = datetime.strptime(completion_date, "%Y-%m-%d").date()
            overlap = RefitMaintenancePeriod.objects.filter(
                occasion=refit_type,
                actual_start_date__lte=e,
                actual_end_date__gte=s,
            ).first()
            if overlap:
                return Response(
                    {
                        "status": "error",
                        "message": f"Overlap detected with {overlap.name}",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            year = s.year
            last_year_record = (
                RefitMaintenancePeriod.objects.filter(
                    occasion=refit_type, actual_start_date__year=year
                )
                .order_by("-id")
                .first()
            )

            new_seq = 1
            if last_year_record:
                try:
                    if refit_type in ["AMP", "SMP", "EAMP"]:
                        match = re.search(r"([A-Za-z]+)(\d+)-", last_year_record.name)
                        if match:
                            new_seq = int(match.group(2)) + 1
                        else:
                            parts = last_year_record.name.split("-")
                            new_seq = int(parts[1]) + 1 if len(parts) > 2 else 1
                    else:
                        parts = last_year_record.name.split("-")
                        new_seq = int(parts[1]) + 1 if len(parts) > 2 else 1
                except Exception:
                    new_seq = 1

            if refit_type in ["AMP", "SMP", "EAMP"]:
                ops_name = f"{refit_type}{new_seq}-{year}"
                mp_type = "OPERATIONAL"
            else:
                ops_name = f"{refit_type}-{year}"
                mp_type = "REFIT"

            created_period = RefitMaintenancePeriod.objects.create(
                name=ops_name,
                maintenance_period=mp_type,
                occasion=refit_type,
                actual_start_date=s,
                actual_end_date=e,
            )
            return Response(
                {
                    "status": "success",
                    "message": f"Nomenclature Generated: {ops_name}",
                    "data": {
                        "id": created_period.id,
                        "name": created_period.name,
                        "occasion": created_period.occasion,
                        "actual_start_date": str(created_period.actual_start_date),
                        "actual_end_date": str(created_period.actual_end_date),
                    },
                },
                status=status.HTTP_200_OK,
            )
        except Exception as ex:
            return Response(
                {"status": "error", "message": str(ex)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @extend_schema(
        request=RefitOccasionEditSerializer,
        responses={200: OpenApiResponse(description="Period dates updated")},
    )
    def put(self, request, *args, **kwargs):
        """Edit dates of an existing refit/operational period."""
        try:
            period_id = request.data.get("period_id")
            start_date = request.data.get("start_date")
            completion_date = request.data.get("completion_date")
            if not period_id or not start_date or not completion_date:
                return Response(
                    {"status": "error", "message": "All fields are required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            s = datetime.strptime(start_date, "%Y-%m-%d").date()
            e = datetime.strptime(completion_date, "%Y-%m-%d").date()
            try:
                period = RefitMaintenancePeriod.objects.get(id=period_id)
            except RefitMaintenancePeriod.DoesNotExist:
                return Response(
                    {
                        "status": "error",
                        "message": f"Period with ID {period_id} not found.",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            overlap = (
                RefitMaintenancePeriod.objects.filter(
                    occasion=period.occasion,
                    actual_start_date__lte=e,
                    actual_end_date__gte=s,
                )
                .exclude(id=period_id)
                .first()
            )
            if overlap:
                return Response(
                    {
                        "status": "error",
                        "message": f"Overlap detected with {overlap.name}",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            period.actual_start_date = s
            period.actual_end_date = e
            period.save()
            return Response(
                {
                    "status": "success",
                    "message": f"Dates updated for {period.name}.",
                    "data": {
                        "id": period.id,
                        "name": period.name,
                        "occasion": period.occasion,
                        "actual_start_date": str(period.actual_start_date),
                        "actual_end_date": str(period.actual_end_date),
                    },
                },
                status=status.HTTP_200_OK,
            )
        except Exception as ex:
            return Response(
                {"status": "error", "message": str(ex)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["DART"])
class GeneratedDLIIReportsAPIView(APIView):
    """
    List generated DL-II report groups.

    Replaces the reference project's template view: dart/generate_report/.
    """

    @extend_schema(responses={200: GeneratedDLIIReportSerializer(many=True)})
    def get(self, request, *args, **kwargs):
        reports = (
            RADLMaster.objects.annotate(
                total_dl_rows=Count(
                    "initiateradl",
                    filter=Q(initiateradl__dl_type="DL-II")
                    & ~Q(initiateradl__status="DELETED"),
                    distinct=True,
                )
            )
            .filter(total_dl_rows__gt=0)
            .order_by("-id")
        )
        serializer = GeneratedDLIIReportSerializer(reports, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class GeneratedDLIIReportRowsAPIView(APIView):
    """
    Return the generated DL-II rows inside one RA/DL master report.

    Replaces the reference project's report_inner_rows/<id>/ JSON endpoint.
    """

    @extend_schema(responses={200: GeneratedDLIIReportRowSerializer(many=True)})
    def get(self, request, report_id, *args, **kwargs):
        rows = get_generated_dl_ii_rows(report_id)
        serializer = GeneratedDLIIReportRowSerializer(rows, many=True)
        return Response(
            {"status": True, "data": serializer.data},
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class GeneratedDLIIReportExportAPIView(APIView):
    """
    Download generated DL-II report rows.

    The reference project exports ACCDB using Java/JAR assets. Those assets are
    not present in this DRF project yet, so this endpoint exports the same row
    data as CSV and keeps the report workflow available.
    """

    def get(self, request, report_id, *args, **kwargs):
        rows = []
        for row in get_generated_dl_ii_rows(report_id):
            rows.append(
                [
                    row.id,
                    row.dl_no,
                    row.dl_key,
                    row.dl_type,
                    row.status,
                    (
                        row.initiate_dart.equipment_ship.nomenclature
                        if row.initiate_dart and row.initiate_dart.equipment_ship
                        else ""
                    ),
                    (
                        row.initiate_dart.defective_discriptions
                        if row.initiate_dart
                        else ""
                    ),
                    getattr(row.initiate_dart, "defective_component", ""),
                ]
            )
        return build_csv_response(
            f"generated_dl_ii_report_{report_id}.csv",
            [
                "ID",
                "DL No",
                "RA/DL Name",
                "DL Type",
                "Status",
                "Equipment",
                "Description",
                "Defective Component",
            ],
            rows,
        )


def build_csv_response(filename, headers, rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def build_xlsx_response(filename, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DART"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@extend_schema(tags=["DART"])
class PendingDefectsExportAPIView(APIView):
    export_headers = [
        "Sr No",
        "DART No",
        "Defect Date",
        "Equipment",
        "Nomenclature",
        "Defective Component",
        "Description",
        "Severity",
        "Occasion",
    ]
    export_format = "csv"

    def get(self, request, *args, **kwargs):
        rows = build_pending_defect_export_rows()
        if self.export_format == "xlsx":
            return build_xlsx_response(
                "pending_defects.xlsx",
                self.export_headers,
                rows,
            )
        return build_csv_response("pending_defects.csv", self.export_headers, rows)

    @extend_schema(
        request=ExportPendingDefectsACCDBSerializer,
        responses={200: ExportPendingDefectsACCDBResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        serializer = ExportPendingDefectsACCDBSerializer(data=request.data)
        if serializer.is_valid():
            val = serializer.validated_data
            dart_ids = val.get("dart_ids", [])
            yard = val.get("yard", "")
            export_format = val.get("export_format", "CSV")
            dl_type = val.get("dl_type", "RA")
            remarks_data = val.get("remarks_data", {})
            ss_remarks_data = val.get("ss_remarks_data", {})

            dtg_date = val.get("dtg_date", "")
            dtg_hour = val.get("dtg_hour", "")
            dtg_minute = val.get("dtg_minute", "")
            dtg = val.get("dtg", "")

            if not dtg and dtg_date:
                time_str = f"{dtg_hour}:{dtg_minute}".strip(":")
                dtg = f"{dtg_date} {time_str}".strip()

            if dart_ids:
                ra_group_id = f"RA-{now().strftime('%Y%m%d%H%M%S')}"

                InitiateDart.objects.filter(id__in=dart_ids).update(
                    is_ra_draft=False, is_ra_initiate=True
                )

                darts = InitiateDart.objects.filter(id__in=dart_ids)
                for d in darts:
                    d_id_str = str(d.id)
                    d_id_int = d.id
                    ss_rem_val = (
                        ss_remarks_data.get(d_id_str)
                        or ss_remarks_data.get(d_id_int)
                        or ss_remarks_data.get(str(d_id_int))
                    )
                    if ss_rem_val:
                        if str(ss_rem_val).isdigit():
                            ss_obj = ChMasterShipRemarksBy.objects.filter(
                                id=int(ss_rem_val)
                            ).first()
                        else:
                            ss_obj = ChMasterShipRemarksBy.objects.filter(
                                Q(description__iexact=str(ss_rem_val))
                                | Q(description__icontains=str(ss_rem_val))
                            ).first()
                        if ss_obj:
                            d.remark_code = ss_obj
                            d.save(update_fields=["remark_code"])

                    rem_text = (
                        remarks_data.get(d_id_str) or remarks_data.get(d_id_int) or ""
                    )
                    add_rem = f"DTG: {dtg}" if dtg else ""
                    InitiateRADL.objects.update_or_create(
                        initiate_dart=d,
                        dl_type=dl_type or "RA",
                        defaults={
                            "ra_grup_id": ra_group_id,
                            "remarks": rem_text,
                            "additional_remarks": add_rem,
                            "status": "GENERATED",
                            "is_active": True,
                        },
                    )

                export_dir = os.path.join(settings.MEDIA_ROOT, "exports")
                os.makedirs(export_dir, exist_ok=True)

                fmt = str(export_format or "CSV").lower()
                ext = fmt if fmt in ["csv", "xlsx", "pdf"] else "csv"
                filename = f"RA_Export_{ra_group_id}.{ext}"
                filepath = os.path.join(export_dir, filename)

                export_headers = [
                    "Ser No",
                    "DART No",
                    "Defect Date",
                    "Yard",
                    "Status",
                    "Equipment",
                    "Nomenclature",
                    "Description",
                    "Additional Remarks",
                    "SS Remarks",
                ]

                export_rows = []
                for idx, d in enumerate(darts, start=1):
                    eq_name = ""
                    nom_name = ""
                    if d.equipment_ship:
                        nom_name = d.equipment_ship.nomenclature or ""
                        if d.equipment_ship.equipment:
                            eq_name = str(
                                d.equipment_ship.equipment.equipment_class or ""
                            )
                    elif d.equipment_ems:
                        eq_name = d.equipment_ems.name or ""

                    d_id_str = str(d.id)
                    d_id_int = d.id
                    rem_text = (
                        remarks_data.get(d_id_str) or remarks_data.get(d_id_int) or ""
                    )
                    ss_rem_desc = d.remark_code.description if d.remark_code else ""

                    export_rows.append(
                        [
                            idx,
                            d.dart_number or "",
                            d.dart_date.strftime("%Y-%m-%d") if d.dart_date else "",
                            yard,
                            "Closed" if d.is_closed else "Open",
                            eq_name,
                            nom_name,
                            d.defective_discriptions or "",
                            rem_text,
                            ss_rem_desc,
                        ]
                    )

                if ext == "xlsx":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "RA Export"
                    ws.append(export_headers)
                    for r in export_rows:
                        ws.append(r)
                    wb.save(filepath)
                else:
                    import csv

                    with open(filepath, "w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerow(export_headers)
                        writer.writerows(export_rows)

                download_url = request.build_absolute_uri(
                    f"{settings.MEDIA_URL}exports/{filename}"
                )

                return Response(
                    {
                        "status": "success",
                        "message": "RA records saved in database successfully.",
                        "data": {
                            "ra_group_id": ra_group_id,
                            "total_records": len(dart_ids),
                            "yard": yard,
                            "dtg": dtg,
                            "export_format": export_format,
                            "dl_type": dl_type or "RA",
                            "download_url": download_url,
                        },
                    },
                    status=status.HTTP_200_OK,
                )

        return self.get(request, *args, **kwargs)


class PendingDefectsXLSXExportAPIView(PendingDefectsExportAPIView):
    export_format = "xlsx"


@extend_schema(tags=["DART"])
class DLIISelectedExportAPIView(APIView):
    export_headers = [
        "Sr No",
        "DART No",
        "Equipment",
        "Nomenclature",
        "Defective Component",
        "Description",
        "Severity",
        "Maintenance Period",
        "Occasion",
    ]
    filename = "pending_defects_dl2.csv"

    def post(self, request, *args, **kwargs):
        try:
            dart_ids = get_export_dart_ids(request.data)
        except (TypeError, ValueError, json.JSONDecodeError):
            return Response(
                {"status": "error", "message": "Invalid export payload"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        rows = build_dl_ii_export_rows(dart_ids)
        if not rows:
            return Response(
                {"status": "error", "message": "No data found"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return build_csv_response(self.filename, self.export_headers, rows)


@extend_schema(tags=["DART"])
class DartHistoryPDFAPIView(APIView):
    def get(self, request, id, *args, **kwargs):
        try:
            defect = InitiateDart.objects.select_related("equipment_ship").get(id=id)
        except InitiateDart.DoesNotExist:
            return Response(
                {"status": "error", "message": "DART not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        pdf = build_simple_pdf(
            f"DART History - {defect.dart_number or defect.id}",
            [
                f"DART No: {defect.dart_number or ''}",
                f"Defect Date: {defect.dart_date or ''}",
                f"Description: {defect.defective_discriptions or ''}",
                f"Maintenance Period: {defect.maintenance_period or ''}",
                f"Occasion: {defect.dart_occasion or ''}",
            ],
        )
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="dart_history_{defect.id}.pdf"'
        )
        return response


@extend_schema(tags=["DART"])
class GuaranteeDefectPDFAPIView(APIView):
    def get(self, request, dart_id, *args, **kwargs):
        try:
            defect = InitiateDart.objects.select_related("equipment_ship").get(
                id=dart_id
            )
        except InitiateDart.DoesNotExist:
            return Response(
                {"status": "error", "message": "DART not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        equipment_ship = defect.equipment_ship
        pdf = build_simple_pdf(
            f"Guarantee Defect - {defect.dart_number or defect.id}",
            [
                f"DART No: {defect.dart_number or ''}",
                f"Nomenclature: {equipment_ship.nomenclature if equipment_ship else ''}",
                f"Component: {defect.defective_component or ''}",
                f"Description: {defect.defective_discriptions or ''}",
                f"Repair Date: {defect.guarantee_repair_date or ''}",
                f"Completion Date: {defect.guarantee_completion_date or ''}",
            ],
        )
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="gd_{defect.id}.pdf"'
        return response


@extend_schema(tags=["DART"])
class SaveOEMSpareAPIView(APIView):
    """
    Save OEM spare details.
    """

    @extend_schema(
        request=SaveOEMSpareSerializer,
        responses={200: OpenApiResponse(description="Success")},
    )
    def post(self, request, *args, **kwargs):
        serializer = SaveOEMSpareSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {"success": False, "error": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            val = serializer.validated_data
            denomination = val["denomination_id"]
            pattern_number = val["pattern_number"].strip()
            description = val["description"].strip()

            pil_class = EquipmentClass.objects.filter(name="SPARES PIL").first()
            pil_authority = Authority.objects.filter(name="PIL").first()

            spare, created = Spares.objects.get_or_create(
                pattern_number=pattern_number,
                defaults={
                    "description": description,
                    "denomination": denomination,
                    "equipment_class": pil_class,
                    "authority": pil_authority,
                },
            )
            return Response(
                {
                    "success": True,
                    "pattern_number": spare.pattern_number,
                    "description": spare.description,
                    "denomination_name": (
                        spare.denomination.name if spare.denomination else None
                    ),
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                {"success": False, "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["DART"])
class GetSubDeptDefectsAPIView(APIView):
    """
    Get defects for a specific sub-department.
    """

    @extend_schema(
        parameters=[
            OpenApiParameter("sub_dept", type=str, description="Sub-department name")
        ],
        responses={200: GetSubDeptDefectsResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        sub_dept_name = request.query_params.get("sub_dept")
        if not sub_dept_name:
            return Response(
                {"success": False, "error": "No sub-dept provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        darts = InitiateDart.objects.filter(
            is_closed=False,
            equipment_ship__sub_department_f_key__name=sub_dept_name,
            severity_code__severity_name__icontains="Non-Ops",
        ).select_related("equipment_ship", "equipment_ship__equipment", "severity_code")

        data = []
        today = timezone.now().date()
        for d in darts:
            days = (today - d.dart_date).days if d.dart_date else 0
            data.append(
                {
                    "dart_number": d.dart_number,
                    "nomenclature": (
                        d.equipment_ship.nomenclature if d.equipment_ship else "N/A"
                    ),
                    "category": (
                        d.severity_code.severity_name if d.severity_code else "N/A"
                    ),
                    "defective_since": f"{days} days",
                }
            )
        return Response({"success": True, "data": data}, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class GetMaintenancePeriodNameAPIView(APIView):
    """
    Get period names and start/end dates filtered by maintenance_period (e.g. Operational, Refit).
    """

    @extend_schema(
        parameters=[
            OpenApiParameter(
                "maintenance_period",
                type=str,
                description="Maintenance Period Type (e.g. Operational, Refit)",
            )
        ],
        responses={200: GetMaintenancePeriodNameResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        m_period = request.query_params.get("maintenance_period")
        if not m_period:
            return Response(
                {
                    "status": "error",
                    "message": "maintenance_period query parameter is required",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Normalize to uppercase (e.g. "Operational" -> "OPERATIONAL", "Refit" -> "REFIT")
        m_period_upper = m_period.upper()

        periods = RefitMaintenancePeriod.objects.filter(
            maintenance_period=m_period_upper
        )
        data = []
        for p in periods:
            start_date = p.actual_start_date or p.plan_start_date
            end_date = p.actual_end_date or p.plan_end_date
            data.append(
                {
                    "id": p.id,
                    "name": p.name or "",
                    "start_date": (
                        start_date.strftime("%Y-%m-%d") if start_date else None
                    ),
                    "end_date": end_date.strftime("%Y-%m-%d") if end_date else None,
                }
            )
        return Response({"status": "success", "data": data}, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class CreateRAAjaxAPIView(APIView):
    """
    Create an RA entry via AJAX.
    """

    @extend_schema(
        request=CreateRAAjaxSerializer,
        responses={200: OpenApiResponse(description="RA created")},
    )
    def post(self, request, *args, **kwargs):
        from django.utils.timezone import now

        serializer = CreateRAAjaxSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {"error": "Validation failed", "details": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            val = serializer.validated_data
            dart_ids = val.get("dart_ids")
            dl_type = val.get("dl_type")

            ra_group_id = f"RA-{now().strftime('%Y%m%d%H%M%S')}"
            darts = InitiateDart.objects.filter(id__in=dart_ids)

            for d in darts:
                if d.is_ra_initiate:
                    continue
                InitiateRADL.objects.create(
                    initiate_dart=d,
                    dl_type=dl_type,
                    ra_grup_id=ra_group_id,
                    status="DRAFT",
                )
                d.is_ra_initiate = True
                d.save(update_fields=["is_ra_initiate"])

            return Response(
                {"success": True, "ra_group_id": ra_group_id}, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"success": False, "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["DART"])
class DartHistoryDefaultDataAPIView(APIView):
    """
    Returns all closed DART records for history table + unique filter options.
    Replaces: dart_history_default_data in old views.
    """

    @extend_schema(responses={200: DartHistoryDefaultDataResponseSerializer})
    def get(self, request, *args, **kwargs):
        cache_key = make_cache_key("dart:history-default", request)
        cached_payload = cache.get(cache_key)
        if cached_payload is not None:
            return Response(cached_payload, status=status.HTTP_200_OK)

        darts = get_closed_dart_history_queryset()

        data = []
        dart_occasions = set()
        sub_departments = set()
        dart_types = set()
        departments = set()
        dart_maintenance_period = set()
        equipments = set()
        equipment_nomenclatures = set()

        for dart in darts:
            row = build_dart_history_row(
                dart, "maintenanceType", maintenance_default="ops"
            )
            data.append(row)

            all_dart_occasions = {"All", "Normal Defect", "Normal RA"}
            all_maintenance_period = {"All", "Operational", "Refit"}
            dart_maintenance_period.update(all_maintenance_period)
            dart_occasions.update(all_dart_occasions)
            dart_types.update(all_dart_occasions)
            departments.add(dart.department_id.name if dart.department_id else "")
            if row["subDepartment"]:
                sub_departments.add(row["subDepartment"])
            equipments.add(row["equipmentName"])
            equipment_nomenclatures.add(row["equipmentNomenclature"])

        payload = {
            "data": data,
            "filters": {
                "dartMaintenancePeriod": sorted(dart_maintenance_period),
                "dartOccasions": sorted(dart_occasions),
                "subDepartments": sorted(sub_departments),
                "dartTypes": sorted(dart_types),
                "departments": sorted(departments),
                "equipments": sorted(equipments),
                "equipment_nomenclatures": sorted(equipment_nomenclatures),
            },
        }
        cache.set(cache_key, payload, timeout=300)
        return Response(payload, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class AddDartFromOBSAPIView(APIView):
    """
    Get metadata needed for add_dart_from_obs page (spares data, equipment, serial).
    """

    @extend_schema(
        parameters=[
            OpenApiParameter("pks", type=str, description="Comma-separated spare PKs"),
            OpenApiParameter("equipment_class", type=str),
            OpenApiParameter(
                "descriptions", type=str, description="|| separated descriptions"
            ),
            OpenApiParameter(
                "patterns", type=str, description="Comma-separated patterns"
            ),
            OpenApiParameter("inventory_type", type=str),
            OpenApiParameter("equipment_code", type=str),
        ],
        responses={200: OpenApiResponse(description="Dart from OBS metadata")},
    )
    def get(self, request, *args, **kwargs):
        pks = request.query_params.get("pks", "")
        equipment_class = request.query_params.get("equipment_class", "")
        descriptions = request.query_params.get("descriptions", "")
        patterns = request.query_params.get("patterns", "")
        inventory_type = request.query_params.get("inventory_type", "OBS")

        pk_list = pks.split(",") if pks else []
        desc_list = descriptions.split("||") if descriptions else []
        pattern_list = patterns.split(",") if patterns else []
        spares_data = list(zip(pk_list, desc_list, pattern_list))

        return Response(
            {
                "spares_data": spares_data,
                "equipment_class": equipment_class,
                "inventory_type": inventory_type,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class CreateDLFunAPIView(APIView):
    """
    Create DL-II entries for defects and mark them as DL draft.
    Replaces: create_dl_fun POST in old views.
    """

    @extend_schema(
        request=CreateDLFunSerializer,
        responses={200: CreateDLFunResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        serializer = CreateDLFunSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {
                    "status": "error",
                    "message": "Validation failed",
                    "details": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        val = serializer.validated_data
        defect_ids = val.get("dl_defect_ids")

        try:
            dl_no = InitiateRADL.objects.filter(dl_type="DL-II")
            dart_id = [dart.initiate_dart_id for dart in dl_no]
            dart_objects = InitiateDart.objects.filter(id__in=defect_ids).exclude(
                id__in=dart_id
            )
            dart_objects.update(is_dl_draft=True)

            for d in dart_objects:
                InitiateRADL.objects.update_or_create(
                    initiate_dart=d, dl_type="DL-II", status="DRAFT"
                )

            draft_ra = InitiateRADL.objects.filter(dl_type="DL-II", status="DRAFT")
            draft_serializer = CreateDLDefectDataSerializer(draft_ra, many=True)

            refit_qs = RefitMaintenancePeriod.objects.all()
            refit_serializer = CreateDLRefitItemSerializer(refit_qs, many=True)

            ship_remarks = ChMasterShipRemarksBy.objects.filter(active=1)
            ship_remarks_data = [
                {"id": sr.id, "description": sr.description or ""}
                for sr in ship_remarks
            ]

            return Response(
                {
                    "status": "success",
                    "message": "DL initiated successfully.",
                    "data": {
                        "draft_data": draft_serializer.data,
                        "refit_list": refit_serializer.data,
                        "ship_remarks_list": ship_remarks_data,
                    },
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["DART"])
class SaveDLRowsAPIView(APIView):
    """
    Save DL row remarks for draft DL entries.
    Replaces: save_dl_rows POST in old views.
    """

    @extend_schema(
        request=SaveDLRowsSerializer,
        responses={200: SaveDLRowsResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        import json

        rows = request.data.get("rows", [])
        if isinstance(rows, str):
            try:
                rows = json.loads(rows)
            except Exception:
                rows = []

        if not rows:
            return Response(
                {"status": "error", "message": "No rows provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for row in rows:
            dart_id = row.get("dart_id") or row.get("dl_id")
            additional_remarks = row.get("additional_remarks") or row.get(
                "additional_remark", ""
            )
            remarks = row.get("remarks") or row.get("ss_remark", "")

            if dart_id:
                dl_obj = InitiateRADL.objects.filter(id=dart_id).first()
                if dl_obj:
                    dl_obj.additional_remarks = additional_remarks
                    dl_obj.remarks = remarks
                    dl_obj.is_active = True
                    dl_obj.save(
                        update_fields=[
                            "additional_remarks",
                            "remarks",
                            "is_active",
                            "updated_date",
                        ]
                    )

                    if remarks and dl_obj.initiate_dart:
                        if str(remarks).isdigit():
                            ss_obj = ChMasterShipRemarksBy.objects.filter(
                                id=int(remarks)
                            ).first()
                        else:
                            ss_obj = ChMasterShipRemarksBy.objects.filter(
                                Q(description__iexact=str(remarks))
                                | Q(description__icontains=str(remarks))
                            ).first()
                        if ss_obj:
                            dl_obj.initiate_dart.remark_code = ss_obj
                            dl_obj.initiate_dart.save(update_fields=["remark_code"])

        return Response(
            {"status": "success", "message": "DL rows saved successfully"},
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class DeleteDLRowAPIView(APIView):
    """
    Delete a draft DL row.
    Replaces: delete_dl_row POST in old views.
    """

    @extend_schema(
        request=DeleteDLRowSerializer,
        responses={200: GenericSuccessResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        dl_id = request.data.get("dl_id") or request.POST.get("dl_id")
        if not dl_id:
            return Response(
                {"status": "error", "message": "DL row id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            dl_id = int(dl_id)
        except (TypeError, ValueError):
            return Response(
                {"status": "error", "message": "Invalid DL row id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            dl_row = InitiateRADL.objects.get(id=dl_id, dl_type="DL-II")
        except InitiateRADL.DoesNotExist:
            return Response(
                {"status": "error", "message": "DL row not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        initiate_dart_id = dl_row.initiate_dart_id
        dl_row.status = "DELETED"
        dl_row.is_active = False
        dl_row.save(update_fields=["status", "is_active", "updated_date"])

        still_in_draft = InitiateRADL.objects.filter(
            initiate_dart_id=initiate_dart_id, dl_type="DL-II", status="DRAFT"
        ).exists()
        if not still_in_draft:
            InitiateDart.objects.filter(id=initiate_dart_id).update(is_dl_draft=False)

        return Response(
            {"status": "success", "message": "DL row deleted successfully"},
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class ExportPendingDefectsDL2APIView(APIView):
    """
    Export DL-II pending defects, assign DL numbers, create RADLMaster, and generate export file.
    Replaces: export_pending_defects_dl2_* POST endpoints.
    """

    @extend_schema(
        request=ExportPendingDefectsDL2Serializer,
        responses={200: ExportPendingDefectsDL2ResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        serializer = ExportPendingDefectsDL2Serializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {
                    "status": "error",
                    "message": "Validation failed",
                    "details": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        val = serializer.validated_data
        yard = val.get("yard", "")
        export_format = val.get("export_format", "CSV")
        refit_type_id = val.get("refit_Type")
        row_data = val.get("row_data", [])

        if isinstance(row_data, str):
            try:
                row_data = json.loads(row_data)
            except Exception:
                row_data = []

        if not row_data:
            return Response(
                {"status": "error", "message": "No row data provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        refit_obj = None
        refit_name = ""
        if refit_type_id:
            refit_obj = RefitMaintenancePeriod.objects.filter(id=refit_type_id).first()
            if refit_obj:
                refit_name = refit_obj.occasion or ""

        date_str = date.today().strftime("%d%m%Y")
        ra_dl_name = f"DLII-2-{date_str}"

        created_ra_dl_obj = RADLMaster.objects.create(
            ra_dl_name=ra_dl_name,
            refit_type_f_key=refit_obj,
            refit_type_name=refit_name,
            dockyard_name=yard,
        )

        processed_ids = []
        export_rows = []

        headers = [
            "Ser No",
            "DL No",
            "DART No",
            "Defect Date",
            "Closing Date",
            "Status",
            "Equipment",
            "Description",
            "Additional Remarks",
            "SS Remarks",
        ]

        for idx, r in enumerate(row_data, start=1):
            dart_id = r.get("dart_id")
            dl_id = r.get("dl_id")
            dl_number = r.get("dl_number") or r.get("dl_no", "")
            add_remark = r.get("additional_remark") or r.get("additional_remarks", "")
            ss_remark = r.get("ss_remark") or r.get("remarks", "")

            dl_obj = None
            if dl_id:
                dl_obj = InitiateRADL.objects.filter(id=dl_id).first()
            elif dart_id:
                dl_obj = InitiateRADL.objects.filter(
                    initiate_dart_id=dart_id, dl_type="DL-II"
                ).first()

            dart_obj = None
            if dl_obj:
                dart_obj = dl_obj.initiate_dart
                dl_obj.dl_no = str(dl_number)
                dl_obj.additional_remarks = add_remark
                dl_obj.remarks = ss_remark
                dl_obj.status = "GENERATED"
                dl_obj.radl_master = created_ra_dl_obj
                dl_obj.ra_grup_id = ra_dl_name
                dl_obj.is_active = True
                dl_obj.save(
                    update_fields=[
                        "dl_no",
                        "additional_remarks",
                        "remarks",
                        "status",
                        "radl_master",
                        "ra_grup_id",
                        "is_active",
                        "updated_date",
                    ]
                )
            elif dart_id:
                dart_obj = InitiateDart.objects.filter(id=dart_id).first()
                if dart_obj:
                    dl_obj = InitiateRADL.objects.create(
                        initiate_dart=dart_obj,
                        dl_no=str(dl_number),
                        dl_type="DL-II",
                        additional_remarks=add_remark,
                        remarks=ss_remark,
                        status="GENERATED",
                        radl_master=created_ra_dl_obj,
                        ra_grup_id=ra_dl_name,
                        is_active=True,
                    )

            if dart_obj:
                dart_obj.is_dl_initiate = True
                dart_obj.is_dl_draft = False

                if ss_remark:
                    if str(ss_remark).isdigit():
                        ss_obj = ChMasterShipRemarksBy.objects.filter(
                            id=int(ss_remark)
                        ).first()
                    else:
                        ss_obj = ChMasterShipRemarksBy.objects.filter(
                            Q(description__iexact=str(ss_remark))
                            | Q(description__icontains=str(ss_remark))
                        ).first()
                    if ss_obj:
                        dart_obj.remark_code = ss_obj

                dart_obj.save(
                    update_fields=["is_dl_initiate", "is_dl_draft", "remark_code"]
                )

                eq_name = ""
                if dart_obj.equipment_ship and dart_obj.equipment_ship.equipment:
                    eq_name = str(
                        dart_obj.equipment_ship.equipment.equipment_class or ""
                    )
                elif dart_obj.equipment_ems:
                    eq_name = dart_obj.equipment_ems.name or ""

                export_rows.append(
                    [
                        idx,
                        dl_number,
                        dart_obj.dart_number or "",
                        dart_obj.dart_date.strftime("%Y-%m-%d")
                        if dart_obj.dart_date
                        else "",
                        dart_obj.rectification_date.strftime("%Y-%m-%d")
                        if dart_obj.rectification_date
                        else "",
                        "Closed" if dart_obj.is_closed else "Open",
                        eq_name,
                        dart_obj.defective_discriptions or "",
                        add_remark,
                        ss_remark,
                    ]
                )
                processed_ids.append(dart_obj.id)

        export_dir = os.path.join(settings.MEDIA_ROOT, "exports")
        os.makedirs(export_dir, exist_ok=True)

        fmt = str(export_format or "CSV").lower()
        ext = fmt if fmt in ["csv", "xlsx", "pdf"] else "csv"
        filename = f"DL2_Export_{ra_dl_name}_{yard}.{ext}"
        filepath = os.path.join(export_dir, filename)

        if ext == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "DL2 Export"
            ws.append(headers)
            for r in export_rows:
                ws.append(r)
            wb.save(filepath)
        else:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(export_rows)

        download_url = request.build_absolute_uri(
            f"{settings.MEDIA_URL}exports/{filename}"
        )

        return Response(
            {
                "status": "success",
                "message": "DL-II records assigned and exported successfully.",
                "data": {
                    "ra_dl_name": ra_dl_name,
                    "total_records": len(processed_ids),
                    "yard": yard,
                    "download_url": download_url,
                },
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class CreateRADLFunAPIView(APIView):
    """
    Create RA/DL: marks defects as RA initiated.
    Replaces: create_ra_dl_fun POST in old views.
    """

    @extend_schema(
        request=CreateRADLFunSerializer,
        responses={200: CreateRADLFunResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        defect_ids = request.data.get("defect_ids", [])
        if isinstance(defect_ids, str):
            defect_ids = json.loads(defect_ids)

        InitiateDart.objects.filter(id__in=defect_ids).update(
            is_ra_draft=False, is_ra_initiate=True
        )

        dart_objects = InitiateDart.objects.filter(id__in=defect_ids).select_related(
            "equipment_ems", "equipment_ship"
        )
        serializer = CreateRADLDefectDataSerializer(dart_objects, many=True)

        ship_remarks = ChMasterShipRemarksBy.objects.filter(active=1)
        ship_remarks_data = [
            {"id": sr.id, "description": sr.description or ""} for sr in ship_remarks
        ]

        return Response(
            {
                "status": "success",
                "message": "RA initiated successfully.",
                "data": {
                    "dart_objects": serializer.data,
                    "type": "RA",
                    "ship_remarks_list": ship_remarks_data,
                },
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class AllDataOfRAAPIView(APIView):
    """
    Get all data of RA for export.
    Replaces: all_data_of_ra in old views.
    """

    @extend_schema(
        request=AllDataOfRASerializer,
        responses={200: OpenApiResponse(description="RA Data")},
    )
    def post(self, request, *args, **kwargs):
        import json

        yard = request.data.get("yard", "")
        export_format = request.data.get("export_format", "")
        dart_ids = request.data.get("dart_ids", [])
        if isinstance(dart_ids, str):
            dart_ids = json.loads(dart_ids)

        dispatch = dispatch_task(export_all_ra_data_task, yard, export_format, dart_ids)
        if dispatch["queued"]:
            return accepted_task_response(
                request,
                dispatch["task"],
                "RA export data build queued for background processing.",
            )
        return sync_task_response(
            dispatch["result"],
            "Background task service unavailable. RA export data build completed synchronously.",
        )

        dart_qs = InitiateDart.objects.filter(id__in=dart_ids)
        dart_data = []
        for obj in dart_qs:
            data = model_to_dict(obj)
            data["equipment_ship"] = (
                obj.equipment_ship.id if obj.equipment_ship else None
            )
            data["equipment_ems"] = (
                obj.equipment_ems.id if getattr(obj, "equipment_ems", None) else None
            )
            data["remark_code"] = str(obj.remark_code) if obj.remark_code else None
            dart_data.append(data)

        return Response(
            {
                "yard": yard,
                "export_format": export_format,
                "total_records": len(dart_data),
                "dart_data": dart_data,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class DartHistoryFilteredDataAPIView(APIView):
    """
    Get filtered historical data for DART based on POST body filters.
    Replaces: dart_history_filtered_data in old views.
    """

    @extend_schema(
        request=DartHistoryFilteredDataSerializer,
        responses={200: DartHistoryFilteredDataResponseSerializer},
    )
    def post(self, request, *args, **kwargs):
        body = request.data

        darts = get_closed_dart_history_queryset()

        # Maintenance filter
        maintenance = body.get("maintenancePeriod")
        if maintenance == "Refit":
            darts = darts.filter(maintenance_period="REFIT")

        # Dart Type filter
        dart_type = body.get("dartType")
        if dart_type and dart_type != "ALL":
            darts = darts.filter(dart_occasion__iexact=dart_type.strip())

        # Equipment search
        equipment_search = body.get("equipmentSearch") or body.get("equipmentName")
        if equipment_search:
            search_term = equipment_search.strip()
            darts = darts.filter(
                Q(equipment_ship__equipment__equipment_class__icontains=search_term)
                | Q(equipment_ship__nomenclature__icontains=search_term)
            )

        # Sub-Department filter
        sub_department = body.get("subDepartment")
        if sub_department and sub_department != "ALL":
            darts = darts.filter(
                equipment_ship__sub_department_f_key__name__iexact=sub_department.strip()
            )

        # Department filter
        department = body.get("department")
        if department and department != "ALL":
            darts = darts.filter(department_id__name__iexact=department.strip())

        # Defect Date filter
        defect_date_from = body.get("defectDateFrom")
        defect_date_to = body.get("defectDateTo")
        if defect_date_from:
            try:
                from_date = datetime.strptime(defect_date_from, "%Y-%m-%d")
                darts = darts.filter(dart_date__gte=from_date)
            except Exception:
                pass
        if defect_date_to:
            try:
                to_date = datetime.strptime(defect_date_to, "%Y-%m-%d")
                darts = darts.filter(dart_date__lte=to_date)
            except Exception:
                pass

        data = [build_dart_history_row(dart, "maintenancePeriod") for dart in darts]

        return Response({"data": data}, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class GetDartSparesDataAPIView(APIView):
    """
    Get Dart spares data by type: mo, mo_mapped, wed, wed_mapped.
    """

    @extend_schema(
        parameters=[
            OpenApiParameter(
                "type",
                type=str,
                description="Spare type: mo, mo_mapped, wed, wed_mapped, obs",
            ),
            OpenApiParameter("equipment_id", type=str, required=False),
        ],
        responses={200: GetDartSparesDataResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        spare_type = request.query_params.get("type")
        data = []

        if spare_type == "mo":
            items = Item.objects.all()
            for item in items:
                data.append(
                    {
                        "item_code": str(item.item_code) if item.item_code else "",
                        "item_desc": str(item.item_desc) if item.item_desc else "",
                        "status": (
                            str(item.incat_yn)
                            if hasattr(item, "incat_yn") and item.incat_yn
                            else ""
                        ),
                        "crp_category": (
                            str(item.crp_category) if item.crp_category else ""
                        ),
                        "denomination": str(item.item_deno) if item.item_deno else "",
                        "price": "",
                        "price_date": "",
                        "ilms_eqpt_code": str(item.item_code) if item.item_code else "",
                        "ilms_eqpt_desc": "NA",
                        "vendor_name": "-",
                        "pk": item.pk,
                    }
                )

        elif spare_type == "mo_mapped":
            items = Item.objects.filter(
                momappingtable__ilms_spare_id__isnull=False
            ).distinct()
            for item in items:
                data.append(
                    {
                        "item_code": str(item.item_code) if item.item_code else "",
                        "item_desc": str(item.item_desc) if item.item_desc else "",
                        "status": (
                            str(item.incat_yn)
                            if hasattr(item, "incat_yn") and item.incat_yn
                            else ""
                        ),
                        "crp_category": (
                            str(item.crp_category) if item.crp_category else ""
                        ),
                        "denomination": str(item.item_deno) if item.item_deno else "",
                        "price": "",
                        "price_date": "",
                        "ilms_eqpt_code": str(item.item_code) if item.item_code else "",
                        "ilms_eqpt_desc": "NA",
                        "vendor_name": "-",
                        "pk": item.pk,
                    }
                )

        elif spare_type == "wed":
            from wlms.models import WLMSSpare

            wed_spares = WLMSSpare.objects.select_related("eqpt").all()
            for spare in wed_spares:
                eqpt_name = ""
                try:
                    eqpt_name = str(spare.eqpt.eqpt_name) if spare.eqpt else ""
                except Exception:
                    pass
                data.append(
                    {
                        "eqpt": eqpt_name,
                        "item_code": str(spare.item_code) if spare.item_code else "",
                        "item_desc": str(spare.item_desc) if spare.item_desc else "",
                        "denomination": str(spare.denom_id) if spare.denom_id else "",
                        "category": str(spare.category) if spare.category else "",
                        "typeofspare": (
                            str(spare.typeofspare) if spare.typeofspare else ""
                        ),
                        "avail_status": (
                            "Yes" if getattr(spare, "is_active", False) else "No"
                        ),
                        "pk": spare.pk,
                    }
                )

        elif spare_type == "wed_mapped":
            from django.db.models import Prefetch

            from wlms.models import SpareDataMap, WLMSSpare

            wed_spares = WLMSSpare.objects.prefetch_related(
                Prefetch(
                    "eqpt__spare_data_maps",
                    queryset=SpareDataMap.objects.select_related("equipment"),
                    to_attr="mapped_equipments",
                )
            ).select_related("eqpt")
            for spare in wed_spares:
                eqpt_name = ""
                insma_name = ""
                try:
                    eqpt_name = str(spare.eqpt.eqpt_name) if spare.eqpt else ""
                    mapped = spare.eqpt.mapped_equipments if spare.eqpt else []
                    if mapped:
                        insma_name = str(mapped[0].equipment.nomenclature)
                except Exception:
                    pass
                data.append(
                    {
                        "item_code": str(spare.item_code) if spare.item_code else "",
                        "item_desc": str(spare.item_desc) if spare.item_desc else "",
                        "denomination": str(spare.denom_id) if spare.denom_id else "",
                        "category": str(spare.category) if spare.category else "",
                        "typeofspare": (
                            str(spare.typeofspare) if spare.typeofspare else ""
                        ),
                        "avail_status": (
                            "Yes" if getattr(spare, "is_active", False) else "No"
                        ),
                        "wed_eqpt_name": eqpt_name,
                        "sfd_equipment": insma_name,
                        "pk": spare.pk,
                    }
                )

        elif spare_type == "obs":
            from obs.models import Spares

            spares = Spares.objects.all().select_related(
                "equipment_class", "denomination", "authority"
            )
            for spare in spares:
                data.append(
                    {
                        "equipmentClass": (
                            str(spare.equipment_class.name)
                            if spare.equipment_class
                            else ""
                        ),
                        "patternNo": (
                            str(spare.pattern_number) if spare.pattern_number else ""
                        ),
                        "itemDesc": str(spare.description) if spare.description else "",
                        "availabilityStatus": (
                            "Yes" if spare.quantity_available > 0 else "No"
                        ),
                        "denomination": (
                            str(spare.denomination.name) if spare.denomination else ""
                        ),
                        "heldQty": spare.quantity_available,
                        "crpCategory": str(spare.category) if spare.category else "",
                        "authority": (
                            str(spare.authority.name) if spare.authority else ""
                        ),
                        "pk": spare.pk,
                    }
                )

        serializer = GetDartSparesDataResponseSerializer(
            {"success": True, "data": data}
        )
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class DartHistoryDetailAPIView(APIView):
    """
    Get detailed history of a specific Dart including closure + spares used + RA/DL entries.
    Replaces: dart_history_detail in old views.
    """

    @extend_schema(responses={200: DartHistoryDetailResponseSerializer})
    def get(self, request, id, *args, **kwargs):
        from .models import CompleteDefectDart

        try:
            defect = InitiateDart.objects.get(id=id)
        except InitiateDart.DoesNotExist:
            return Response(
                {"error": "Dart not found"}, status=status.HTTP_404_NOT_FOUND
            )

        closure = CompleteDefectDart.objects.filter(dart_details=defect).first()
        spares_used = []
        if closure and hasattr(closure, "spares_used"):
            for s in closure.spares_used.all():
                spares_used.append(
                    {
                        "pattern_no": getattr(s, "pattern_no", ""),
                        "description": getattr(s, "description", ""),
                        "quantity": getattr(s, "quantity", 0),
                    }
                )

        ra_dl_entries = []
        if hasattr(defect, "ra_dl_entries"):
            for entry in defect.ra_dl_entries.all():
                ra_dl_entries.append(
                    {
                        "id": entry.id,
                        "dl_type": entry.dl_type,
                        "dl_key": getattr(entry, "dl_key", ""),
                        "status": entry.status,
                        "remarks": getattr(entry, "remarks", ""),
                    }
                )

        closure_data = None
        if closure:
            closure_data = {
                "id": closure.id,
                "serial_no": getattr(closure, "serial_no", ""),
                "rectified_date": (
                    str(closure.rectified_date) if closure.rectified_date else ""
                ),
                "days_delay": getattr(closure, "days_delay", 0),
                "lesson_learnt": getattr(closure, "lesson_learnt", ""),
                "other_reasons": getattr(closure, "other_reasons", ""),
            }

        return Response(
            {
                "defect": {
                    "id": defect.id,
                    "dart_number": defect.dart_number or "",
                    "dart_date": (
                        defect.dart_date.strftime(DATE_FORMAT)
                        if defect.dart_date
                        else ""
                    ),
                    "defective_discriptions": defect.defective_discriptions or "",
                    "maintenance_period": defect.maintenance_period or "",
                    "dart_occasion": defect.dart_occasion or "",
                    "is_closed": defect.is_closed,
                },
                "closure": closure_data,
                "spares_used": spares_used,
                "ra_dl_entries": ra_dl_entries,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["DART"])
class MaintenanceOverviewAPIView(APIView):
    """
    Returns the maintenance overview metrics for the dashboard for 6M, 1Y, and 2Y.
    Strictly queries database models dynamically. No fake fallback data is returned.
    """

    serializer_class = AllPeriodsMaintenanceOverviewSerializer

    @extend_schema(
        responses={200: AllPeriodsMaintenanceOverviewSerializer},
    )
    def get(self, request, *args, **kwargs):
        cache_key = make_cache_key("dart:maintenance-overview", request)
        cached_payload = cache.get(cache_key)
        if cached_payload is not None:
            serializer = AllPeriodsMaintenanceOverviewSerializer(cached_payload)
            return Response(serializer.data, status=status.HTTP_200_OK)

        periods = ["6M", "1Y", "2Y"]
        today = timezone.now().date()
        response_data = {}

        # ── 1. Operational Readiness (Calculated globally based on current state) ──
        try:
            total_equipment = EquipmentName.objects.count()
            if total_equipment == 0:
                total_equipment = ShipEquipment.objects.count()
        except Exception:
            total_equipment = 0

        try:
            non_ops_eq_count = (
                InitiateDart.objects.open().non_ops().distinct_equipment_ship_count()
            )
        except Exception:
            non_ops_eq_count = 0

        if total_equipment > 0:
            operational_readiness_val = int(
                ((total_equipment - non_ops_eq_count) / total_equipment) * 100
            )
            operational_readiness_val = max(0, min(100, operational_readiness_val))
        else:
            operational_readiness_val = 100  # 100% ready if no equipment exists

        operational_readiness_status = get_readiness_status(operational_readiness_val)

        for period in periods:
            if period == "1Y":
                delta = relativedelta(years=1)
            elif period == "2Y":
                delta = relativedelta(years=2)
            else:  # 6M
                delta = relativedelta(months=6)

            current_start = today - delta
            prev_start = current_start - delta

            # ── 2. Critical Defects ──
            crit_qs = InitiateDart.objects.open().non_ops()
            current_critical = crit_qs.filter(dart_date__gte=current_start).count()
            prev_critical = crit_qs.filter(
                dart_date__range=(prev_start, current_start)
            ).count()

            critical_change = calculate_percentage_change(
                current_critical, prev_critical
            )

            # ── 3. Overdue MAINTOPS ──
            current_overdue = 0
            prev_overdue = 0
            try:
                from ems.models import RoutineDescription

                overdue_qs = RoutineDescription.objects.overdue(today)
                current_overdue = overdue_qs.filter(due_date__gte=current_start).count()
                prev_overdue = overdue_qs.filter(
                    due_date__range=(prev_start, current_start)
                ).count()
            except Exception:
                pass

            overdue_change = calculate_percentage_change(current_overdue, prev_overdue)

            # ── 4. Equipment Under Maintenance ──
            maint_qs = InitiateDart.objects.open()
            current_maint = maint_qs.filter(
                dart_date__gte=current_start
            ).distinct_equipment_ship_count()
            prev_maint = maint_qs.filter(
                dart_date__range=(prev_start, current_start)
            ).distinct_equipment_ship_count()

            maint_change = calculate_percentage_change(current_maint, prev_maint)

            # ── 5. Repeated Failures ──
            rep_qs = InitiateDart.objects.filter(dart_date__gte=current_start)
            current_repeated = rep_qs.repeated_failures_count()

            prev_rep_qs = InitiateDart.objects.filter(
                dart_date__range=(prev_start, current_start)
            )
            prev_repeated = prev_rep_qs.repeated_failures_count()

            repeated_change = calculate_percentage_change(
                current_repeated, prev_repeated
            )

            response_data[period] = {
                "operational_readiness": {
                    "value": operational_readiness_val,
                    "status": operational_readiness_status,
                    "change_percentage": None,
                },
                "critical_defects": {
                    "value": current_critical,
                    "status": None,
                    "change_percentage": critical_change,
                },
                "overdue_maintops": {
                    "value": current_overdue,
                    "status": None,
                    "change_percentage": overdue_change,
                },
                "equipment_under_maintenance": {
                    "value": current_maint,
                    "status": None,
                    "change_percentage": maint_change,
                },
                "repeated_failures": {
                    "value": current_repeated,
                    "status": None,
                    "change_percentage": repeated_change,
                },
            }

        cache.set(cache_key, response_data, timeout=300)
        serializer = AllPeriodsMaintenanceOverviewSerializer(response_data)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class MaintenanceOverviewKPIAPIView(APIView):
    serializer_class = MaintenanceOverviewKPIResponseSerializer

    KPI_METADATA = [
        ("operational_readiness", "Operational Readiness", "percent"),
        ("critical_defects", "Critical Defects", "count"),
        ("overdue_maintops", "Overdue MAINTOPS", "count"),
        ("equipment_under_maintenance", "Equipment Under Maintenance", "count"),
        ("repeated_failures", "Repeated Failures", "count"),
    ]

    def _normalize_period(self, period):
        period = str(period or "6m").strip().lower()
        return period if period in {"6m", "1y", "2y"} else "6m"

    @extend_schema(
        parameters=[
            OpenApiParameter(
                name="period",
                type=str,
                location=OpenApiParameter.QUERY,
                required=False,
                description="KPI trend period: 6m, 1y, or 2y. Defaults to 6m.",
            )
        ],
        responses={200: MaintenanceOverviewKPIResponseSerializer},
    )
    def get(self, request, *args, **kwargs):
        selected_period = self._normalize_period(request.query_params.get("period"))
        cache_key = make_cache_key(
            "dart:maintenance-overview-kpis",
            request,
            extra={"period": selected_period},
            vary_on_user=True,
        )
        cached_payload = cache.get(cache_key)
        if cached_payload is not None:
            serializer = self.serializer_class(cached_payload)
            return Response(serializer.data, status=status.HTTP_200_OK)
        period_map = {
            "6m": "6M",
            "1y": "1Y",
            "2y": "2Y",
        }
        overview_response = MaintenanceOverviewAPIView().get(request, *args, **kwargs)
        overview_data = dict(overview_response.data)

        kpis = []
        for key, title, unit in self.KPI_METADATA:
            grouped_periods = {}
            for period_key, source_key in period_map.items():
                period_payload = overview_data.get(source_key, {}).get(key, {})
                change_percentage = period_payload.get("change_percentage")
                if change_percentage is None:
                    change_percentage = 0
                direction = "flat"
                if change_percentage > 0:
                    direction = "up"
                elif change_percentage < 0:
                    direction = "down"

                grouped_periods[period_key] = {
                    "value": period_payload.get("value", 0),
                    "status": period_payload.get("status"),
                    "trend": {
                        "period": period_key,
                        "percentage": change_percentage,
                        "direction": direction,
                    },
                }

            kpi_item = {
                "key": key,
                "title": title,
                "unit": unit,
                "6m": grouped_periods["6m"],
                "1y": grouped_periods["1y"],
                "2y": grouped_periods["2y"],
            }
            kpi_item.update(grouped_periods[selected_period])
            kpis.append(kpi_item)

        payload = {
            "kpis": kpis,
            "periods": ["6m", "1y", "2y"],
            "default_period": selected_period,
        }
        cache.set(cache_key, payload, timeout=300)
        serializer = self.serializer_class(payload)
        return Response(serializer.data, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class MaintenanceOverviewDetailsAPIView(APIView):
    """
    Returns detailed maintenance overview widget data for the dashboard.
    Strictly queries database models dynamically.
    """

    serializer_class = MaintenanceOverviewDetailsSerializer

    @extend_schema(
        responses={200: MaintenanceOverviewDetailsSerializer},
    )
    def get(self, request, *args, **kwargs):
        cache_key = make_cache_key("dart:maintenance-overview-details", request)
        cached_payload = cache.get(cache_key)
        if cached_payload is not None:
            serializer = MaintenanceOverviewDetailsSerializer(cached_payload)
            return Response(serializer.data, status=status.HTTP_200_OK)

        today = timezone.now().date()

        def month_starts():
            current_month = today.replace(day=1)
            return [
                current_month - relativedelta(months=offset)
                for offset in range(5, -1, -1)
            ]

        def month_count(queryset, date_field):
            starts = month_starts()
            index_map = {value: index for index, value in enumerate(starts)}
            trend = [0 for _ in starts]
            for value in queryset.values_list(date_field, flat=True):
                if not value:
                    continue
                month_value = value.date() if hasattr(value, "date") else value
                month_key = month_value.replace(day=1)
                if month_key in index_map:
                    trend[index_map[month_key]] += 1
            return trend

        def month_sum(queryset, date_field, value_field):
            starts = month_starts()
            index_map = {value: index for index, value in enumerate(starts)}
            trend = [0 for _ in starts]
            for date_value, metric_value in queryset.values_list(
                date_field, value_field
            ):
                if not date_value:
                    continue
                month_value = (
                    date_value.date() if hasattr(date_value, "date") else date_value
                )
                month_key = month_value.replace(day=1)
                if month_key in index_map:
                    trend[index_map[month_key]] += metric_value or 0
            return trend

        def month_avg(queryset, date_field, value_field):
            starts = month_starts()
            index_map = {value: index for index, value in enumerate(starts)}
            totals = [0 for _ in starts]
            counts = [0 for _ in starts]
            for date_value, metric_value in queryset.values_list(
                date_field, value_field
            ):
                if not date_value or metric_value is None:
                    continue
                month_value = (
                    date_value.date() if hasattr(date_value, "date") else date_value
                )
                month_key = month_value.replace(day=1)
                if month_key in index_map:
                    index = index_map[month_key]
                    totals[index] += metric_value
                    counts[index] += 1
            return [
                int(totals[index] / counts[index]) if counts[index] else 0
                for index in range(len(starts))
            ]

        def latest_total_trend(total):
            trend = [0 for _ in month_starts()]
            trend[-1] = total
            return trend

        # ── 1. Configuration Validation ──
        # ── 2. Maintenance Prioritisation ──
        prioritisation = []
        try:
            open_defects = InitiateDart.objects.open().with_priority_relations()
            for idx, defect in enumerate(open_defects, 1):
                eq_name = (
                    defect.equipment_ems.name
                    if defect.equipment_ems
                    else (
                        defect.equipment_ship.nomenclature
                        if defect.equipment_ship
                        else ""
                    )
                )
                severity_name = (
                    defect.severity_code.severity_name
                    if defect.severity_code
                    else "Medium"
                )

                # Determine score and priority rank based on severity
                if "Non-Ops" in severity_name or "Critical" in severity_name:
                    priority = "Critical"
                    score = 94 - idx
                elif "High" in severity_name:
                    priority = "High"
                    score = 86 - idx
                else:
                    priority = "Medium"
                    score = 70 - idx

                prioritisation.append(
                    {
                        "rank": idx,
                        "equipment": eq_name,
                        "defect": defect.defective_discriptions
                        or "No description provided",
                        "priority": priority,
                        "score": max(10, score),
                    }
                )
        except Exception:
            pass

        # ── 3. Equipment Health & Condition Monitoring ──
        health_data = {}
        try:
            for equipment in EquipmentName.objects.order_by("id"):
                fault_count = (
                    InitiateDart.objects.open().for_equipment_ems(equipment).count()
                )
                health_score = 0 if fault_count else 100
                section = getattr(equipment, "section", None)
                sub_department = getattr(equipment, "sub_department", None)
                sfd_equipment = getattr(equipment, "sfd_equipment", None)
                health_data[f"ems-{equipment.id}"] = {
                    "label": equipment.name,
                    "health_index": health_score,
                    "alarms": fault_count,
                    "status_text": "Open defect exists" if fault_count else "Normal",
                    "system": getattr(section, "name", "")
                    or getattr(sub_department, "name", ""),
                    "location": equipment.started_at_location
                    or getattr(sfd_equipment, "location_on_board", ""),
                    "image": "",
                    "charts": {
                        "defects": month_count(
                            InitiateDart.objects.for_equipment_ems(equipment),
                            "dart_date",
                        ),
                        "maintenance_history": month_count(
                            CompleteDefectDart.objects.for_equipment_ems(equipment),
                            "rectified_date",
                        ),
                    },
                }
        except Exception:
            health_data = {}

        try:
            for equipment in ShipEquipment.objects.order_by("id"):
                fault_count = (
                    InitiateDart.objects.open().for_equipment_ship(equipment).count()
                )
                health_score = 0 if fault_count else 100
                equipment_name = getattr(equipment, "nomenclature", None) or str(
                    equipment
                )
                equipment_category = getattr(equipment, "equipment_category", None)
                health_data[f"ship-{equipment.id}"] = {
                    "label": equipment_name,
                    "health_index": health_score,
                    "alarms": fault_count,
                    "status_text": "Open defect exists" if fault_count else "Normal",
                    "system": getattr(equipment_category, "name", ""),
                    "location": equipment.location_on_board
                    or equipment.compartment
                    or "",
                    "image": "",
                    "charts": {
                        "defects": month_count(
                            InitiateDart.objects.for_equipment_ship(equipment),
                            "dart_date",
                        ),
                        "maintenance_history": month_count(
                            CompleteDefectDart.objects.for_equipment_ship(equipment),
                            "rectified_date",
                        ),
                    },
                }
        except Exception:
            pass

        # ── 4. Maintenance History & Learning Layer ──
        maint_history = []
        try:
            closed_defects = CompleteDefectDart.objects.with_dart_details().ordered_by_rectified_date()
            for cd in closed_defects:
                maint_history.append(
                    {
                        "date": (
                            cd.rectified_date.isoformat()
                            if cd.rectified_date
                            else today.isoformat()
                        ),
                        "title": "Defect Rectified",
                        "description": cd.dart_details.defective_discriptions
                        or "Defect closed successfully",
                        "type": "history",
                    }
                )
        except Exception:
            pass

        try:
            routines = CompletedRoutine.objects.ordered_by_completion()
            for r in routines:
                maint_history.append(
                    {
                        "date": (
                            r.date_of_completion.isoformat()
                            if r.date_of_completion
                            else today.isoformat()
                        ),
                        "title": "Routine Completed",
                        "description": r.completion_details or "Routine task performed",
                        "type": "history",
                    }
                )
        except Exception:
            pass

        # ── 5. Preventive & Predictive Maintenance ──
        upcoming_maintops = []
        try:
            from ems.models import RoutineDescription

            routines_qs = (
                RoutineDescription.objects.open_items()
                .select_related("equipment_name", "department_f_key")
                .order_by("due_date")
            )
            for r in routines_qs:
                days_left = (r.due_date - today).days if r.due_date else 0
                if days_left < 0:
                    status_str = f"Overdue {abs(days_left)} Days"
                else:
                    status_str = f"Due in {days_left} Days"

                upcoming_maintops.append(
                    {
                        "equipment": (
                            r.equipment_name.name if r.equipment_name else ""
                        ),
                        "from_dept": (
                            r.department_f_key.name if r.department_f_key else ""
                        ),
                        "status": status_str,
                        "days_left": days_left,
                    }
                )
        except Exception:
            pass

        # ── 6. Maintenance Constraints & Dependencies ──
        try:
            spare_constraints_count = (
                InitiateDart.objects.open().spares_required().count()
            )
        except Exception:
            spare_constraints_count = 0

        try:
            trials_pending_count = InitiateDart.objects.open().trial_required().count()
        except Exception:
            trials_pending_count = 0

        constraints_list = []
        try:
            spares_defects = (
                InitiateDart.objects.open()
                .spares_required()
                .select_related("equipment_ems")
            )
            for d in spares_defects:
                eq_name = d.equipment_ems.name if d.equipment_ems else ""
                constraints_list.append(
                    {
                        "type": "Spares Required",
                        "title": eq_name,
                        "description": d.defective_discriptions or "",
                        "reference_id": d.dart_number or "",
                    }
                )
        except Exception:
            pass

        _maintenance_constraints = {
            "badges": {
                "spare_constraints": spare_constraints_count,
                "fmu_oem_dep": 0,
                "trials_pending": trials_pending_count,
                "manpower_gaps": 0,
            },
            "list": constraints_list,
        }

        # ── 7. Reliability & Degradation Trends ──
        try:
            avg_mttr = CompleteDefectDart.objects.average_days_delay()
        except Exception:
            avg_mttr = 0

        try:
            repeated_failures = InitiateDart.objects.repeated_failures_count()
        except Exception:
            repeated_failures = 0

        try:
            total_routine_hours = CompletedRoutine.objects.total_hours()
            total_failures = InitiateDart.objects.count()
            mtbf = int(total_routine_hours / total_failures) if total_failures else 0
        except Exception:
            mtbf = 0

        try:
            maintenance_burden = CompletedRoutine.objects.total_manpower()
        except Exception:
            maintenance_burden = 0

        routine_hours_trend = month_sum(
            CompletedRoutine.objects.all(), "date_of_completion", "hours"
        )
        defect_trend = month_count(InitiateDart.objects.all(), "dart_date")
        mtbf_trend = [
            int(hours / failures) if failures else 0
            for hours, failures in zip(routine_hours_trend, defect_trend)
        ]
        mttr_trend = month_avg(
            CompleteDefectDart.objects.all(), "rectified_date", "days_delay"
        )
        maintenance_burden_trend = month_sum(
            CompletedRoutine.objects.all(), "date_of_completion", "total_manpower"
        )

        _reliability_degradation_trends = {
            "mtbf": {"value": mtbf, "trend": mtbf_trend},
            "mttr": {
                "value": int(avg_mttr) if avg_mttr else 0,
                "trend": mttr_trend,
            },
            "failure_recurrence": {
                "value": repeated_failures,
                "trend": defect_trend,
            },
            "maint_burden": {
                "value": maintenance_burden,
                "trend": maintenance_burden_trend,
            },
        }

        # ── 8. Trials, Validation & Post-Maintenance ──
        restored_count = 0
        validation_pending = 0
        failed_count = 0
        trials_list = []

        try:
            trial_defects = InitiateDart.objects.trial_required().select_related(
                "equipment_ems"
            )
            for d in trial_defects:
                eq_name = d.equipment_ems.name if d.equipment_ems else ""
                if d.is_closed:
                    restored_count += 1
                    status_lbl = "RESTORED"
                else:
                    validation_pending += 1
                    status_lbl = "PENDING"

                trials_list.append(
                    {
                        "equipment": eq_name,
                        "trial_type": d.dart_occasion or "",
                        "date": (
                            d.rectification_date.strftime("%d %b")
                            if d.rectification_date
                            else ""
                        ),
                        "status": status_lbl,
                    }
                )
        except Exception:
            pass

        _trials_validation = {
            "summary": {
                "restored": restored_count,
                "validation_pending": validation_pending,
                "failed": failed_count,
            },
            "list": trials_list,
        }

        prioritisation_items = [
            {
                "icon": "",
                "title": item["equipment"],
                "sub_title": item["defect"],
                "priority": item["priority"],
                "percentage": item["score"],
            }
            for item in prioritisation
        ]

        equipment_summary = []
        for equipment_id, values in health_data.items():
            score = values["health_index"]
            equipment_summary.append(
                {
                    "equipment_id": equipment_id,
                    "equipment_name": values["label"],
                    "score": score,
                    "status": get_health_status(score),
                    "alarm_count": values["alarms"],
                    "message": values["status_text"],
                }
            )

        selected_equipment = equipment_summary[0] if equipment_summary else None
        selected_health_detail = {}
        if selected_equipment:
            selected_health_detail = health_data.get(
                selected_equipment["equipment_id"], {}
            )
        selected_equipment_data = {
            "equipment_name": (
                selected_equipment["equipment_name"] if selected_equipment else ""
            ),
            "system": selected_health_detail.get("system", ""),
            "location": selected_health_detail.get("location", ""),
            "image": selected_health_detail.get("image", ""),
            "condition_parameters": [
                {
                    "parameter": name.replace("_", " ").title(),
                    "current_value": sum(values),
                    "unit": "Qty",
                    "trend_data": values,
                }
                for name, values in selected_health_detail.get("charts", {}).items()
            ],
        }

        movement_history = [
            {
                "date": item["date"],
                "event_type": item["title"],
                "title": item["title"],
                "description": item["description"],
            }
            for item in maint_history
        ]

        constraint_items = [
            {
                "constraint_type": item["type"],
                "description": item["description"],
                "reference_id": item["reference_id"],
            }
            for item in _maintenance_constraints["list"]
        ]

        spare_records_count = DartSpare.objects.filter(is_delete=False).count()
        contextual_metrics = [
            {
                "title": "Defects",
                "value": InitiateDart.objects.count(),
                "unit": "Qty",
                "trend_data": month_count(InitiateDart.objects.all(), "dart_date"),
            },
            {
                "title": "Maintenance History",
                "value": CompleteDefectDart.objects.count(),
                "unit": "Qty",
                "trend_data": month_count(
                    CompleteDefectDart.objects.all(), "rectified_date"
                ),
            },
            {
                "title": "Spare Records",
                "value": spare_records_count,
                "unit": "Qty",
                "trend_data": latest_total_trend(spare_records_count),
            },
            {
                "title": "Trial Records",
                "value": InitiateDart.objects.trial_required().count(),
                "unit": "Qty",
                "trend_data": month_count(
                    InitiateDart.objects.trial_required(), "dart_date"
                ),
            },
        ]

        reliability_metrics = [
            {
                "metric_name": "MTBF",
                "value": _reliability_degradation_trends["mtbf"]["value"],
                "unit": "Hrs",
                "trend_data": _reliability_degradation_trends["mtbf"]["trend"],
            },
            {
                "metric_name": "MTTR",
                "value": _reliability_degradation_trends["mttr"]["value"],
                "unit": "Hrs",
                "trend_data": _reliability_degradation_trends["mttr"]["trend"],
            },
            {
                "metric_name": "Failure Recurrence",
                "value": _reliability_degradation_trends["failure_recurrence"]["value"],
                "unit": "Qty",
                "trend_data": _reliability_degradation_trends["failure_recurrence"][
                    "trend"
                ],
            },
            {
                "metric_name": "Maintenance Burden",
                "value": _reliability_degradation_trends["maint_burden"]["value"],
                "unit": "Mh",
                "trend_data": _reliability_degradation_trends["maint_burden"]["trend"],
            },
        ]

        trial_activities = [
            {
                "equipment_name": item["equipment"],
                "activity": item["trial_type"],
                "date": item["date"],
                "status": item["status"],
            }
            for item in _trials_validation["list"]
        ]

        try:
            spare_model = apps.get_model("obs", "Spares")
            critical_spares_unavailable = spare_model.objects.filter(
                critical=True,
                quantity_available__lte=0,
            ).count()
        except LookupError:
            critical_spares_unavailable = 0

        command_actions = [
            {
                "title": "New Defects",
                "count": InitiateDart.objects.open().count(),
                "severity": "high",
            },
            {
                "title": "New Maintenance",
                "count": len(upcoming_maintops),
                "severity": "medium",
            },
            {
                "title": "Critical Spares Unavailable",
                "count": critical_spares_unavailable,
                "severity": "critical",
            },
            {
                "title": "Delayed Procurement",
                "count": 0,
                "severity": "high",
            },
        ]

        payload = {
            "command_actions_pending": {
                "heading": "COMMAND ACTIONS PENDING",
                "view_all_flag": True,
                "items": command_actions,
            },
            "maintenance_prioritisation": {
                "heading": "Maintenance Prioritisation",
                "view_all_flag": bool(prioritisation_items),
                "maintenance_prioritisation_items": prioritisation_items,
            },
            "equipment_health_monitoring": {
                "heading": "Equipment Health & Condition Monitoring",
                "view_all_flag": bool(equipment_summary),
                "search_enabled": True,
                "equipment_summary": equipment_summary,
                "selected_equipment": selected_equipment_data,
            },
            "movement_and_configuration_history": {
                "heading": "Movement & Configuration History",
                "view_all_flag": bool(movement_history),
                "movement_and_configuration_histories": movement_history,
            },
            "maintenance_constraints_and_dependencies": {
                "heading": "Maintenance Constraints & Dependencies",
                "view_all_flag": bool(constraint_items),
                "summary": {
                    "spares_constraints": _maintenance_constraints["badges"][
                        "spare_constraints"
                    ],
                    "fmu_oem_dependencies": _maintenance_constraints["badges"][
                        "fmu_oem_dep"
                    ],
                    "trials_pending": _maintenance_constraints["badges"][
                        "trials_pending"
                    ],
                    "manpower_gaps": _maintenance_constraints["badges"][
                        "manpower_gaps"
                    ],
                },
                "constraints": constraint_items,
            },
            "contextual_search_and_drill_down": {
                "heading": "Contextual Search & Drill-Down",
                "view_all_flag": True,
                "search_enabled": True,
                "metrics": contextual_metrics,
            },
            "reliability_and_degradation_trend": {
                "heading": "Reliability and Degradation Trend",
                "view_all_flag": True,
                "search_enabled": True,
                "metrics": reliability_metrics,
            },
            "trials_validation_and_post_maintenance": {
                "heading": "Trials, Validation & Post-Maintenance",
                "view_all_flag": bool(trial_activities),
                "summary": _trials_validation["summary"],
                "activities": trial_activities,
            },
        }

        cache.set(cache_key, payload, timeout=300)
        serializer = MaintenanceOverviewDetailsSerializer(payload)
        return Response(serializer.data, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────
# CMMS / CANDEF / OPDEF Integration Views
# ─────────────────────────────────────────────────────────────


class DummySweetify:
    def info(self, request, title="", icon="info", text="", persistent=False):
        pass

    def success(self, request, title="", icon="success", text="", persistent=False):
        pass

    def error(self, request, title="", icon="error", text="", persistent=False):
        pass


sweetify = DummySweetify()


def get_this_ship():
    from master.models import Ship

    ship_id = os.getenv("DEFAULT_SHIP_ID", "12")
    ship_obj = Ship.objects.filter(universal_id_m_ship=ship_id).first()
    if not ship_obj:
        ship_obj = Ship.objects.first()
    return ship_obj


def get_mssql_connection():
    class DummyCursor:
        def execute(self, query, params=None):
            pass

        def fetchall(self):
            return []

        def close(self):
            pass

    class DummyConnection:
        def cursor(self):
            return DummyCursor()

        def commit(self):
            pass

        def rollback(self):
            pass

        def close(self):
            pass

    return DummyConnection()


@extend_schema(tags=["DART"])
class DartPayloadView(APIView):
    """CMMS DART integration.
    GET  → Returns pending defects data (delegates to dart/pending_defect).
    POST → Returns the DART sync payload.
    """

    def get(self, request):
        """Return pending defects data from PendingDefectAPIView."""
        pending_view = PendingDefectAPIView()
        pending_view.request = request
        pending_view.format_kwarg = None
        pending_view.kwargs = {}
        return pending_view.get(request)

    def post(self, request):
        """Return same pending defects data as GET."""
        return self.get(request)


@extend_schema(tags=["DART"])
class CompletedDartView(APIView):
    """CMMS Completed DART integration (renamed from DefectsListView).
    GET  → Returns completed/closed DART history data (delegates to dart/history/data).
    POST → Accepts completed DART data in the same format as GET response.
    """

    def get(self, request):
        """Return completed DART history data from DartHistoryDefaultDataAPIView."""
        history_view = DartHistoryDefaultDataAPIView()
        history_view.request = request
        history_view.format_kwarg = None
        history_view.kwargs = {}
        return history_view.get(request)

    def post(self, request):
        """Accept completed DART data (same format as GET response) and acknowledge receipt."""
        data = request.data
        t_dart_records = data.get("T_DART", [])

        results = []
        for record in t_dart_records:
            defect_data = record.get("defect", {})
            closure_data = record.get("closure", None)
            spares_used = record.get("spares_used", [])
            ra_dl_entries = record.get("ra_dl_entries", [])

            results.append(
                {
                    "defect": defect_data,
                    "closure": closure_data,
                    "spares_used": spares_used,
                    "ra_dl_entries": ra_dl_entries,
                }
            )

        return Response({"T_DART": results}, status=status.HTTP_200_OK)


@extend_schema(tags=["DART"])
class DefectDetailView(APIView):
    """Return a validation-only defect representation for the requested ID."""

    def get(self, request, defect_id):
        resp_serializer = DefectResponseSerializer(data={"id": defect_id})
        resp_serializer.is_valid()
        return Response(resp_serializer.data)


@extend_schema(tags=["DART"])
class DefectRectifyView(APIView):
    """Validate a defect rectification payload and acknowledge receipt."""

    def post(self, request, defect_id):
        serializer = DefectRectifyRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        resp_serializer = GenericSuccessResponseSerializer(
            data={
                "success": True,
                "message": "Defect rectification payload validated successfully.",
                "data": {
                    "defect_id": defect_id,
                    "rectified_date": serializer.validated_data.get("rectified_date"),
                },
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data)


@extend_schema(tags=["DART"])
class CompletedRoutinesView(APIView):
    """Return an empty completed routines list."""

    def get(self, request):
        return Response([])


@extend_schema(tags=["DART"])
class CompleteRoutineView(APIView):
    """Validate a completed maintenance routine payload and acknowledge receipt."""

    def post(self, request):
        serializer = CompletedRoutineCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        from dart.tasks import validate_routine_completion_task

        dispatch = dispatch_task(
            validate_routine_completion_task, serializer.validated_data
        )
        if dispatch["queued"]:
            resp_serializer = GenericSuccessResponseSerializer(
                data={
                    "success": True,
                    "message": "Routine completion payload validation dispatched to Celery.",
                    "data": {"task_id": dispatch["task"].id},
                }
            )
            resp_serializer.is_valid()
            return Response(resp_serializer.data, status=status.HTTP_202_ACCEPTED)
        result = dispatch["result"]
        return Response(
            result,
            status=result.get("status_code", status.HTTP_201_CREATED),
        )


@extend_schema(tags=["DART"])
class OpdefView(APIView):
    """OPDEF integration - GET for sync payload, POST for initiation validation."""

    def get(self, request):
        serializer = OpdefSyncPayloadResponseSerializer(data={})
        serializer.is_valid()
        return Response(serializer.data)

    def post(self, request):
        serializer = OpdefInitiateRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        val = serializer.validated_data
        resp_data = {
            "OpdefMainID": 0,
            "Universal_ID_T_OpdefMain": f"U-OPD-{val.get('ship_id')}-{val.get('fitted_equipment_id')}-"
            + f"{val.get('opdef_number')}",
        }
        resp_serializer = OpdefInitiateResponseSerializer(data=resp_data)
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["DART"])
class OpdefAnalysisView(APIView):
    """Validate an OPDEF analysis payload and acknowledge receipt."""

    def post(self, request, opdef_id):
        serializer = OpdefAnalysisRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        resp_serializer = GenericSuccessResponseSerializer(
            data={
                "success": True,
                "message": "Defect analysis payload validated successfully.",
                "data": {
                    "opdef_id": opdef_id,
                    "analysis_date": serializer.validated_data.get("analysis_date"),
                },
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["DART"])
class OpdefSparesView(APIView):
    """Validate a spare consumption payload and acknowledge receipt."""

    def post(self, request, opdef_id):
        serializer = OpdefSpareRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        resp_serializer = GenericSuccessResponseSerializer(
            data={
                "success": True,
                "message": "Spare consumption payload validated successfully.",
                "data": {
                    "opdef_id": opdef_id,
                    "spare_item_code": serializer.validated_data.get("spare_item_code"),
                },
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["DART"])
class OpdefTrialsView(APIView):
    """Validate trial conducted parameters and acknowledge receipt."""

    def post(self, request, opdef_id):
        serializer = OpdefTrialRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        resp_serializer = GenericSuccessResponseSerializer(
            data={
                "success": True,
                "message": "Trial parameters payload validated successfully.",
                "data": {
                    "opdef_id": opdef_id,
                    "trial_date": serializer.validated_data.get("trial_date"),
                },
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["DART"])
class OpdefPriorParamsView(APIView):
    """Validate prior reading parameters and acknowledge receipt."""

    def post(self, request, opdef_id):
        serializer = OpdefPriorParamRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        resp_serializer = GenericSuccessResponseSerializer(
            data={
                "success": True,
                "message": "Prior parameters payload validated successfully.",
                "data": {
                    "opdef_id": opdef_id,
                    "reading_time": serializer.validated_data.get("reading_time"),
                },
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["DART"])
class OpdefPhotographsView(APIView):
    """Validate photograph upload metadata and return a derived file path."""

    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, opdef_id):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response(
                {"detail": "File is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        file_path = f"/media/opdef_photos/{file_obj.name}"
        resp_serializer = OpdefPhotoResponseSerializer(
            data={
                "success": True,
                "file_path": file_path,
                "message": f"Photograph metadata validated at {datetime.now().isoformat()}.",
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["DART"])
class InitiateCANDEFDataSaveAPI(APIView):
    @transaction.atomic
    def post(self, request):
        try:
            data = request.data
            initiate_candefs = data.get("initiate_candefs", [])
            ship_obj = get_this_ship()
            ship_sr_no = ship_obj.sr_no if ship_obj else ""
            universal_id_m_ship = ship_obj.universal_id_m_ship if ship_obj else 0

            inserted_ids = []
            success_ids = []

            for candef in initiate_candefs:
                candef_id = f"SWMM-{candef.get('id')}"
                candef_number = candef.get("candef_number")

                candef_parts = (candef_number or "").split("-")
                candef_section = candef_parts[1] if len(candef_parts) > 1 else ""
                candef_no = candef_parts[2] if len(candef_parts) > 2 else ""
                sr_no = f"{candef_section}-{candef_no}"

                obj, created = TCandef.objects.update_or_create(
                    universal_id_t_candef=candef_id,
                    defaults={
                        "candef_number": candef.get("candef_number"),
                        "sr_no": sr_no,
                        "ship_sr_no": ship_sr_no,
                        "equipment_code": candef.get("equipment_code"),
                        "department_id": candef.get("department_DepartmentID"),
                        "department": candef.get("department_Department"),
                        "ex_dept": candef.get("department_ExDept"),
                        "ex_dept_id": candef.get("department_ExDeptID"),
                        "candef_date": candef.get("candef_date"),
                        "defect_date": candef.get("candef_date"),
                        "schedule_date": candef.get("rectified_date"),
                        "universal_id_ch_master_symptoms": candef.get(
                            "Universal_ID_Ch_Master_Symptoms"
                        ),
                        "severity_id": candef.get("serverity_id"),
                        "severity_code": candef.get("serverity_code"),
                        "universal_id_m_severity": candef.get(
                            "Universal_ID_M_Severity"
                        ),
                        "universal_id_ch_master_ship_remarks_by": candef.get(
                            "Universal_ID_Ch_Master_Ship_Remarks_By"
                        ),
                        "universal_id_m_required_assistance": candef.get(
                            "Universal_ID_M_RequiredAssistance"
                        ),
                        "is_ost_observation": candef.get("trial_required"),
                        "universal_id_m_ost_list": candef.get("Universal_ID_M_OSTList"),
                        "defective_component": candef.get("defective_component"),
                        "remarks": candef.get("defective_discriptions"),
                        "defect_description": candef.get("defective_discriptions"),
                        "is_closed": 0,
                        "is_defect": 1,
                        "active": 1,
                        "nil_dart": 0,
                        "is_amp": 0,
                        "is_signal_drafted": 0,
                        "is_refit": 0,
                        "is_routine": 0,
                        "is_dl_ii_drafted": 0,
                        "is_refit_ra_draft": 0,
                        "is_gd_form": 0,
                        "universal_id_m_ship": universal_id_m_ship,
                        "universal_id_m_department": candef.get(
                            "Universal_ID_M_Department"
                        ),
                        "universal_id_t_equipment_ship_detail": candef.get(
                            "t_ship_details"
                        ),
                        "created_date": candef.get("created_date"),
                        "routine_defect": 2,
                        "is_final_submit": 1,
                        "serial_number": candef_no,
                        "is_operational": candef.get("Is_Operational"),
                        "universal_id_t_ref_comp": candef.get("Universal_ID_T_RefComp"),
                    },
                )

                if created:
                    inserted_ids.append(candef_id)
                success_ids.append(candef.get("id"))

            return Response(
                {
                    "status": True,
                    "message": "CANDEF records saved successfully",
                    "inserted_count": len(inserted_ids),
                    "data": inserted_ids,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            logger.exception("CANDEF bulk insert failed")
            return Response(
                {
                    "status": False,
                    "message": "Failed to save CANDEF data",
                    "error": str(e),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["DART"])
class CompleteCANDEFUpdateAPI(APIView):
    @transaction.atomic
    def post(self, request):
        try:
            complete_candefs = request.data.get("complete_candefs", [])

            if not complete_candefs:
                return Response(
                    {"status": False, "error": "complete_candefs is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            updated_ids = []

            for comp in complete_candefs:
                universal_id = f"SWMM-{comp['initiate_candef_id']}"

                TCandef.objects.filter(universal_id_t_candef=universal_id).update(
                    rectified_date=comp.get("rectified_date"),
                    cancel_date=comp.get("rectified_date"),
                    diagnostic_code=comp.get("DiagnosticCode"),
                    repair_code=comp.get("RepairCode"),
                    delay_reason_days=comp.get("days_delay"),
                    additional_remarks=comp.get("rectification_details"),
                    spares_availability=comp.get("spares_delay"),
                    diagnostic_id=comp.get("DiagnosticID"),
                    repair_id=comp.get("RepairID"),
                    repair_agency_id=comp.get("RepairAgencyID"),
                    agency_code=comp.get("AgencyCode"),
                    delay_id=comp.get("DelayID"),
                    delay_code=comp.get("DelayCode"),
                    universal_id_m_repair=comp.get("Universal_ID_M_Repair"),
                    universal_id_m_delay=comp.get("Universal_ID_M_Delay"),
                    universal_id_m_diagnostic=comp.get("Universal_ID_M_Diagnostic"),
                    universal_id_m_repair_agency=comp.get(
                        "Universal_ID_M_RepairAgency"
                    ),
                    is_auto_generated_dart=comp.get("Is_Auto_Generated_Dart"),
                    is_closed=1,
                )
                updated_ids.append(universal_id)

            return Response(
                {
                    "status": True,
                    "message": "CANDEF completion updated successfully",
                    "updated_count": len(updated_ids),
                    "updated_ids": updated_ids,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            logger.exception("CANDEF completion update failed")
            return Response(
                {"status": False, "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CANDEFSyncAPIView(APIView):
    task = None
    queued_message = ""
    fallback_message = ""

    def get(self, request, pk, *args, **kwargs):
        if self.task is None:
            raise NotImplementedError("task must be set on the subclass.")
        dispatch = dispatch_task(self.task, pk)
        if dispatch["queued"]:
            return accepted_task_response(
                request,
                dispatch["task"],
                self.queued_message,
            )
        return sync_task_response(dispatch["result"], self.fallback_message)


@extend_schema(tags=["DART"])
class InitiateCandefSyncAPIView(CANDEFSyncAPIView):
    task = initiate_candef_sync_task
    queued_message = "CANDEF initiation sync queued for background processing."
    fallback_message = "Background task service unavailable. CANDEF initiation sync completed synchronously."


@extend_schema(tags=["DART"])
class CompleteCandefSyncAPIView(CANDEFSyncAPIView):
    task = complete_candef_sync_task
    queued_message = "CANDEF completion sync queued for background processing."
    fallback_message = "Background task service unavailable. CANDEF completion sync completed synchronously."


@extend_schema(tags=["DART"])
class TaskStatusView(APIView):
    """Check status of a Celery background task."""

    def get(self, request, task_id):
        res = AsyncResult(task_id, app=celery_app)
        response_data = {
            "task_id": task_id,
            "status": res.status,
            "result": res.result if res.ready() else None,
        }
        return Response(response_data)
