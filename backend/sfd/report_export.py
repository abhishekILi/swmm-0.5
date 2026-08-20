"""Report export rendering — used by sfd.tasks.generate_report_export_task.

Reuses the SAME ListAPIView classes (queryset + filterset_class + serializer_class) the
on-screen report grids already hit, rather than re-deriving each report's filter/query logic
here — a fake DRF Request built from the export's filter dict drives filter_queryset() exactly
like a real GET request would, so this can never drift from what the grid shows.
"""

import html as html_lib
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.request import Request
from rest_framework.test import APIRequestFactory
from weasyprint import HTML

from .source_contract import (
    ApprovalStatusReportAPIView,
    EquipmentInstallationReportAPIView,
    EquipmentLocationReportAPIView,
    RemovedEquipmentReportAPIView,
    SFDTransactionReportAPIView,
    ShipEquipmentConfigurationReportAPIView,
)

REPORT_KEY_TO_VIEW = {
    "ship-equipment-configuration": ShipEquipmentConfigurationReportAPIView,
    "sfd-transactions": SFDTransactionReportAPIView,
    "removed-equipment": RemovedEquipmentReportAPIView,
    "sfd-installations": EquipmentInstallationReportAPIView,
    "sfd-locations": EquipmentLocationReportAPIView,
    "approval-status": ApprovalStatusReportAPIView,
}

REPORT_TITLES = {
    "ship-equipment-configuration": "Ship Equipment Configuration Report",
    "sfd-transactions": "Transaction Report",
    "removed-equipment": "Removed Equipment Report",
    "sfd-installations": "Equipment Installation Report",
    "sfd-locations": "Equipment Location Report",
    "approval-status": "Approval Status Report",
}

# (serializer field, display header) pairs, in export column order — mirrors the frontend's
# REPORTS[].columns / *_FIELD_BY_COLUMN tables (sfd-reports.data.ts) so the export matches what
# the on-screen report shows.
REPORT_COLUMNS = {
    "ship-equipment-configuration": [
        ("department", "Department"),
        ("sub_department", "Sub Department"),
        ("equipment_name", "Equipment Name"),
        ("transaction_category", "Equipment Type"),
        ("manufacture", "OEM"),
        ("equipment_sr_no", "Serial Number"),
        ("equipment_nomenclature", "Equipment Nomenclature"),
        ("deck_no", "Deck Number"),
        ("frame_station", "Frame Station"),
        ("compartment", "Compartment"),
        ("location", "Location"),
        ("equipment_model", "Model"),
        ("maintop_id", "Maintop ID"),
        ("transaction_type", "Transaction Type"),
        ("is_system", "Is System"),
        ("oem_part_no", "OEM Part No"),
        ("installation_date", "Installation Date"),
        ("supplier", "Supplier Name"),
        ("service_life", "Shelf Life"),
        ("approval_status", "Approval Status"),
    ],
    "sfd-transactions": [
        ("equipment_name", "Equipment Name"),
        ("department", "Department"),
        ("sfd_category", "SFD Category"),
        ("transaction_type", "Transaction Type"),
        ("serial_no", "Serial No"),
        ("transaction_date", "Transaction Date"),
        ("status", "Status"),
    ],
    "removed-equipment": [
        ("equipment_code", "Equipment Code"),
        ("equipment_name", "Equipment Name"),
        ("serial_no", "Serial No"),
        ("removal_date", "Removal Date"),
        ("removal_remark", "Removal Remark"),
        ("removal_authority", "Removal Authority"),
        ("installation_authority", "Installation Authority"),
        ("status", "Status"),
        ("is_sync", "Is Sync"),
    ],
    "sfd-installations": [
        ("equipment_name", "Equipment Name"),
        ("serial_no", "Serial Number"),
        ("oem", "OEM"),
        ("supplier", "Supplier"),
        ("installation_date", "Installation Date"),
        ("installation_authority", "Installation Authority"),
        ("deck_no", "Deck No"),
        ("frame_station", "Frame Station"),
        ("compartment", "Compartment"),
    ],
    "sfd-locations": [
        ("equipment_name", "Equipment Name"),
        ("equipment_code", "Equipment Code"),
        ("deck_no", "Deck No"),
        ("frame_station", "Frame Station"),
        ("location", "Location"),
        ("compartment", "Compartment Name"),
        ("qty_fitted", "Qty Fitted"),
    ],
    "approval-status": [
        ("equipment_code", "Equipment Code"),
        ("equipment_name", "Equipment Name"),
        ("serial_no", "Serial No"),
        ("approval_request_type", "Request Type"),
        ("status", "Status"),
        ("install_date", "Installation Date"),
        ("rh_at_installation", "RH at Installation"),
        ("installation_remark", "Installation Remark"),
        ("installation_authority", "Installation Authority"),
        ("removal_date", "Removal Date"),
        ("removal_remark", "Removal Remark"),
        ("removal_authority", "Removal Authority"),
        ("is_sync", "Is Sync"),
    ],
}

_request_factory = APIRequestFactory()


def fetch_report_rows(report_key: str, filters: dict) -> list[dict]:
    """Filtered, serialized rows for a report — the full (unpaginated) result set, exactly as
    the on-screen grid's underlying queryset/serializer would produce for the same filters."""
    view_class = REPORT_KEY_TO_VIEW.get(report_key)
    if not view_class:
        raise ValueError(f"Unknown report_key: {report_key}")

    view = view_class()
    django_request = _request_factory.get("/", filters or {})
    view.request = Request(django_request)
    view.format_kwarg = None
    view.kwargs = {}

    queryset = view.filter_queryset(view.get_queryset())
    serializer = view.get_serializer(queryset, many=True)
    return list(serializer.data)


def _cell_text(value) -> str:
    return "" if value is None else str(value)


def render_excel(report_key: str, title: str, rows: list[dict]) -> bytes:
    columns = REPORT_COLUMNS.get(report_key, [])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (title or "Report")[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1D96E9", end_color="1D96E9", fill_type="solid"
    )

    sheet.append([label for _, label in columns])
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        sheet.append([_cell_text(row.get(field)) for field, _ in columns])

    for col_cells in sheet.columns:
        longest = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        sheet.column_dimensions[col_cells[0].column_letter].width = min(
            max(longest + 2, 10), 40
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_pdf(report_key: str, title: str, rows: list[dict]) -> bytes:
    columns = REPORT_COLUMNS.get(report_key, [])
    header_html = "".join(f"<th>{html_lib.escape(label)}</th>" for _, label in columns)
    body_html = "".join(
        "<tr>"
        + "".join(
            f"<td>{html_lib.escape(_cell_text(row.get(field)))}</td>"
            for field, _ in columns
        )
        + "</tr>"
        for row in rows
    )
    html_string = f"""
    <html>
      <head>
        <style>
          @page {{ size: A4 landscape; margin: 1.5cm; }}
          body {{ font-family: Arial, sans-serif; color: #111; }}
          h1 {{ font-size: 16px; margin: 0 0 12px; }}
          table {{ width: 100%; border-collapse: collapse; font-size: 9px; }}
          th, td {{ border: 1px solid #ccc; padding: 4px 6px; text-align: left; }}
          th {{ background: #1D96E9; color: #fff; }}
          tr:nth-child(even) {{ background: #f7f7f7; }}
        </style>
      </head>
      <body>
        <h1>{html_lib.escape(title)}</h1>
        <table>
          <thead><tr>{header_html}</tr></thead>
          <tbody>{body_html}</tbody>
        </table>
      </body>
    </html>
    """
    return HTML(string=html_string).write_pdf()
