import math
import re
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

import numpy as np
import pdfkit
from celery.result import AsyncResult
from dart.models import CompletedRoutine, CompletedRoutineSpare
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.db.models import Count, Exists, F, IntegerField, Max, OuterRef, Q
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone as dj_timezone
from django.utils.dateparse import parse_datetime
from django.utils.timezone import is_naive, make_aware
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiTypes,
    extend_schema,
)
from ilms.models import (
    Item,
    MoMappingTable,
    MOPlannedSparesDescription,
    PlannedMOSpareList,
    Vendor,
)
from master.models import (
    Department,
    MDeferment,
    MEstablishment,
    MInability,
    MMaterialOrganizations,
    MRanklist,
    MReason,
    RefitMaintenancePeriod,
    SubDepartment,
)
from obs.models import (
    Authority,
    Denomination,
    EquipmentClass,
    Issue,
    PlannedRoutineSpareList,
    Spares,
    SparesMapping,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.exceptions import ValidationError
from rest_framework.generics import GenericAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from sfd.models import Equipment, ShipEquipment
from swmm.async_jobs import dispatch_task, sync_task_response
from swmm.celery import app as celery_app
from wlms.models import (
    PlannedSparesDescription,
    PlannedWEDSpareList,
    SpareDataMap,
    WLMSSpare,
)

from .models import (
    AddRoutineDetails,
    DataPointsAirprHPC,
    DataPointsExtTemp,
    DataPointsExtTempGTG,
    DataPointsGTLPC,
    EquipmentName,
    FussRaiseDetails,
    FussSpare,
    LessAddRoutineDetails,
    LessRoutineDescription,
    MaintopDetail,
    MaintopHeader,
    MeasurementFuelsondingFinal,
    PlannedRoutineDescription,
    PostCalculateGTG,
    PostCalculateLPC,
    PostEquipmentStateChangeHistorySave,
    PostRoutineDetails,
    RADLMaster,
    RADLRoutineDescription,
    RoutineDescription,
    SectionName,
    SlipLimit,
    UniqueRoutineName,
)
from .serializers import (
    AddRoutineDescriptionRequestSerializer,
    AddRoutineDetailsRequestSerializer,
    CalculateGtSlipRequestSerializer,
    CompleteFussRequestSerializer,
    CompleteRoutineRequestSerializer,
    CreateEquipmentNameRequestSerializer,
    CreateRoutineFrequencyRequestSerializer,
    DeleteDlDraftRowRequestSerializer,
    EditEquipmentNameRequestSerializer,
    EditRoutineNameRequestSerializer,
    EmsEquipmentCreateRequestSerializer,
    EmsMonthlyRunningHoursSaveRequestSerializer,
    EmsSectionCreateRequestSerializer,
    EmsTotalRunningHoursCreateRequestSerializer,
    EquipmentStatusResponseSerializer,
    FrontendFussRaisedDetailResponseSerializer,
    FrontendFussRaisedSearchResponseSerializer,
    FrontendPlannedRoutineDetailResponseSerializer,
    FrontendPlannedRoutineSearchResponseSerializer,
    FrontendRoutinePlanDetailResponseSerializer,
    FrontendRoutinePlanFiltersSerializer,
    FrontendRoutinePlanFormResponseSerializer,
    FrontendRoutinePlanSaveResponseSerializer,
    FrontendRoutinePlanSearchResponseSerializer,
    FussMastersResponseSerializer,
    FussRaiseRequestSerializer,
    FussSyncPayloadResponseSerializer,
    GenerateDL1RequestSerializer,
    GenerateDl1RequestSerializer,
    GenericSuccessResponseSerializer,
    GetEquipmentHistoryResponseSerializer,
    GetEquipmentNameResponseSerializer,
    GetEquipmentNameWithoutRHSINullRowsResponseSerializer,
    GetManualNamesRequestSerializer,
    GetRoutineNameResponseSerializer,
    GetSectionNameResponseSerializer,
    GetsrarEquipmentNameResponseSerializer,
    InitiateCloseRoutineResponseSerializer,
    LookupFuelSoundingRequestSerializer,
    MaintopDistributionRequestSerializer,
    MaintopJICRequestSerializer,
    MaintopSyncRequestSerializer,
    MulRaiseFussRequestSerializer,
    MulRaiseFussResponseSerializer,
    PlanRoutineMultiSaveRequestSerializer,
    PlanRoutineSaveRequestSerializer,
    RefitMaintenancePeriodSerializer,
    RoutineDescriptionSerializer,
    SaveDlDraftRowsRequestSerializer,
    SaveOemDataPointRequestSerializer,
    SaveOemSpareRequestSerializer,
    SaveRoutineInitBulkRequestSerializer,
    SaveRoutineInitializationRowRequestSerializer,
    SaveSlipLimitRequestSerializer,
    UpdateEquipmentStateRequestSerializer,
)
from .utils import (
    add_hours_to_pseudo as ems_add_hours_to_pseudo,
)
from .utils import (
    calculate_volume_weight,
    generate_routine_dart_number,
)
from .utils import (
    float_hours_to_pseudo as ems_float_to_pseudo,
)
from .utils import (
    get_due_status as get_due_status_backend,
)
from .utils import (
    pseudo_hours_to_hhmm as ems_pseudo_to_hhmm,
)

# Local Timezone
local_tz = timezone(timedelta(hours=5, minutes=30))


class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class _LazyPandas:
    def __getattr__(self, name):
        import pandas

        return getattr(pandas, name)


pd = _LazyPandas()

# --- API Endpoints ---


def _get_ems_department(request):
    if request.user.is_authenticated and hasattr(request.user, "user_profile"):
        return request.user.department
    return Department.objects.first()


def _build_due_bucket(status_color):
    if status_color == "#FF9999":
        return "due"
    if status_color == "#f7e687":
        return "due_lt_3m_500h"
    if status_color == "orange":
        return "due_3_6m_1000h"
    return "other"


def _build_due_counts(items):
    counts = {
        "due": 0,
        "due_lt_3m_500h": 0,
        "due_3_6m_1000h": 0,
        "other": 0,
        "total": len(items),
    }
    for item in items:
        counts[item["due_bucket"]] += 1
    return counts


def _serialize_frontend_sections(department):
    queryset = SectionName.objects.ordered_by_name()
    if department:
        queryset = queryset.for_department_id(department.id)
    return [{"id": item.id, "label": item.name} for item in queryset]


def _serialize_frontend_equipment(section_id=None, department=None):
    if section_id and str(section_id) != "0":
        queryset = EquipmentName.objects.for_section(section_id).ordered_by_name()
    elif department:
        queryset = EquipmentName.objects.for_department(department).ordered_by_name()
    else:
        queryset = EquipmentName.objects.ordered_by_name()
    return [{"id": item.id, "label": item.name} for item in queryset]


def _serialize_frontend_routine_types():
    return [
        {"value": choice[0], "label": choice[0]}
        for choice in AddRoutineDetails.routine_category_choices
    ]


def _serialize_frontend_routine_names(request, department=None):
    category = request.query_params.get("routine_type") or request.query_params.get(
        "category"
    )
    section_id = request.query_params.get("section_id") or request.query_params.get(
        "sectionId"
    )
    equipment_id = request.query_params.get("equipment_id") or request.query_params.get(
        "equipment_name"
    )

    routine_names = AddRoutineDetails.objects.all()
    if department:
        routine_names = routine_names.filter(
            equipment_name__section__department=department
        )
    if category and category != "0":
        routine_names = routine_names.filter(routine_category__iexact=category)
    if section_id and str(section_id) != "0":
        routine_names = routine_names.filter(equipment_name__section_id=section_id)
    if equipment_id and str(equipment_id) != "0":
        routine_names = routine_names.filter(equipment_name_id=equipment_id)

    routine_names = (
        routine_names.exclude(frequency__isnull=True)
        .exclude(frequency__exact="")
        .values_list("frequency", flat=True)
        .distinct()
        .order_by("frequency")
    )
    return [{"value": value, "label": value} for value in routine_names]


def _build_frontend_routine_plan_results(request):
    current_time = make_aware(datetime.now())
    department = _get_ems_department(request)
    department_id = department.id if department else None

    routines = AddRoutineDetails.objects.with_dashboard_relations()
    if department_id:
        routines = routines.filter(equipment_name__section__department_id=department_id)

    section_id = request.query_params.get("section_id") or request.query_params.get(
        "section"
    )
    equipment_name_id = request.query_params.get(
        "equipment_id"
    ) or request.query_params.get("equipment_name")
    routine_category = request.query_params.get(
        "routine_type"
    ) or request.query_params.get("routine_category")
    routine_name = request.query_params.get("routine_name")

    routines = routines.exclude_planned().filter_master_request(
        section_id=section_id,
        equipment_name_id=equipment_name_id,
        routine_category=routine_category,
        routine_name=routine_name,
    )
    routines = routines.order_by("equipment_name__name")

    items = []
    for routine in routines:
        next_due_date = None
        next_due_running_hrs = None
        running_hrs_available = None

        if routine.last_routine_completion_date and routine.routine_category in [
            "CALENDAR BASED",
            "ALTERNATE PERIODIC",
        ]:
            next_due_date = routine.last_routine_completion_date + timedelta(
                days=routine.frequency_in_months * 30
            )

        if (
            routine.last_routine_completion_atrunning_hrs is not None
            and routine.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
        ):
            next_due_running_hrs = (
                float(routine.last_routine_completion_atrunning_hrs)
                + routine.frequency_in_hours
            )

        start_timedate = routine.equipment_name.start_timedate
        if start_timedate and start_timedate.tzinfo is None:
            start_timedate = make_aware(start_timedate)

        if routine.equipment_name.state == "ACTIVE" and start_timedate:
            if routine.equipment_name.rhsi is not None:
                dynamic_rhsi = round(
                    routine.equipment_name.rhsi
                    + (current_time - start_timedate).total_seconds() / 3600,
                    1,
                )
            else:
                dynamic_rhsi = None
        else:
            dynamic_rhsi = routine.equipment_name.rhsi

        if dynamic_rhsi is not None and next_due_running_hrs is not None:
            running_hrs_available = round(next_due_running_hrs - dynamic_rhsi, 1)

        children = RoutineDescription.objects.active_for_add_routine(routine.id)
        worst_status = "-"
        worst_color = "white"
        color_priority = {"#FF9999": 1, "#f7e687": 2, "orange": 3, "white": 4, "-": 5}
        representative_due_date = next_due_date

        for child in children:
            child_status = "-"
            child_color = "white"
            if (
                routine.routine_category in ["CALENDAR BASED", "ALTERNATE PERIODIC"]
                and child.due_date
            ):
                child_status, child_color = get_due_status_backend(child.due_date)
                if not representative_due_date:
                    representative_due_date = child.due_date

            if (
                routine.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
                and child.due_at_rh is not None
                and dynamic_rhsi is not None
            ):
                rh_avail = float(child.due_at_rh) - float(dynamic_rhsi)
                if rh_avail <= 0:
                    hour_status, hour_color = "Routine Due", "#FF9999"
                elif rh_avail <= 500:
                    hour_status, hour_color = (
                        f"Due In {round(rh_avail, 1)} Hrs",
                        "#f7e687",
                    )
                elif rh_avail <= 1000:
                    hour_status, hour_color = (
                        f"Due In {round(rh_avail, 1)} Hrs",
                        "orange",
                    )
                else:
                    hour_status, hour_color = (
                        f"Due In {round(rh_avail, 1)} Hrs",
                        "white",
                    )

                if color_priority.get(hour_color, 5) < color_priority.get(
                    child_color, 5
                ):
                    child_status, child_color = hour_status, hour_color

            if color_priority.get(child_color, 5) < color_priority.get(worst_color, 5):
                worst_status, worst_color = child_status, child_color

        if representative_due_date and not next_due_date:
            next_due_date = representative_due_date

        routine_other_details = RoutineDescription.objects.for_add_routine_id(
            routine.id
        ).first()
        due_bucket = _build_due_bucket(worst_color)
        items.append(
            {
                "id": routine.id,
                "section_id": getattr(routine.equipment_name.section, "id", None),
                "section_name": (
                    routine.equipment_name.section.name
                    if routine.equipment_name and routine.equipment_name.section
                    else ""
                ),
                "equipment_id": getattr(routine.equipment_name, "id", None),
                "equipment_name": (
                    routine.equipment_name.name if routine.equipment_name else ""
                ),
                "routine_type": routine.routine_category or "",
                "routine_name": (
                    routine.routine_name.name if routine.routine_name else ""
                ),
                "maintop_no": (
                    routine_other_details.maintop_no if routine_other_details else "NA"
                ),
                "last_routine_date": (
                    routine.last_routine_completion_date.strftime("%d %b, %Y")
                    if routine.last_routine_completion_date
                    else "NA"
                ),
                "next_due_date": (
                    next_due_date.strftime("%d %b, %Y")
                    if next_due_date and not isinstance(next_due_date, str)
                    else next_due_date
                    if next_due_date
                    else "NA"
                ),
                "last_routine_running_hours": (
                    str(routine.last_routine_completion_atrunning_hrs)
                    if routine.last_routine_completion_atrunning_hrs is not None
                    else "NA"
                ),
                "next_due_running_hours": (
                    str(next_due_running_hrs)
                    if next_due_running_hrs is not None
                    else "NA"
                ),
                "total_running_hours": (
                    str(round(dynamic_rhsi, 2)) if dynamic_rhsi is not None else "NA"
                ),
                "running_hours_updated_till": (
                    routine.equipment_name.rhsi_updated_until.strftime("%d %b, %Y")
                    if routine.equipment_name
                    and routine.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "running_hours_available": (
                    str(running_hrs_available)
                    if running_hrs_available is not None
                    else "NA"
                ),
                "due_status": worst_status,
                "due_bucket": due_bucket,
                "status_color": worst_color,
            }
        )

    return items


def _build_frontend_routine_plan_detail(request, pk):
    routine_detail = get_object_or_404(AddRoutineDetails, pk=pk)
    category = (
        routine_detail.routine_category.upper().strip()
        if routine_detail.routine_category
        else ""
    )

    current_time = make_aware(datetime.now())
    start_timedate = routine_detail.equipment_name.start_timedate
    if start_timedate and start_timedate.tzinfo is None:
        start_timedate = make_aware(start_timedate)

    if routine_detail.equipment_name.state == "ACTIVE" and start_timedate:
        if routine_detail.equipment_name.rhsi is not None:
            dynamic_rhsi = round(
                routine_detail.equipment_name.rhsi
                + (current_time - start_timedate).total_seconds() / 3600,
                1,
            )
        else:
            dynamic_rhsi = None
    else:
        dynamic_rhsi = routine_detail.equipment_name.rhsi

    remove_list = PlannedRoutineDescription.objects.routine_description_ids()
    routine_descriptions = (
        RoutineDescription.objects.active_for_add_routine(routine_detail)
        .order_by("routine_no")
        .exclude(id__in=remove_list)
        .non_fuss_non_draft()
    )

    items = []
    for item in routine_descriptions:
        due_status = "-"
        status_color = "white"

        if category in ["CALENDAR BASED", "ALTERNATE PERIODIC"] and item.due_date:
            due_status, status_color = get_due_status_backend(item.due_date)
        elif (
            category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
            and item.due_at_rh is not None
            and str(item.due_at_rh).strip() != ""
            and dynamic_rhsi is not None
        ):
            try:
                rh_avail = float(item.due_at_rh) - float(dynamic_rhsi)
                if rh_avail <= 0:
                    due_status, status_color = "Routine Due", "#FF9999"
                elif rh_avail <= 500:
                    due_status, status_color = (
                        f"Due in {round(rh_avail, 1)} Hrs",
                        "#f7e687",
                    )
                elif rh_avail <= 1000:
                    due_status, status_color = (
                        f"Due in {round(rh_avail, 1)} Hrs",
                        "orange",
                    )
                else:
                    due_status, status_color = (
                        f"Due in {round(rh_avail, 1)} Hrs",
                        "white",
                    )
            except (ValueError, TypeError):
                pass

        items.append(
            {
                "id": item.pk,
                "routine_name": item.routine_name.name if item.routine_name else "",
                "equipment_name": (
                    item.equipment_name.name if item.equipment_name else ""
                ),
                "maintop_no": item.maintop_no,
                "dart_number": item.dart_number,
                "routine_description": item.routine_description,
                "routine_no": item.routine_no,
                "previous_routine_completed_date": (
                    item.previous_completed_date.strftime("%Y-%m-%d")
                    if item.previous_completed_date
                    else ""
                ),
                "due_date": item.due_date.strftime("%Y-%m-%d") if item.due_date else "",
                "due_at_rh": str(item.due_at_rh or ""),
                "previous_completed_at_rh": str(item.previous_completed_at_rh or ""),
                "action_by": item.by_whom,
                "due_status": due_status,
                "due_bucket": _build_due_bucket(status_color),
                "status_color": status_color,
            }
        )

    return {
        "routine": {
            "add_routine_id": routine_detail.id,
            "routine_name": (
                routine_detail.routine_name.name if routine_detail.routine_name else ""
            ),
            "equipment_name": (
                routine_detail.equipment_name.name
                if routine_detail.equipment_name
                else ""
            ),
            "maintop_no": routine_detail.maintop_no or "",
            "routine_type": routine_detail.routine_category or "",
            "section_name": (
                routine_detail.equipment_name.section.name
                if routine_detail.equipment_name
                and routine_detail.equipment_name.section
                else ""
            ),
            "dynamic_running_hours": dynamic_rhsi,
        },
        "counts": _build_due_counts(items),
        "items": items,
    }


def _build_frontend_plan_routine_form(request, pk):
    routine_description_obj = get_object_or_404(RoutineDescription, id=pk)
    department = _get_ems_department(request)
    department_id = department.id if department else None

    exclude_q = (
        Q(authority__name="B & D", is_obs__isnull=True)
        | Q(authority__name="MO ITEM", is_obs__isnull=True)
        | Q(authority__name="SHIPYARD ALLOWANCE", is_obs__isnull=True)
        | Q(authority__name="MO ALLOWANCE", is_obs__isnull=True)
    )

    spare_obj = Spares.objects.filter(
        equipment_class__spares_mappings__isnull=False
    ).distinct()

    without_map_spare_obj = (
        Spares.objects.filter(equipment_class__spare_class__department_id=department_id)
        .exclude(exclude_q)
        .distinct()
        if department
        else Spares.objects.exclude(exclude_q).distinct()
    )

    ilms_objs = Item.objects.all()
    mapped_ilms_objs = Item.objects.filter(
        momappingtable__ilms_spare_id__isnull=False
    ).distinct()

    wed_spares = WLMSSpare.objects.all()
    wed_map_spares = SpareDataMap.objects.select_related("equipment__equipment").all()
    map_dict = {
        obj.wed_equipment: (
            obj.equipment.equipment.equipment_class
            if obj.equipment and obj.equipment.equipment
            else None
        )
        for obj in wed_map_spares
    }
    with_map_wed_spares = WLMSSpare.objects.filter(eqpt__in=map_dict.keys())

    denominations = (
        Denomination.objects.filter(department_id=department_id)
        if department_id
        else Denomination.objects.all()
    )

    equipment_obj = routine_description_obj.equipment_name
    add_routine_obj = routine_description_obj.add_routine_details
    planned_routine_obj = (
        PlannedRoutineDescription.objects.filter(
            routine_description_id=routine_description_obj
        )
        .order_by("-id")
        .first()
    )

    def _format_date(value):
        return value.strftime("%Y-%m-%d") if value else ""

    def _stringify(value):
        return "" if value in (None, "") else str(value)

    routine_payload = {
        "id": routine_description_obj.id,
        "add_routine_id": getattr(add_routine_obj, "id", None),
        "sub_department": (
            equipment_obj.sub_department.name
            if equipment_obj and equipment_obj.sub_department
            else ""
        ),
        "equipment_name": equipment_obj.name if equipment_obj else "",
        "equipment_nomenclature": (
            equipment_obj.nomenclature
            if equipment_obj and equipment_obj.nomenclature
            else ""
        ),
        "section_name": (
            equipment_obj.section.name
            if equipment_obj and equipment_obj.section
            else ""
        ),
        "rhsi": (
            _stringify(equipment_obj.rhsi)
            if equipment_obj
            else _stringify(routine_description_obj.rhs_i)
        ),
        "maintop_no": routine_description_obj.maintop_no or "",
        "routine_no": routine_description_obj.routine_no or "",
        "dart_no": routine_description_obj.dart_number or "",
        "routine_name": (
            routine_description_obj.routine_name.name
            if routine_description_obj.routine_name
            else ""
        ),
        "routine_description": routine_description_obj.routine_description or "",
        "due_date": _format_date(routine_description_obj.due_date),
        "due_at_rh": _stringify(routine_description_obj.due_at_rh),
        "routine_due_rh": _stringify(routine_description_obj.due_at_rh),
        "previous_completed_date": _format_date(
            routine_description_obj.previous_completed_date
        ),
        "previous_completed_at_rh": _stringify(
            routine_description_obj.previous_completed_at_rh
        ),
        "action_by": routine_description_obj.by_whom or "",
        "planned_commencement_date": _format_date(
            planned_routine_obj.planned_commencement_date
            if planned_routine_obj
            else None
        ),
        "spares_required": (
            "YES"
            if planned_routine_obj and planned_routine_obj.spares_required
            else "NO"
        ),
    }

    spares_payload = []

    if planned_routine_obj:
        obs_spares = PlannedRoutineSpareList.objects.filter(
            planned_routine_description=planned_routine_obj,
            is_deleted=False,
        ).order_by("pattern_number")
        obs_spares_by_pattern = {
            spare.pattern_number: spare
            for spare in Spares.objects.filter(
                pattern_number__in=obs_spares.values_list("pattern_number", flat=True)
            )
        }
        for spare in obs_spares:
            obs_item = obs_spares_by_pattern.get(spare.pattern_number)
            spares_payload.append(
                {
                    "pattern_number": spare.pattern_number or "",
                    "oem_part_number": spare.pattern_number or "",
                    "spare_description": obs_item.description if obs_item else "",
                    "inventory_type": "OBS",
                    "wed_inventory_type": "",
                    "quantity_required": spare.quantity_required or 0,
                    "action": "",
                }
            )

        mo_description = MOPlannedSparesDescription.objects.filter(
            routine_description_id=routine_description_obj,
            is_deleted=False,
        ).first()
        if mo_description:
            mo_spares = PlannedMOSpareList.objects.filter(
                planned_spares_description=mo_description,
                is_deleted=False,
            ).order_by("pattern_number")
            mo_items_by_code = {
                item.item_code: item
                for item in Item.objects.filter(
                    item_code__in=mo_spares.values_list("pattern_number", flat=True)
                )
            }
            for spare in mo_spares:
                mo_item = mo_items_by_code.get(spare.pattern_number)
                spares_payload.append(
                    {
                        "pattern_number": spare.pattern_number or "",
                        "oem_part_number": spare.pattern_number or "",
                        "spare_description": mo_item.item_desc if mo_item else "",
                        "inventory_type": "MO",
                        "wed_inventory_type": "",
                        "quantity_required": spare.quantity_required or 0,
                        "action": "",
                    }
                )

        wed_description = PlannedSparesDescription.objects.filter(
            routine_description_id=routine_description_obj,
            is_deleted=False,
        ).first()
        if wed_description:
            wed_spare_rows = PlannedWEDSpareList.objects.filter(
                planned_spares_description=wed_description,
                is_deleted=False,
            ).order_by("pattern_number")
            wed_items_by_code = {
                item.item_code: item
                for item in WLMSSpare.objects.filter(
                    item_code__in=wed_spare_rows.values_list(
                        "pattern_number", flat=True
                    )
                )
            }
            for spare in wed_spare_rows:
                wed_item = wed_items_by_code.get(spare.pattern_number)
                spares_payload.append(
                    {
                        "pattern_number": spare.pattern_number or "",
                        "oem_part_number": spare.pattern_number or "",
                        "spare_description": wed_item.item_desc if wed_item else "",
                        "inventory_type": "WED",
                        "wed_inventory_type": (
                            wed_item.wlms_inventory if wed_item else ""
                        ),
                        "quantity_required": spare.quantity_required or 0,
                        "action": "",
                    }
                )

    return {
        "routine": routine_payload,
        "lookup": {
            "obs_pil_mapped": [
                {
                    "code": spare.pattern_number,
                    "name": spare.description or "",
                    "description": spare.description or "",
                }
                for spare in spare_obj
            ],
            "obs_pil_unmapped": [
                {
                    "code": spare.pattern_number,
                    "name": spare.description or "",
                    "description": spare.description or "",
                }
                for spare in without_map_spare_obj
            ],
            "mo_all": [
                {
                    "code": item.item_code,
                    "name": item.item_desc or "",
                    "description": item.item_desc or "",
                }
                for item in ilms_objs
            ],
            "mo_mapped": [
                {
                    "code": item.item_code,
                    "name": item.item_desc or "",
                    "description": item.item_desc or "",
                }
                for item in mapped_ilms_objs
            ],
            "wed_all": [
                {
                    "code": spare.item_code,
                    "name": spare.item_desc or "",
                    "description": spare.item_desc or "",
                }
                for spare in wed_spares
            ],
            "wed_mapped": [
                {
                    "code": spare.item_code,
                    "name": spare.item_desc or "",
                    "mapped_equipment_class": map_dict.get(spare.eqpt),
                }
                for spare in with_map_wed_spares
            ],
            "denominations": [
                {"id": denomination.id, "name": denomination.name}
                for denomination in denominations
            ],
        },
        "spares": spares_payload,
    }


def _save_planned_routine(pk, validated_data):
    spares_required = True if validated_data.get("spares_required") == "YES" else False
    planned_date = validated_data.get("planned_commencement_date")
    if planned_date == "":
        planned_date = None

    spares = validated_data.get("spares", [])
    get_routine_obj = get_object_or_404(RoutineDescription, pk=pk)

    if spares_required and not spares:
        spares_required = False

    planned_routine_obj = PlannedRoutineDescription.objects.create(
        routine_description_id=get_routine_obj,
        spares_required=spares_required,
        planned_commencement_date=planned_date,
    )

    if spares:
        for spare in spares:
            pattern = spare.get("pattern")
            qty = spare.get("qty", 0)
            inventory_type = spare.get("inventory_type", "")

            if inventory_type == "WED":
                wed_spare_obj = WLMSSpare.objects.filter(item_code=pattern).first()
                obj, _planned_spares_desc = (
                    PlannedSparesDescription.objects.update_or_create(
                        spares_required=spares_required,
                        wlms_spare_id=wed_spare_obj,
                        routine_description_id=get_routine_obj,
                    )
                )
                PlannedWEDSpareList.objects.update_or_create(
                    pattern_number=pattern,
                    quantity_required=qty,
                    planned_spares_description=obj,
                )

            elif inventory_type == "MO":
                item_obj = Item.objects.filter(item_code=pattern).first()
                if item_obj:
                    obj, _mo_planned = (
                        MOPlannedSparesDescription.objects.update_or_create(
                            item_id_id=item_obj.pk,
                            routine_description_id_id=get_routine_obj.id,
                        )
                    )
                    PlannedMOSpareList.objects.update_or_create(
                        pattern_number=pattern,
                        quantity_required=qty,
                        planned_spares_description_id=obj.id,
                    )

            else:
                PlannedRoutineSpareList.objects.update_or_create(
                    pattern_number=pattern,
                    quantity_required=qty,
                    planned_routine_description=planned_routine_obj,
                )

    equipment_name_obj = get_routine_obj.equipment_name
    equipment_obj = equipment_name_obj.sfd_equipment if equipment_name_obj else None

    obs_pil_patterns = [
        spare.get("pattern")
        for spare in spares
        if spare.get("inventory_type") in ("OBS", "PIL", "")
    ]
    if obs_pil_patterns and equipment_obj:
        section_name_obj = equipment_name_obj.section
        spare_objs = Spares.objects.filter(pattern_number__in=obs_pil_patterns)
        from obs.models import EquipmentClass

        equipment_classes = EquipmentClass.objects.filter(
            id__in=spare_objs.values_list("equipment_class_id", flat=True)
        ).distinct()

        for eq_class in equipment_classes:
            mapping_exists = SparesMapping.objects.filter(
                equipment_class=eq_class,
                equipment=equipment_obj,
                section_name_id=section_name_obj,
            ).exists()
            if not mapping_exists:
                SparesMapping.objects.create(
                    equipment_class=eq_class,
                    equipment=equipment_obj,
                    section_name_id=section_name_obj,
                )

    wed_patterns = [
        spare.get("pattern") for spare in spares if spare.get("inventory_type") == "WED"
    ]
    if wed_patterns and equipment_obj:
        for pattern in wed_patterns:
            wed_spare_obj = WLMSSpare.objects.filter(item_code=pattern).first()
            if not wed_spare_obj or not wed_spare_obj.eqpt:
                continue
            wed_equipment_obj = wed_spare_obj.eqpt
            map_exists = SpareDataMap.objects.filter(
                equipment=equipment_obj,
                wed_equipment=wed_equipment_obj,
            ).exists()
            if not map_exists:
                SpareDataMap.objects.create(
                    equipment=equipment_obj,
                    wed_equipment=wed_equipment_obj,
                )

    mo_patterns = [
        spare.get("pattern") for spare in spares if spare.get("inventory_type") == "MO"
    ]
    if mo_patterns and equipment_obj:
        for pattern in mo_patterns:
            item_obj = Item.objects.filter(item_code=pattern).first()
            if not item_obj:
                continue

            vendor = None
            if pattern and "-" in pattern:
                vendor_prefix = pattern.split("-", 1)[0]
                modified_vendor_code = vendor_prefix[1:]
                vendor = Vendor.objects.filter(vendor_code=modified_vendor_code).first()

            map_exists = MoMappingTable.objects.filter(
                ilms_spare_id=item_obj,
                equipment=equipment_obj,
            ).exists()
            if not map_exists:
                MoMappingTable.objects.create(
                    ilms_spare_id=item_obj,
                    equipment=equipment_obj,
                    vendor_id=vendor,
                )

    return planned_routine_obj


def _build_frontend_planned_routine_results(request):
    current_time = make_aware(datetime.now())
    department = _get_ems_department(request)
    department_id = department.id if department else None

    routines = AddRoutineDetails.objects.planned_only().with_dashboard_relations()
    if department_id:
        routines = routines.filter(equipment_name__section__department_id=department_id)

    section_id = request.query_params.get("section_id") or request.query_params.get(
        "section"
    )
    equipment_name_id = request.query_params.get(
        "equipment_id"
    ) or request.query_params.get("equipment_name")
    routine_category = request.query_params.get(
        "routine_type"
    ) or request.query_params.get("routine_category")
    routine_name = request.query_params.get("routine_name")

    routines = routines.filter_master_request(
        section_id=section_id,
        equipment_name_id=equipment_name_id,
        routine_category=routine_category,
        routine_name=routine_name,
    )

    items = []
    for routine in routines:
        next_due_date = None
        next_due_running_hrs = None
        running_hrs_available = None

        if routine.last_routine_completion_date and routine.routine_category in [
            "CALENDAR BASED",
            "ALTERNATE PERIODIC",
        ]:
            next_due_date = routine.last_routine_completion_date + timedelta(
                days=routine.frequency_in_months * 30
            )

        if (
            routine.last_routine_completion_atrunning_hrs is not None
            and routine.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
        ):
            next_due_running_hrs = (
                float(routine.last_routine_completion_atrunning_hrs)
                + routine.frequency_in_hours
            )

        start_timedate = routine.equipment_name.start_timedate
        if start_timedate and start_timedate.tzinfo is None:
            start_timedate = make_aware(start_timedate)

        if routine.equipment_name.state == "ACTIVE" and start_timedate:
            if routine.equipment_name.rhsi is not None:
                dynamic_rhsi = round(
                    routine.equipment_name.rhsi
                    + (current_time - start_timedate).total_seconds() / 3600,
                    1,
                )
            else:
                dynamic_rhsi = None
        else:
            dynamic_rhsi = routine.equipment_name.rhsi

        if dynamic_rhsi is not None and next_due_running_hrs is not None:
            running_hrs_available = round(next_due_running_hrs - dynamic_rhsi, 1)

        routine_other_details = RoutineDescription.objects.for_add_routine_id(
            routine.id
        ).first()
        routine_pair_qs = RoutineDescription.objects.for_routine_and_equipment(
            routine.routine_name,
            routine.equipment_name,
        )

        items.append(
            {
                "id": routine.pk,
                "section_id": getattr(routine.equipment_name.section, "id", None),
                "section_name": (
                    routine.equipment_name.section.name
                    if routine.equipment_name and routine.equipment_name.section
                    else ""
                ),
                "equipment_id": getattr(routine.equipment_name, "id", None),
                "equipment_name": (
                    routine.equipment_name.name if routine.equipment_name else ""
                ),
                "routine_type": routine.routine_category or "",
                "routine_name": (
                    routine.routine_name.name if routine.routine_name else ""
                ),
                "maintop_no": (
                    routine_other_details.maintop_no if routine_other_details else "NA"
                ),
                "last_routine_date": (
                    routine.last_routine_completion_date.strftime("%d %b, %Y")
                    if routine.last_routine_completion_date
                    else "NA"
                ),
                "next_due_date": (
                    next_due_date.strftime("%d %b, %Y") if next_due_date else "NA"
                ),
                "last_routine_running_hours": (
                    str(routine.last_routine_completion_atrunning_hrs)
                    if routine.last_routine_completion_atrunning_hrs is not None
                    else "NA"
                ),
                "next_due_running_hours": (
                    str(next_due_running_hrs)
                    if next_due_running_hrs is not None
                    else "NA"
                ),
                "total_running_hours": (
                    str(round(dynamic_rhsi, 2)) if dynamic_rhsi is not None else "NA"
                ),
                "running_hours_updated_till": (
                    routine.equipment_name.rhsi_updated_until.strftime("%d %b, %Y")
                    if routine.equipment_name
                    and routine.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "running_hours_available": (
                    str(running_hrs_available)
                    if running_hrs_available is not None
                    else "NA"
                ),
                "total_routines": routine_pair_qs.count(),
                "dyd_routines": routine_pair_qs.dyd().count(),
                "remarks": routine.remarks or "",
            }
        )
    return items


def _build_frontend_planned_routine_detail(pk):
    routine_detail = get_object_or_404(AddRoutineDetails, pk=pk)
    routine_obj = PlannedRoutineDescription.objects.filter(
        routine_description_id__add_routine_details__id=pk,
        is_deleted=False,
    ).values_list("routine_description_id__id", flat=True)

    routine_descriptions = RoutineDescription.objects.filter(
        id__in=routine_obj,
        equipment_name=routine_detail.equipment_name,
        routine_name=routine_detail.routine_name,
    ).order_by("routine_no")

    items = []
    for item in routine_descriptions:
        planned_item = PlannedRoutineDescription.objects.filter(
            routine_description_id__id=item.id,
            is_deleted=False,
        ).first()
        items.append(
            {
                "id": item.pk,
                "routine_name": item.routine_name.name if item.routine_name else "",
                "equipment_name": (
                    item.equipment_name.name if item.equipment_name else ""
                ),
                "maintop_no": item.maintop_no,
                "dart_number": item.dart_number,
                "routine_description": item.routine_description,
                "routine_no": item.routine_no,
                "planned_commencement_date": (
                    planned_item.planned_commencement_date.strftime("%Y-%m-%d")
                    if planned_item and planned_item.planned_commencement_date
                    else "NA"
                ),
                "rhsi": (
                    str(item.equipment_name.rhsi)
                    if item.equipment_name.rhsi is not None
                    else "NA"
                ),
                "rhsi_updated_until": (
                    item.equipment_name.rhsi_updated_until.strftime("%Y-%m-%d")
                    if item.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "previous_routine_completed_date": (
                    item.previous_completed_date.strftime("%Y-%m-%d")
                    if item.previous_completed_date
                    else ""
                ),
                "due_date": item.due_date.strftime("%Y-%m-%d") if item.due_date else "",
                "due_at_rh": (
                    str(item.due_at_rh) if item.due_at_rh is not None else "NA"
                ),
                "previous_completed_at_rh": (
                    str(item.previous_completed_at_rh)
                    if item.previous_completed_at_rh is not None
                    else "NA"
                ),
                "action_by": item.by_whom,
                "spare_req": (
                    "YES" if planned_item and planned_item.spares_required else "NO"
                ),
                "category_data": routine_detail.routine_category,
            }
        )

    return {
        "routine": {
            "add_routine_id": routine_detail.id,
            "routine_name": (
                routine_detail.routine_name.name if routine_detail.routine_name else ""
            ),
            "equipment_name": (
                routine_detail.equipment_name.name
                if routine_detail.equipment_name
                else ""
            ),
            "routine_type": routine_detail.routine_category or "",
            "section_name": (
                routine_detail.equipment_name.section.name
                if routine_detail.equipment_name
                and routine_detail.equipment_name.section
                else ""
            ),
            "dynamic_running_hours": (
                float(routine_detail.equipment_name.rhsi)
                if routine_detail.equipment_name.rhsi is not None
                else None
            ),
        },
        "items": items,
    }


def _build_frontend_fuss_raised_search(request):
    today = date.today()
    fuss_raised_list = (
        FussRaiseDetails.objects.select_related(
            "routine_description_id",
            "routine_description_id__routine_name",
            "routine_description_id__equipment_name",
        )
        .order_by("-created_at")
        .filter(isclosed_fuss=False)
    )

    unique_departments = sorted(
        list(set(filter(None, fuss_raised_list.values_list("department", flat=True))))
    )
    unique_equipments = sorted(
        list(set(filter(None, fuss_raised_list.values_list("equipment", flat=True))))
    )
    unique_routine_types = sorted(
        list(
            set(
                filter(
                    None,
                    fuss_raised_list.values_list(
                        "routine_description_id__routine_name__name",
                        flat=True,
                    ),
                )
            )
        )
    )

    dept_equipment_map = defaultdict(list)
    for fuss in fuss_raised_list:
        if fuss.department and fuss.equipment:
            if fuss.equipment not in dept_equipment_map[fuss.department]:
                dept_equipment_map[fuss.department].append(fuss.equipment)
    dept_equipment_map = {
        dept: sorted(equips) for dept, equips in dept_equipment_map.items()
    }

    selected_year = request.query_params.get("year", "0000")
    selected_month = request.query_params.get("month", "00")

    items = []
    seen = set()
    for fuss in fuss_raised_list:
        routine = fuss.routine_description_id
        if not routine:
            continue

        unique_key = (
            fuss.department,
            fuss.equipment,
            routine.routine_name.name if routine.routine_name else "",
        )
        if unique_key in seen:
            continue
        seen.add(unique_key)

        next_due_date = None
        if fuss.last_undertaken and fuss.frequency:
            try:
                freq_months = int(float(fuss.frequency))
                next_due_date = fuss.last_undertaken + relativedelta(months=freq_months)
            except (ValueError, TypeError):
                next_due_date = None

        should_append = True
        if next_due_date:
            due_month = next_due_date.strftime("%m")
            due_year = next_due_date.strftime("%Y")
            if selected_month != "00" and selected_year != "0000":
                should_append = (
                    selected_month == due_month and selected_year == due_year
                )
            elif selected_month != "00" and selected_year == "0000":
                should_append = selected_month == due_month
            elif selected_month == "00" and selected_year != "0000":
                should_append = selected_year == due_year

        if should_append:
            total_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=routine.equipment_name,
            ).count()
            dyd_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=routine.equipment_name,
                by_whom="DYD",
            ).count()
            ss_routines = total_routines - dyd_routines
            routines_due = 1 if next_due_date and next_due_date <= today else 0

            items.append(
                {
                    "id": fuss.id,
                    "department": fuss.department,
                    "equipment": fuss.equipment,
                    "routine_name": (
                        routine.routine_name.name if routine.routine_name else ""
                    ),
                    "status": (
                        "Active"
                        if fuss.due_date and fuss.due_date >= today
                        else "Overdue"
                    ),
                    "maintop_no": fuss.maintop_no or "NA",
                    "routines_by_dyd": dyd_routines,
                    "routines_by_ss": ss_routines,
                    "total_routines": total_routines,
                    "routines_due": routines_due,
                    "due_date": (
                        next_due_date.strftime("%d %b, %Y") if next_due_date else "NA"
                    ),
                }
            )

    return {
        "filters": {
            "departments": unique_departments,
            "equipments": unique_equipments,
            "routine_types": unique_routine_types,
            "dept_equipment_map": dept_equipment_map,
        },
        "items": items,
    }


def _build_frontend_fuss_raised_detail(request):
    routine_name = request.query_params.get("routine")
    fuss_list = FussRaiseDetails.objects.filter(isclosed_fuss=False).select_related(
        "routine_description_id",
        "routine_description_id__routine_name",
        "routine_description_id__equipment_name",
    )
    if routine_name:
        fuss_list = fuss_list.filter(
            routine_description_id__routine_name__name=routine_name
        )

    items = []
    departments = set()
    equipments = set()
    routine_types = set()
    dept_equipment_map = defaultdict(list)
    for fuss in fuss_list:
        routine = fuss.routine_description_id
        if fuss.department:
            departments.add(fuss.department)
        if fuss.equipment:
            equipments.add(fuss.equipment)
            if (
                fuss.department
                and fuss.equipment not in dept_equipment_map[fuss.department]
            ):
                dept_equipment_map[fuss.department].append(fuss.equipment)
        if routine and routine.routine_name and routine.routine_name.name:
            routine_types.add(routine.routine_name.name)
        items.append(
            {
                "id": fuss.id,
                "ship": fuss.ship,
                "department": fuss.department,
                "serial_no": fuss.serial_no,
                "fuss_date": (
                    fuss.fuss_date.strftime("%Y-%m-%d") if fuss.fuss_date else ""
                ),
                "due_date": fuss.due_date.strftime("%Y-%m-%d") if fuss.due_date else "",
                "equipment": fuss.equipment,
                "maintop_no": fuss.maintop_no,
                "frequency": fuss.frequency,
                "routine_id": routine.id if routine else None,
                "routine_no": routine.routine_no if routine else "",
                "routine_name": (
                    routine.routine_name.name
                    if routine and routine.routine_name
                    else ""
                ),
            }
        )

    return {
        "filters": {
            "departments": sorted(departments),
            "equipments": sorted(equipments),
            "routine_types": sorted(routine_types),
            "dept_equipment_map": {
                dept: sorted(equips) for dept, equips in dept_equipment_map.items()
            },
        },
        "items": items,
    }


@extend_schema(
    tags=["EMS"],
    responses={200: EquipmentStatusResponseSerializer},
)
class EquipmentStatusView(GenericAPIView):
    serializer_class = EquipmentStatusResponseSerializer
    section_codes = ("AER", "AMR", "FER", "OMS")

    def get_queryset(self):
        return EquipmentName.objects.status_sections(
            self.section_codes
        ).ordered_by_name()

    def get(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(
            {
                f"{section}_equipment_list": queryset.filter(section__name=section)
                for section in self.section_codes
            }
        )
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: GetSectionNameResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="department_id",
            type=int,
            location=OpenApiParameter.PATH,
            description="Department ID",
            required=False,
        )
    ],
)
@api_view(["GET"])
def GetSectionNameView(request, department_id=None):
    if not department_id:
        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            department_id = request.user.department_id
        else:
            dept = Department.objects.first()
            department_id = dept.id if dept else None

    if department_id:
        sections = SectionName.objects.for_department_id(
            department_id
        ).ordered_by_name()
    else:
        sections = SectionName.objects.ordered_by_name()

    section_dict = {s.name: s.id for s in sections}
    return Response({"section_name": section_dict})


@extend_schema(
    tags=["EMS"],
    responses={200: GetEquipmentNameResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="section_id",
            type=int,
            location=OpenApiParameter.PATH,
            description="Section ID",
            required=False,
        )
    ],
)
@api_view(["GET"])
def GetEquipmentNameView(request, section_id=None):
    if section_id and section_id != 0:
        equipment = EquipmentName.objects.for_section(section_id).ordered_by_name()
    else:
        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dept = request.user.CustomUser_profile.department
            equipment = EquipmentName.objects.for_department(dept).ordered_by_id()
        else:
            equipment = EquipmentName.objects.ordered_by_id()

    equipment_dict = {e.name: e.id for e in equipment}
    return Response({"equipment_name": equipment_dict})


@extend_schema(
    tags=["EMS"],
    responses={200: GetsrarEquipmentNameResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="section_id",
            type=int,
            location=OpenApiParameter.PATH,
            description="Section ID",
            required=False,
        )
    ],
)
@api_view(["GET"])
def GetsrarEquipmentNameView(request, section_id=None):
    if section_id and section_id != 0:
        equipment = (
            EquipmentName.objects.for_section(section_id).srar_only().ordered_by_name()
        )
    else:
        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dept = request.user.CustomUser_profile.department
            equipment = (
                EquipmentName.objects.for_department(dept).srar_only().ordered_by_id()
            )
        else:
            equipment = EquipmentName.objects.srar_only().ordered_by_id()

    equipment_dict = {e.name: e.id for e in equipment}
    return Response({"equipment_name": equipment_dict})


@extend_schema(
    tags=["EMS"],
    responses={200: GetEquipmentNameWithoutRHSINullRowsResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="section_id",
            type=int,
            location=OpenApiParameter.PATH,
            description="Section ID",
            required=False,
        )
    ],
)
@api_view(["GET"])
def GetEquipmentNameWithoutRHSINullRowsView(request, section_id=None):
    if section_id and section_id != 0:
        equipment = (
            EquipmentName.objects.for_section(section_id).with_rhsi().ordered_by_name()
        )
    else:
        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dept = request.user.CustomUser_profile.department
            equipment = (
                EquipmentName.objects.for_department(dept).with_rhsi().ordered_by_id()
            )
        else:
            equipment = EquipmentName.objects.with_rhsi().ordered_by_id()

    equipment_dict = {e.name: e.id for e in equipment}
    return Response({"equipment_name": equipment_dict})


@extend_schema(
    tags=["EMS"],
    responses={200: GetRoutineNameResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="equipment_name_id",
            type=int,
            location=OpenApiParameter.PATH,
            description="Equipment Name ID",
            required=False,
        ),
        OpenApiParameter(
            name="routine_category",
            type=str,
            location=OpenApiParameter.PATH,
            description="Routine Category",
            required=False,
        ),
    ],
)
@api_view(["GET"])
def GetRoutineNameView(request, equipment_name_id=None, routine_category=None):
    if not equipment_name_id or not routine_category:
        return Response({"routine_name": {}}, status=status.HTTP_400_BAD_REQUEST)

    routine_entries = AddRoutineDetails.objects.for_equipment_and_category(
        equipment_name_id,
        routine_category,
    ).order_by("routine_name")

    routine_name_dict = {
        entry.routine_name.name: entry.routine_name_id for entry in routine_entries
    }
    return Response({"routine_name": routine_name_dict})


@extend_schema(
    tags=["EMS"],
    request=EmsSectionCreateRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def SectionCreateView(request):
    serializer = EmsSectionCreateRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    form_id = data.get("form_id")
    section_id = data.get("section_id")
    sub_dep = data.get("sub_dep")
    section_name = data.get("section_name")

    if request.user.is_authenticated and hasattr(request.user, "user_profile"):
        dep_id = request.user.CustomUser_profile.department_id
        dep_obj = request.user.CustomUser_profile.department
    else:
        dept = Department.objects.first()
        dep_id = dept.id if dept else None
        dep_obj = dept

    if form_id == "1":
        if section_id and section_name:
            SectionName.objects.filter(id=section_id, department=dep_id).update(
                name=section_name.upper()
            )
            SubDepartment.objects.filter(id=sub_dep, department_name=dep_id).update(
                name=section_name.upper()
            )
            return Response({"message": "Updated section successfully."})
        else:
            name = section_name.upper() if section_name else ""
            if not name.strip():
                return Response(
                    {"message": "This field cannot be left blank."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for instance in SectionName.objects.all():
                a = re.sub(r"[^a-zA-Z0-9]", "", instance.name).upper()
                b = re.sub(r"[^a-zA-Z0-9]", "", name).upper()
                if a == b:
                    return Response(
                        {"message": "Section Name entry already exists."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            SectionName.objects.create(name=name, department=dep_obj)
            return Response({"message": "New section entry created."})

    # Fallback to returning list
    sections = (
        SectionName.objects.filter(department=dep_id)
        if dep_id
        else SectionName.objects.all()
    )
    section_names = [s.name for s in sections]
    sub_deps = SubDepartment.objects.filter(name__in=section_names)
    sub_dep_map = {sd.name: sd.id for sd in sub_deps}

    result_list = []
    for s in sections:
        result_list.append(
            {"name": s.name, "section_id": s.id, "sub_dep_id": sub_dep_map.get(s.name)}
        )

    return Response({"result": result_list})


@extend_schema(
    tags=["EMS"],
    request=EmsEquipmentCreateRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def EquipmentCreateView(request):
    serializer = EmsEquipmentCreateRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    form_id = data.get("form_id")
    name = data.get("name")
    section_id = data.get("section")

    if form_id == "2":
        if not name or not section_id:
            return Response(
                {"message": "Equipment name and section are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        name = name.upper()
        section = get_object_or_404(SectionName, id=section_id)

        for instance in EquipmentName.objects.filter(section=section):
            a = re.sub(r"[^a-zA-Z0-9]", "", instance.name).upper()
            b = re.sub(r"[^a-zA-Z0-9]", "", name).upper()
            if a == b:
                return Response(
                    {"message": "Equipment name entry already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        EquipmentName.objects.create(name=name, section=section)
        return Response({"message": "New equipment entry created."})

    return Response({"message": "Invalid form id."}, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    tags=["EMS"],
    request=EmsTotalRunningHoursCreateRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def TotalRunningHoursCreateView(request):
    serializer = EmsTotalRunningHoursCreateRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    equipment_id = data.get("equipment") or data.get("name")
    date_upto_str = data.get("rhsi_updated_until")
    total_running_hours = data.get("rhsi")

    equipment = get_object_or_404(EquipmentName, id=equipment_id)

    if equipment.state == "ACTIVE":
        return Response(
            {
                "message": "⚠️ Make equipment state as INACTIVE first to save cumulative running hours."
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        date_upto = parse_datetime(date_upto_str)
        if is_naive(date_upto):
            date_upto = make_aware(date_upto)
    except Exception:
        return Response(
            {"message": "⚠️ Invalid date format."}, status=status.HTTP_400_BAD_REQUEST
        )

    if equipment.rhsi_updated_until and date_upto < equipment.rhsi_updated_until:
        return Response(
            {
                "message": "⚠️ Error: Selected date is earlier than the last recorded date."
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    if equipment.rhsi is not None:
        if float(total_running_hours) <= float(equipment.rhsi):
            return Response(
                {
                    "message": "⚠️ Error: Total running hours must be greater than previously recorded value."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

    equipment.rhsi = total_running_hours
    equipment.rhsi_updated_until = date_upto
    equipment.state = "INACTIVE"
    equipment.save()

    return Response({"message": "✅ Running hours updated successfully."})


@extend_schema(
    tags=["EMS"],
    request=EmsMonthlyRunningHoursSaveRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def MonthlyRunningHoursSaveView(request):
    serializer = EmsMonthlyRunningHoursSaveRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    equipment = get_object_or_404(EquipmentName, id=data.get("equipment_id"))

    for entry in data.get("rows", []):
        s_dt = parse_datetime(entry["start"])
        start_time = make_aware(s_dt) if is_naive(s_dt) else s_dt

        e_dt = parse_datetime(entry["stop"])
        stop_time = make_aware(e_dt) if is_naive(e_dt) else e_dt

        diff_hours = (stop_time - start_time).total_seconds() / 3600
        pseudo_diff = ems_float_to_pseudo(diff_hours)

        PostEquipmentStateChangeHistorySave.objects.create(
            month_name=stop_time.strftime("%b %Y"),
            start_time=start_time,
            stop_time=stop_time,
            started_at_location=entry.get("location"),
            diff_in_hours=pseudo_diff,
            equipment_name=equipment,
        )

        equipment.rhsi = ems_add_hours_to_pseudo(equipment.rhsi, diff_hours)
        equipment.rhsi_updated_until = stop_time
        equipment.state = "INACTIVE"

    equipment.save()
    return Response({"message": "✅ All running hours saved successfully."})


@extend_schema(
    tags=["EMS"],
    responses={200: GetEquipmentHistoryResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            location=OpenApiParameter.PATH,
            description="Equipment ID",
            required=True,
        )
    ],
)
@api_view(["GET"])
def GetEquipmentHistoryJSON(request, pk):
    history = PostEquipmentStateChangeHistorySave.objects.filter(
        equipment_name_id=pk
    ).order_by("-stop_time")[:5]
    result = []
    for log in history:
        result.append(
            {
                "month": log.month_name,
                "start": (
                    log.start_time.strftime("%d %b %H:%M") if log.start_time else "-"
                ),
                "stop": log.stop_time.strftime("%d %b %H:%M") if log.stop_time else "-",
                "duration": ems_pseudo_to_hhmm(log.diff_in_hours),
                "location": log.started_at_location or "-",
            }
        )
    return Response({"history": result})


@extend_schema(
    tags=["EMS"],
    request=UpdateEquipmentStateRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def UpdateEquipmentStateView(request):
    serializer = UpdateEquipmentStateRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    equipment_id = data.get("equipment_id")
    equipment = get_object_or_404(EquipmentName, pk=equipment_id)
    new_state = data.get("state")
    started_at_location = data.get("started_at_location")
    start_timedate = data.get("start_timedate")
    stop_timedate = data.get("stop_timedate")

    if start_timedate:
        start_timedate = (
            pd.to_datetime(start_timedate).tz_localize(local_tz)
            if is_naive(start_timedate)
            else pd.to_datetime(start_timedate).tz_convert(local_tz)
        )
    if stop_timedate:
        stop_timedate = (
            pd.to_datetime(stop_timedate).tz_localize(local_tz)
            if is_naive(stop_timedate)
            else pd.to_datetime(stop_timedate).tz_convert(local_tz)
        )

    if equipment.state == "INACTIVE" and new_state == "ACTIVE":
        if not start_timedate:
            return Response(
                {
                    "success": False,
                    "message": "Start time is required when activating equipment.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if equipment.rhsi_updated_until:
            last_until = pd.to_datetime(equipment.rhsi_updated_until).tz_convert(
                local_tz
            )
            if start_timedate <= last_until:
                return Response(
                    {
                        "success": False,
                        "message": "Start time must be after the last recorded activity"
                        + f"({last_until.strftime('%d %b %Y %H:%M')}).",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        equipment.start_timedate = start_timedate
        equipment.stop_timedate = None

        PostEquipmentStateChangeHistorySave.objects.create(
            month_name=start_timedate.strftime("%b %Y"),
            start_time=start_timedate,
            stop_time=None,
            started_at_location=started_at_location,
            diff_in_hours=0,
            equipment_name=equipment,
        )

    elif equipment.state == "ACTIVE" and new_state == "INACTIVE":
        if not stop_timedate:
            return Response(
                {
                    "success": False,
                    "message": "Stop time is required when deactivating equipment.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        last_start = (
            pd.to_datetime(equipment.start_timedate).tz_convert(local_tz)
            if equipment.start_timedate
            else None
        )
        if last_start and stop_timedate <= last_start:
            return Response(
                {
                    "success": False,
                    "message": f"Stop time must be after the start time ({last_start.strftime('%d %b %Y %H:%M')}).",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        duration_hours = 0
        if equipment.start_timedate:
            duration_hours = round(
                (stop_timedate - last_start).total_seconds() / 3600, 1
            )
            if equipment.rhsi is not None:
                equipment.rhsi += duration_hours

        equipment.state = "INACTIVE"
        equipment.rhsi_updated_until = stop_timedate
        equipment.save()

        history_entry = PostEquipmentStateChangeHistorySave.objects.filter(
            equipment_name=equipment, stop_time=None
        ).first()
        if history_entry:
            history_entry.stop_time = stop_timedate
            history_entry.diff_in_hours = duration_hours
            history_entry.save()
        else:
            PostEquipmentStateChangeHistorySave.objects.create(
                month_name=stop_timedate.strftime("%b %Y"),
                start_time=equipment.start_timedate,
                stop_time=stop_timedate,
                started_at_location=started_at_location,
                diff_in_hours=duration_hours,
                equipment_name=equipment,
            )

    equipment.state = new_state
    equipment.started_at_location = started_at_location
    equipment.save()

    return Response(
        {"success": True, "message": "Equipment state updated successfully."}
    )


# Legacy complete_routine function removed. The logic is now handled in InitiateCloseRoutineAPIView.post.


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            location=OpenApiParameter.PATH,
            description="Routine Description ID",
            required=True,
        )
    ],
)
@api_view(["GET"])
def Planroutine(request, pk):
    res_data = _build_frontend_plan_routine_form(request, pk)
    legacy_payload = {
        "spare_obj": [
            {
                "pattern_number": item["code"],
                "name": item["name"],
                "description": item["description"],
            }
            for item in res_data["lookup"]["obs_pil_mapped"]
        ],
        "without_map_spare_obj": [
            {
                "pattern_number": item["code"],
                "name": item["name"],
                "description": item["description"],
            }
            for item in res_data["lookup"]["obs_pil_unmapped"]
        ],
        "ilms_objs": [
            {
                "item_code": item["code"],
                "name": item["name"],
                "description": item["description"],
            }
            for item in res_data["lookup"]["mo_all"]
        ],
        "mapped_ilms_objs": [
            {
                "item_code": item["code"],
                "name": item["name"],
                "description": item["description"],
            }
            for item in res_data["lookup"]["mo_mapped"]
        ],
        "oem_pil_spares": [
            {
                "pattern_number": item["code"],
                "name": item["name"],
                "description": item["description"],
            }
            for item in res_data["lookup"]["obs_pil_mapped"]
        ],
        "wed_spares": [
            {
                "item_code": item["code"],
                "name": item["name"],
                "description": item["description"],
            }
            for item in res_data["lookup"]["wed_all"]
        ],
        "with_map_wed_spares": [
            {
                "item_code": item["code"],
                "name": item["name"],
                "insmaequp_name": item["mapped_equipment_class"],
            }
            for item in res_data["lookup"]["wed_mapped"]
        ],
        "denominations": res_data["lookup"]["denominations"],
        "routine_description_id": res_data["routine"]["id"],
        "routine_description": res_data["routine"]["routine_description"],
    }
    return Response(legacy_payload)


@extend_schema(
    tags=["EMS"],
    request=PlanRoutineSaveRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            location=OpenApiParameter.PATH,
            description="Routine Description ID",
            required=True,
        )
    ],
)
@api_view(["POST"])
def plan_routine_save(request, pk):
    serializer = PlanRoutineSaveRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    planned_obj = _save_planned_routine(pk, serializer.validated_data)
    return Response(
        {
            "status": "success",
            "message": "Routine planned successfully.",
            "routine_id": pk,
            "planned_routine_id": planned_obj.id if planned_obj else None,
        },
        status=status.HTTP_200_OK,
    )


@extend_schema(tags=["EMS"], request=AddRoutineDetailsRequestSerializer)
@api_view(["POST"])
def add_routine_details(request):
    serializer = AddRoutineDetailsRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    routine_name_obj = get_object_or_404(UniqueRoutineName, pk=data["routine_name"])
    equipment_obj = get_object_or_404(EquipmentName, pk=data["equipment_name"])

    commencement_date = data.get("last_routine_commencement_date")
    completion_date = data.get("last_routine_completion_date")
    completion_hours = data.get("last_routine_completion_atrunning_hrs")
    remarks = data.get("remarks", "")

    existing_routine = AddRoutineDetails.objects.filter(
        equipment_name=equipment_obj,
        routine_name=routine_name_obj,
    ).first()

    if existing_routine:
        same_commencement = (
            existing_routine.last_routine_commencement_date
            and commencement_date
            and existing_routine.last_routine_commencement_date.date()
            == commencement_date.date()
        )
        same_completion = (
            existing_routine.last_routine_completion_date
            and completion_date
            and existing_routine.last_routine_completion_date.date()
            == completion_date.date()
        )
        if same_commencement and same_completion:
            same_time = (
                existing_routine.last_routine_commencement_date.time()
                != commencement_date.time()
                or existing_routine.last_routine_completion_date.time()
                != completion_date.time()
            )
            if not same_time:
                return Response(
                    {
                        "message": (
                            "The entered date range (commencement or completion "
                            "date) already exists. Please change the date range."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        existing_routine.last_routine_commencement_date = commencement_date
        existing_routine.last_routine_completion_date = completion_date
        existing_routine.last_routine_completion_atrunning_hrs = completion_hours
        existing_routine.remarks = remarks
        existing_routine.save()
        routine = existing_routine
        message = "Routine details updated successfully."
    else:
        routine = AddRoutineDetails.objects.create(
            equipment_name=equipment_obj,
            routine_name=routine_name_obj,
            last_routine_commencement_date=commencement_date,
            last_routine_completion_date=completion_date,
            last_routine_completion_atrunning_hrs=completion_hours,
            remarks=remarks,
        )
        message = "Routine details saved successfully."

    last_history = (
        PostRoutineDetails.objects.filter(routine_name=routine).order_by("-id").first()
    )
    if last_history:
        last_to_last_routine_running_hrs = (
            last_history.last_routine_due_running_hrs_actual
        )
        last_to_last_routine_date = last_history.last_routine_date_actual
        last_routine_due_running_hrs_actual = completion_hours
        last_routine_date_ideal = (
            (
                last_history.last_to_last_routine_date
                + timedelta(days=routine.frequency_in_months * 30)
            )
            if routine.frequency_in_months and last_history.last_to_last_routine_date
            else None
        )
        last_routine_due_running_hrs_ideal = (
            (last_history.last_to_last_routine_running_hrs + routine.frequency_in_hours)
            if routine.frequency_in_hours
            and last_history.last_to_last_routine_running_hrs is not None
            else None
        )
        hrs_between_two_routines = (completion_hours or 0) - (
            last_history.last_to_last_routine_running_hrs or 0
        )
    else:
        last_to_last_routine_running_hrs = completion_hours
        last_to_last_routine_date = completion_date
        last_routine_due_running_hrs_actual = completion_hours
        last_routine_date_ideal = completion_date
        last_routine_due_running_hrs_ideal = completion_hours
        hrs_between_two_routines = routine.frequency_in_hours

    no_of_days_to_complete_routine = (
        (completion_date - commencement_date).days
        if commencement_date and completion_date
        else 0
    )

    PostRoutineDetails.objects.create(
        routine_name=routine,
        last_to_last_routine_date=last_to_last_routine_date,
        last_to_last_routine_running_hrs=last_to_last_routine_running_hrs or 0,
        last_routine_due_running_hrs_ideal=last_routine_due_running_hrs_ideal,
        last_routine_due_running_hrs_actual=last_routine_due_running_hrs_actual,
        last_routine_date_ideal=last_routine_date_ideal,
        last_routine_date_actual=completion_date,
        hrs_between_two_routines=hrs_between_two_routines,
        no_of_days_to_complete_routine=no_of_days_to_complete_routine,
        remarks=remarks,
    )

    return Response(
        {"message": message, "id": routine.id}, status=status.HTTP_201_CREATED
    )


@extend_schema(tags=["EMS"], request=AddRoutineDescriptionRequestSerializer)
@api_view(["POST"])
def add_routine_description(request):
    serializer = AddRoutineDescriptionRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    equipment_obj = get_object_or_404(EquipmentName, pk=data["equipment_name"])
    routine_name_obj = get_object_or_404(UniqueRoutineName, pk=data["routine_name"])
    maintop_no = data.get("maintop_no", "")
    routine_no = data.get("routine_no", "")
    routine_description = data["routine_description"]
    by_whom = data.get("by_whom", "")

    add_routine_details_instance = get_object_or_404(
        AddRoutineDetails,
        equipment_name=equipment_obj,
        routine_name=routine_name_obj,
    )

    existing = RoutineDescription.objects.filter(
        equipment_name=equipment_obj,
        routine_name=routine_name_obj,
        maintop_no=maintop_no,
        routine_no=routine_no,
    ).first()

    if existing:
        if existing.routine_description == routine_description:
            return Response(
                {
                    "message": (
                        "Routine Description text entered is exactly matching "
                        "with existing routine."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(
            {"message": "Entry of this routine description already exists."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created = RoutineDescription.objects.create(
        equipment_name=equipment_obj,
        routine_name=routine_name_obj,
        add_routine_details=add_routine_details_instance,
        maintop_no=maintop_no,
        routine_no=routine_no,
        routine_description=routine_description,
        by_whom=by_whom,
    )

    return Response(
        {
            "message": "New routine description entry for selected equipment created.",
            "id": created.id,
        },
        status=status.HTTP_201_CREATED,
    )


@extend_schema(tags=["EMS"], request=EditRoutineNameRequestSerializer)
@api_view(["POST"])
def edit_routine_name(request):
    serializer = EditRoutineNameRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    entry = AddRoutineDetails.objects.filter(
        routine_name_id=data["routine_name"],
        equipment_name_id=data["equipment_name"],
    ).first()
    if not entry:
        return Response(
            {"message": "Cumulative Running hours entry not created."},
            status=status.HTTP_404_NOT_FOUND,
        )

    entry.routine_name.name = data["new_name"]
    entry.routine_name.save(update_fields=["name"])

    routine_category = data["routine_category"]
    update_fields = []
    if routine_category == "RUNNING HOUR BASED":
        entry.frequency_in_hours = data.get("frequency_in_hours")
        update_fields.append("frequency_in_hours")
    elif routine_category == "CALENDAR BASED":
        entry.frequency_in_months = data.get("frequency_in_months")
        update_fields.append("frequency_in_months")
    else:
        entry.frequency_in_months = data.get("frequency_in_months")
        entry.frequency_in_hours = data.get("frequency_in_hours")
        update_fields.extend(["frequency_in_months", "frequency_in_hours"])
    entry.save(update_fields=update_fields)

    return Response({"message": "Routine Name and other details updated."})


@extend_schema(tags=["EMS"], request=EditEquipmentNameRequestSerializer)
@api_view(["POST"])
def edit_equipment_name(request):
    serializer = EditEquipmentNameRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    equipment = get_object_or_404(EquipmentName, pk=data["equipment_name"])
    if not data["new_name"]:
        return Response(
            {"message": "Enter a valid equipment name."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    equipment.name = data["new_name"]
    equipment.save(update_fields=["name"])
    return Response({"message": "Equipment name updated."})


@extend_schema(tags=["EMS"], request=CreateRoutineFrequencyRequestSerializer)
@api_view(["POST"])
def create_routine_frequency(request):
    serializer = CreateRoutineFrequencyRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    equipment = get_object_or_404(EquipmentName, pk=data["equipment_name"])
    routine_name_instance, _created = UniqueRoutineName.objects.get_or_create(
        name=data["routine_name"]
    )

    if AddRoutineDetails.objects.filter(
        equipment_name=equipment,
        routine_name=routine_name_instance,
        frequency_in_months=data.get("frequency_in_months"),
        frequency_in_hours=data.get("frequency_in_hours"),
    ).exists():
        return Response(
            {"message": "This frequency entry already exists for this equipment."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    AddRoutineDetails.objects.create(
        equipment_name=equipment,
        routine_name=routine_name_instance,
        frequency_in_months=data.get("frequency_in_months"),
        frequency_in_hours=data.get("frequency_in_hours"),
        routine_category=data["routine_category"],
    )
    return Response(
        {"message": "Routine Frequency created successfully!"},
        status=status.HTTP_201_CREATED,
    )


@extend_schema(tags=["EMS"], request=CreateEquipmentNameRequestSerializer)
@api_view(["POST"])
def create_equipment_name(request):
    serializer = CreateEquipmentNameRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    section = get_object_or_404(SectionName, pk=data["section"])
    name = data["name"].strip().upper()
    if not name:
        return Response(
            {"message": "This field cannot be left blank."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    normalized = re.sub(r"[^a-zA-Z0-9]", "", name).upper()
    for instance in EquipmentName.objects.filter(section=section):
        existing_normalized = re.sub(r"[^a-zA-Z0-9]", "", instance.name).upper()
        if existing_normalized == normalized:
            return Response(
                {"message": "Equipment name entry already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    equipment = EquipmentName.objects.create(name=name, section=section)
    return Response(
        {"message": "New equipment entry created.", "id": equipment.id},
        status=status.HTTP_201_CREATED,
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def thumbnail_data(request):
    department_id = request.user.department_id

    maintop_count = (
        AddRoutineDetails.objects.filter(
            equipment_name__section__department_id=department_id
        )
        .exclude(frequency__icontains="R")
        .exclude(frequency__icontains="POLICY")
        .values("maintop_no")
        .distinct()
        .count()
    )

    rh_count = 0
    for routine in AddRoutineDetails.objects.filter(
        routine_category="RUNNING HOUR BASED",
        equipment_name__section__department_id=department_id,
    ).select_related("equipment_name"):
        last_rhrs = routine.last_routine_completion_atrunning_hrs
        rhsi = routine.equipment_name.rhsi
        if last_rhrs is not None and rhsi is not None and routine.frequency_in_hours:
            next_due_rhrs = last_rhrs + routine.frequency_in_hours
            if round(next_due_rhrs - rhsi, 1) <= 1000:
                rh_count += 1

    current_datetime = dj_timezone.now()
    six_months_later = current_datetime + relativedelta(months=6)
    one_month_later = current_datetime + relativedelta(months=1)
    calendar_routines = AddRoutineDetails.objects.filter(
        routine_category__in=["CALENDAR BASED", "ALTERNATE PERIODIC"],
        equipment_name__section__department_id=department_id,
    )
    cal_count = 0
    fuss_count = 0
    for routine in calendar_routines:
        if not (routine.last_routine_completion_date and routine.frequency_in_months):
            continue
        next_due_date = routine.last_routine_completion_date + timedelta(
            days=routine.frequency_in_months * 30
        )
        if next_due_date <= six_months_later:
            cal_count += 1
        if next_due_date <= one_month_later:
            fuss_count += 1

    aber_count = ShipEquipment.objects.filter(
        section_f_key__department_id=department_id,
        status__iexact="INACTIVE",
    ).count()

    return Response(
        {
            "data": {
                "count": maintop_count,
                "rh_due_count": rh_count,
                "cal_due_count": cal_count,
                "fuss_due_count": fuss_count,
                "aber_due_count": aber_count,
            }
        }
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def section_rhsi_barchart(request, section_name):
    section = get_object_or_404(SectionName, pk=section_name)
    section_colors = {
        "AC PLANT": "#FF5733",
        "DG SET": "#FFD700",
        "FIRE PUMP": "#1E90FF",
        "MAIN ENGINE": "#32CD32",
        "COOLING SYSTEM": "#FF8C00",
        "COMPRESSOR": "#800080",
        "FUEL SYSTEM": "#A52A2A",
    }
    inactive_base_color = section_colors.get(section.name, "#4682B4")
    current_time = dj_timezone.now()

    equipment_list = (
        EquipmentName.objects.filter(section=section, rhsi__isnull=False)
        .exclude(rhsi=0)
        .order_by("name")
    )

    grouped = {}
    for equipment in equipment_list:
        entry = grouped.setdefault(
            equipment.name,
            {"id": equipment.id, "rhsi": 0, "state": None, "updated_until": None},
        )
        entry["rhsi"] = max(entry["rhsi"], equipment.rhsi or 0)
        if equipment.state == "ACTIVE":
            entry["state"] = "ACTIVE"
        elif not entry["state"]:
            entry["state"] = equipment.state
        if equipment.rhsi_updated_until and (
            not entry["updated_until"]
            or equipment.rhsi_updated_until > entry["updated_until"]
        ):
            entry["updated_until"] = equipment.rhsi_updated_until

    equipment_ids, labels, values, formatted_values, tooltips, colors = (
        [],
        [],
        [],
        [],
        [],
        [],
    )
    for name, info in grouped.items():
        equipment_ids.append(info["id"])
        rhsi = float(info["rhsi"] or 0.0)
        state = info["state"]
        updated_until = info["updated_until"]
        rhsi_status = "on" if state == "ACTIVE" else "off"
        color = "green" if state == "ACTIVE" else inactive_base_color
        updated_date = (
            current_time.strftime("%d %b %Y")
            if state == "ACTIVE"
            else (updated_until.strftime("%d %b %Y") if updated_until else "NA")
        )

        labels.append(name)
        values.append(int(rhsi))
        formatted_values.append(ems_pseudo_to_hhmm(rhsi))
        colors.append(color)
        tooltips.append(
            f"RHSI (HH:MM) :- {ems_pseudo_to_hhmm(rhsi)}\n"
            f"RHSI Upto :- {updated_date}\nStatus :- {rhsi_status.upper()}"
        )

    return Response(
        {
            "equipment_ids": equipment_ids,
            "section": section.name,
            "labels": labels,
            "values": values,
            "formatted_values": formatted_values,
            "tooltips": tooltips,
            "colors": colors,
        }
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def active_users(request):
    user = request.user.CustomUser_profile
    if getattr(user.department, "name", None) == "ADMIN":
        queryset = type(user).objects.all()
    else:
        queryset = type(user).objects.filter(
            is_approved=True,
            department_id=user.department_id,
        )

    users = [
        {
            "username": entry.personnel_number,
            "rank": entry.rank,
            "designation": entry.designation,
            "name": entry.get_full_name(),
            "department": entry.department.name if entry.department else "",
        }
        for entry in queryset.select_related("department")
    ]
    return Response({"history_users": users})


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def equipment_state_list(request):
    department_id = request.user.department_id
    sections = SectionName.objects.filter(department_id=department_id).values_list(
        "id", flat=True
    )
    equipments = (
        EquipmentName.objects.filter(section_id__in=sections)
        .select_related("section")
        .order_by("name")
    )

    result = [
        {
            "pk": equipment.pk,
            "section_id": equipment.section_id,
            "section": equipment.section.name if equipment.section else "",
            "equipment_name": equipment.name,
            "state": equipment.state,
            "location": equipment.started_at_location or "",
            "total_running_hrs": equipment.rhsi if equipment.rhsi else "NA",
            "start_time_raw": (
                equipment.start_timedate.strftime("%d %b, %Y %H:%M")
                if equipment.start_timedate
                else None
            ),
            "running_hrs_updated_tilldate": (
                equipment.rhsi_updated_until.strftime("%d %b, %Y %H:%M")
                if equipment.rhsi_updated_until
                else "NA"
            ),
        }
        for equipment in equipments
    ]
    return Response({"result": result})


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            location=OpenApiParameter.PATH,
            description="Routine Description ID",
            required=True,
        )
    ],
)
@api_view(["POST", "DELETE", "GET"])
def DeletePlanned_routine(request, pk):
    planned_routine = get_object_or_404(
        PlannedRoutineDescription, routine_description_id=pk
    )
    planned_routine_spare = PlannedRoutineSpareList.objects.filter(
        planned_routine_description__id=planned_routine.id
    )
    planned_routine_spare.delete()
    planned_routine.delete()
    return Response(
        {"success": True, "message": "Planned routine deleted successfully."}
    )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="section", type=str, description="Section ID", required=False
        ),
        OpenApiParameter(
            name="equipment_name",
            type=str,
            description="Equipment Name ID",
            required=False,
        ),
        OpenApiParameter(
            name="routine_category",
            type=str,
            description="Routine Category",
            required=False,
        ),
        OpenApiParameter(
            name="routine_name",
            type=str,
            description="Routine Name/Frequency query",
            required=False,
        ),
    ],
)
@api_view(["GET"])
def SearchMergedView(request):
    current_time = make_aware(datetime.now())

    if (
        request.user.is_authenticated
        and hasattr(request.user, "user_profile")
        and request.user.user_profile
    ):
        dept_id = getattr(request.user.user_profile, "department_id", None)
    elif request.user.is_authenticated and getattr(request.user, "department_id", None):
        dept_id = request.user.department_id
    else:
        dept = Department.objects.first()
        dept_id = dept.id if dept else None

    routines = AddRoutineDetails.objects.select_related(
        "equipment_name", "equipment_name__section", "routine_name"
    )

    if dept_id:
        routines = routines.filter(equipment_name__section__department_id=dept_id)

    planned_add_routine_ids = PlannedRoutineDescription.objects.values_list(
        "routine_description_id__add_routine_details_id", flat=True
    ).distinct()
    routines = routines.exclude(id__in=planned_add_routine_ids)

    # Request Filters
    section_id = request.query_params.get("section")
    equipment_name_id = request.query_params.get("equipment_name")
    routine_category = request.query_params.get("routine_category")
    routine_name = request.query_params.get("routine_name")

    if section_id and section_id != "0":
        routines = routines.filter(equipment_name__section_id=section_id)
    if equipment_name_id and equipment_name_id != "0":
        routines = routines.filter(equipment_name_id=equipment_name_id)
    if routine_category and routine_category != "0":
        routines = routines.filter(routine_category=routine_category)
    if routine_name:
        routines = routines.filter(frequency__icontains=routine_name)

    routines = routines.order_by("equipment_name__name")

    result = []
    for e in routines:
        next_due_date = None
        next_due_running_hrs = None
        running_hrs_available = None

        if e.last_routine_completion_date and e.routine_category in [
            "CALENDAR BASED",
            "ALTERNATE PERIODIC",
        ]:
            next_due_date = e.last_routine_completion_date + timedelta(
                days=e.frequency_in_months * 30
            )

        if (
            e.last_routine_completion_atrunning_hrs is not None
            and str(e.last_routine_completion_atrunning_hrs).strip() != ""
            and e.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
        ):
            try:
                next_due_running_hrs = float(
                    e.last_routine_completion_atrunning_hrs
                ) + (e.frequency_in_hours or 0)
            except (ValueError, TypeError):
                next_due_running_hrs = None

        start_timedate = e.equipment_name.start_timedate
        if start_timedate and start_timedate.tzinfo is None:
            start_timedate = make_aware(start_timedate)

        if e.equipment_name.state == "ACTIVE" and start_timedate:
            if e.equipment_name.rhsi is not None:
                dynamic_rhsi = round(
                    e.equipment_name.rhsi
                    + (current_time - start_timedate).total_seconds() / 3600,
                    1,
                )
            else:
                dynamic_rhsi = None
        else:
            dynamic_rhsi = e.equipment_name.rhsi

        if dynamic_rhsi is not None and next_due_running_hrs is not None:
            running_hrs_available = round(next_due_running_hrs - dynamic_rhsi, 1)

        # Get worst status
        children = RoutineDescription.objects.filter(
            add_routine_details=e.id, is_close=False
        )
        worst_status = "-"
        worst_color = "white"
        color_priority = {"#FF9999": 1, "#f7e687": 2, "orange": 3, "white": 4, "-": 5}
        representative_due_date = next_due_date

        for child in children:
            c_status, c_color = "-", "white"
            if (
                e.routine_category in ["CALENDAR BASED", "ALTERNATE PERIODIC"]
                and child.due_date
            ):
                c_status, c_color = get_due_status_backend(child.due_date)
                if not representative_due_date:
                    representative_due_date = child.due_date

            if (
                e.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
                and child.due_at_rh is not None
                and str(child.due_at_rh).strip() != ""
                and dynamic_rhsi is not None
            ):
                try:
                    rh_avail = float(child.due_at_rh) - float(dynamic_rhsi)
                    if rh_avail <= 0:
                        h_status, h_color = "Routine Due", "#FF9999"
                    elif rh_avail <= 500:
                        h_status, h_color = (
                            f"Due In {round(rh_avail, 1)} Hrs",
                            "#f7e687",
                        )
                    elif rh_avail <= 1000:
                        h_status, h_color = f"Due In {round(rh_avail, 1)} Hrs", "orange"
                    else:
                        h_status, h_color = f"Due In {round(rh_avail, 1)} Hrs", "white"

                    if color_priority.get(h_color, 5) < color_priority.get(c_color, 5):
                        c_status, c_color = h_status, h_color
                except (ValueError, TypeError):
                    pass

            if color_priority.get(c_color, 5) < color_priority.get(worst_color, 5):
                worst_status, worst_color = c_status, c_color

        due_status = worst_status
        status_color = worst_color
        if representative_due_date and not next_due_date:
            next_due_date = representative_due_date

        # Other details (e.g. first routine description)
        routine_other_details = RoutineDescription.objects.filter(
            add_routine_details=e.id
        ).first()

        result.append(
            {
                "pk": e.pk,
                "routine_name": e.routine_name.name if e.routine_name else "",
                "section": (
                    e.equipment_name.section.name
                    if e.equipment_name and e.equipment_name.section
                    else ""
                ),
                "equipment_name": e.equipment_name.name if e.equipment_name else "",
                "maintop_no": (
                    routine_other_details.maintop_no if routine_other_details else "NA"
                ),
                "last_routine_date": (
                    e.last_routine_completion_date.strftime("%d %b, %Y")
                    if e.last_routine_completion_date
                    else "NA"
                ),
                "date": (
                    next_due_date.strftime("%d %b, %Y")
                    if next_due_date and not isinstance(next_due_date, str)
                    else next_due_date
                    if next_due_date
                    else "NA"
                ),
                "last_routine_running_hrs": (
                    e.last_routine_completion_atrunning_hrs
                    if e.last_routine_completion_atrunning_hrs
                    else "NA"
                ),
                "next_due_running_hrs": (
                    next_due_running_hrs if next_due_running_hrs else "NA"
                ),
                "total_running_hrs": round(dynamic_rhsi, 2) if dynamic_rhsi else "NA",
                "running_hrs_updated_tilldate": (
                    e.equipment_name.rhsi_updated_until.strftime("%d %b, %Y")
                    if e.equipment_name and e.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "running_hrs_available": (
                    running_hrs_available if running_hrs_available else "NA"
                ),
                "valid_upto": (
                    e.equipment_name.rhsi_updated_until.strftime("%d %b, %Y")
                    if e.equipment_name and e.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "due_status": due_status,
                "status_color": status_color,
            }
        )

    return Response(result)


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            location=OpenApiParameter.PATH,
            description="Add Routine Details ID",
            required=False,
        )
    ],
)
@api_view(["GET"])
def SearchDetailView(request, pk=None):
    if not pk:
        return Response({"result": []})
    payload = _build_frontend_routine_plan_detail(request, pk)
    result = [
        {
            "pk": item["id"],
            "routine_name": item["routine_name"],
            "equipment_name": item["equipment_name"],
            "maintop_no": item["maintop_no"],
            "dart_number": item["dart_number"],
            "routine_description": item["routine_description"],
            "routine_no": item["routine_no"],
            "previous_routine_completed_date": item["previous_routine_completed_date"],
            "due_date": item["due_date"],
            "due_at_rh": item["due_at_rh"],
            "previous_completed_at_rh": item["previous_completed_at_rh"],
            "action_by": item["action_by"],
            "due_status": item["due_status"],
            "status_color": item["status_color"],
        }
        for item in payload["items"]
    ]
    return Response({"result": result})


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="section", type=str, description="Section ID", required=False
        ),
        OpenApiParameter(
            name="equipment_name",
            type=str,
            description="Equipment Name ID",
            required=False,
        ),
        OpenApiParameter(
            name="routine_category",
            type=str,
            description="Routine Category",
            required=False,
        ),
        OpenApiParameter(
            name="routine_name",
            type=str,
            description="Routine Name/Frequency query",
            required=False,
        ),
    ],
)
@api_view(["GET"])
def PlannedRoutinesMasterView(request):
    current_time = make_aware(datetime.now())

    if request.user.is_authenticated and hasattr(request.user, "user_profile"):
        dept_id = request.user.CustomUser_profile.department_id
    else:
        dept = Department.objects.first()
        dept_id = dept.id if dept else None

    # Filter to only routines that are planned (i.e. exist in PlannedRoutineDescription)
    routines = AddRoutineDetails.objects.planned_only().with_dashboard_relations()

    if dept_id:
        routines = routines.filter(equipment_name__section__department_id=dept_id)

    # Request Filters
    section_id = request.query_params.get("section")
    equipment_name_id = request.query_params.get("equipment_name")
    routine_category = request.query_params.get("routine_category")
    routine_name = request.query_params.get("routine_name")

    routines = routines.filter_master_request(
        section_id=section_id,
        equipment_name_id=equipment_name_id,
        routine_category=routine_category,
        routine_name=routine_name,
    )

    result = []
    for e in routines:
        next_due_date = None
        next_due_running_hrs = None
        running_hrs_available = None

        if e.last_routine_completion_date and e.routine_category in [
            "CALENDAR BASED",
            "ALTERNATE PERIODIC",
        ]:
            next_due_date = e.last_routine_completion_date + timedelta(
                days=e.frequency_in_months * 30
            )

        if (
            e.last_routine_completion_atrunning_hrs is not None
            and e.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
        ):
            next_due_running_hrs = (
                float(e.last_routine_completion_atrunning_hrs) + e.frequency_in_hours
            )

        start_timedate = e.equipment_name.start_timedate
        if start_timedate and start_timedate.tzinfo is None:
            start_timedate = make_aware(start_timedate)

        if e.equipment_name.state == "ACTIVE" and start_timedate:
            if e.equipment_name.rhsi is not None:
                dynamic_rhsi = round(
                    e.equipment_name.rhsi
                    + (current_time - start_timedate).total_seconds() / 3600,
                    1,
                )
            else:
                dynamic_rhsi = None
        else:
            dynamic_rhsi = e.equipment_name.rhsi

        if dynamic_rhsi is not None and next_due_running_hrs is not None:
            running_hrs_available = round(next_due_running_hrs - dynamic_rhsi, 1)

        routine_other_details = RoutineDescription.objects.for_add_routine_id(
            e.id
        ).first()
        routine_pair_qs = RoutineDescription.objects.for_routine_and_equipment(
            e.routine_name,
            e.equipment_name,
        )
        total_routines = routine_pair_qs.count()
        dyd_routines = routine_pair_qs.dyd().count()

        result.append(
            {
                "pk": e.pk,
                "routine_name": e.routine_name.name if e.routine_name else "",
                "section": (
                    e.equipment_name.section.name
                    if e.equipment_name and e.equipment_name.section
                    else ""
                ),
                "equipment_name": e.equipment_name.name if e.equipment_name else "",
                "maintop_no": (
                    routine_other_details.maintop_no if routine_other_details else "NA"
                ),
                "last_routine_date": (
                    e.last_routine_completion_date.strftime("%d %b, %Y")
                    if e.last_routine_completion_date
                    else "NA"
                ),
                "date": next_due_date.strftime("%d %b, %Y") if next_due_date else "NA",
                "last_routine_running_hrs": (
                    e.last_routine_completion_atrunning_hrs
                    if e.last_routine_completion_atrunning_hrs is not None
                    else "NA"
                ),
                "next_due_running_hrs": (
                    next_due_running_hrs if next_due_running_hrs is not None else "NA"
                ),
                "total_running_hrs": (
                    round(dynamic_rhsi, 2) if dynamic_rhsi is not None else "NA"
                ),
                "running_hrs_updated_tilldate": (
                    e.equipment_name.rhsi_updated_until.strftime("%d %b, %Y")
                    if e.equipment_name and e.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "running_hrs_available": (
                    running_hrs_available if running_hrs_available is not None else "NA"
                ),
                "valid_upto": (
                    e.equipment_name.rhsi_updated_until.strftime("%d %b, %Y")
                    if e.equipment_name and e.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "total_routines": total_routines,
                "dyd_routines": dyd_routines,
                "remarks": e.remarks or "",
            }
        )

    return Response(result)


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="category", type=str, description="Category", required=False
        )
    ],
)
@api_view(["GET"])
def get_routine_names(request):
    choices = [choice[0] for choice in LessAddRoutineDetails.routine_category_choices]
    return Response({"routine_names": choices})


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="category", type=str, description="Category filter", required=False
        ),
        OpenApiParameter(
            name="sectionId", type=int, description="Section ID filter", required=False
        ),
        OpenApiParameter(
            name="equipment_id",
            type=int,
            description="Equipment ID filter",
            required=False,
        ),
    ],
)
@api_view(["GET"])
def get_routine_name(request):
    try:
        category = request.query_params.get("category")
        sectionId = request.query_params.get("sectionId")
        equipment_id = request.query_params.get("equipment_id")

        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dept = request.user.CustomUser_profile.department
        else:
            dept = Department.objects.first()

        routine_names = AddRoutineDetails.objects.filter(
            equipment_name__section__department=dept
        )

        if category and category != "0":
            routine_names = routine_names.filter(routine_category__iexact=category)
        if sectionId and sectionId != "0":
            routine_names = routine_names.filter(equipment_name__section_id=sectionId)
        if equipment_id and equipment_id != "0":
            routine_names = routine_names.filter(equipment_name_id=equipment_id)

        routine_names = routine_names.select_related("frequency_f_key")

        result = {}
        for entry in routine_names:
            freq = entry.frequency or ""
            desc = entry.frequency_f_key.Description if entry.frequency_f_key else ""
            label = f"{freq} | {desc} ".strip()
            result[freq] = label

        return Response({"success": True, "routine_names": result})
    except Exception as e:
        return Response(
            {"success": False, "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="section_id", type=int, description="Section ID", required=True
        )
    ],
)
@api_view(["GET"])
def get_equipment_by_section(request):
    section_id = request.query_params.get("section_id")
    equipments = EquipmentName.objects.filter(section_id=section_id).values(
        "id", "name"
    )
    return Response({"equipments": list(equipments)})


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendRoutinePlanFiltersSerializer},
    parameters=[
        OpenApiParameter(
            name="section_id",
            type=int,
            description="Selected section/sub-department ID",
            required=False,
        ),
        OpenApiParameter(
            name="equipment_id",
            type=int,
            description="Selected equipment ID",
            required=False,
        ),
        OpenApiParameter(
            name="routine_type",
            type=str,
            description="Selected routine type",
            required=False,
        ),
    ],
)
class FrontendRoutinePlanFiltersAPIView(APIView):
    def get(self, request):
        department = _get_ems_department(request)
        section_id = request.query_params.get("section_id")
        payload = {
            "sections": _serialize_frontend_sections(department),
            "equipment": _serialize_frontend_equipment(section_id, department),
            "routine_types": _serialize_frontend_routine_types(),
            "routine_names": _serialize_frontend_routine_names(request, department),
        }
        serializer = FrontendRoutinePlanFiltersSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendRoutinePlanSearchResponseSerializer},
    parameters=[
        OpenApiParameter(
            name="section_id",
            type=int,
            description="Selected section/sub-department ID",
            required=False,
        ),
        OpenApiParameter(
            name="equipment_id",
            type=int,
            description="Selected equipment ID",
            required=False,
        ),
        OpenApiParameter(
            name="routine_type",
            type=str,
            description="Routine type/category",
            required=False,
        ),
        OpenApiParameter(
            name="routine_name",
            type=str,
            description="Routine name/frequency option",
            required=False,
        ),
    ],
)
class FrontendRoutinePlanSearchAPIView(APIView):
    def get(self, request):
        department = _get_ems_department(request)
        section_id = request.query_params.get("section_id")
        items = _build_frontend_routine_plan_results(request)
        payload = {
            "filters": {
                "sections": _serialize_frontend_sections(department),
                "equipment": _serialize_frontend_equipment(section_id, department),
                "routine_types": _serialize_frontend_routine_types(),
                "routine_names": _serialize_frontend_routine_names(request, department),
            },
            "counts": _build_due_counts(items),
            "items": items,
        }
        serializer = FrontendRoutinePlanSearchResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendRoutinePlanDetailResponseSerializer},
)
class FrontendRoutinePlanDetailAPIView(APIView):
    def get(self, request, pk):
        payload = _build_frontend_routine_plan_detail(request, pk)
        serializer = FrontendRoutinePlanDetailResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendRoutinePlanFormResponseSerializer},
)
class FrontendPlanRoutineAPIView(APIView):
    def get(self, request, pk):
        payload = _build_frontend_plan_routine_form(request, pk)
        serializer = FrontendRoutinePlanFormResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    request=PlanRoutineSaveRequestSerializer,
    responses={200: FrontendRoutinePlanSaveResponseSerializer},
)
class FrontendPlanRoutineSaveAPIView(APIView):
    def post(self, request, pk):
        serializer = PlanRoutineSaveRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        planned_routine = _save_planned_routine(pk, serializer.validated_data)
        response_serializer = FrontendRoutinePlanSaveResponseSerializer(
            {
                "success": True,
                "message": "Routine planned successfully.",
                "planned_routine_id": planned_routine.id,
            }
        )
        return Response(response_serializer.data, status=status.HTTP_200_OK)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendPlannedRoutineSearchResponseSerializer},
    parameters=[
        OpenApiParameter(name="section_id", type=int, required=False),
        OpenApiParameter(name="equipment_id", type=int, required=False),
        OpenApiParameter(name="routine_type", type=str, required=False),
        OpenApiParameter(name="routine_name", type=str, required=False),
    ],
)
class FrontendPlannedRoutineSearchAPIView(APIView):
    def get(self, request):
        department = _get_ems_department(request)
        section_id = request.query_params.get("section_id")
        payload = {
            "filters": {
                "sections": _serialize_frontend_sections(department),
                "equipment": _serialize_frontend_equipment(section_id, department),
                "routine_types": _serialize_frontend_routine_types(),
                "routine_names": _serialize_frontend_routine_names(request, department),
            },
            "items": _build_frontend_planned_routine_results(request),
        }
        serializer = FrontendPlannedRoutineSearchResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendPlannedRoutineDetailResponseSerializer},
)
class FrontendPlannedRoutineDetailAPIView(APIView):
    def get(self, request, pk):
        payload = _build_frontend_planned_routine_detail(pk)
        serializer = FrontendPlannedRoutineDetailResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendFussRaisedSearchResponseSerializer},
    parameters=[
        OpenApiParameter(name="year", type=str, required=False),
        OpenApiParameter(name="month", type=str, required=False),
    ],
)
class FrontendFussRaisedSearchAPIView(APIView):
    def get(self, request):
        payload = _build_frontend_fuss_raised_search(request)
        serializer = FrontendFussRaisedSearchResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: FrontendFussRaisedDetailResponseSerializer},
    parameters=[OpenApiParameter(name="routine", type=str, required=False)],
)
class FrontendFussRaisedDetailAPIView(APIView):
    def get(self, request):
        payload = _build_frontend_fuss_raised_detail(request)
        serializer = FrontendFussRaisedDetailResponseSerializer(payload)
        return Response(serializer.data)


@extend_schema(
    tags=["EMS"],
    responses={200: InitiateCloseRoutineResponseSerializer},
)
class InitiateCloseRoutineAPIView(APIView):
    def get(self, request, pk, *args, **kwargs):
        planned_routine_obj = get_object_or_404(RoutineDescription, pk=pk)

        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dep = getattr(request.user.user_profile, "department", None)
        elif request.user.is_authenticated and getattr(
            request.user, "department", None
        ):
            dep = request.user.department
        else:
            dep = Department.objects.first()
        if not dep:
            dep = Department.objects.first()

        dep_universal_id = getattr(dep, "universal_id_m_department", None)

        rank_obj = MRanklist.objects.filter(universal_id_m_department=dep_universal_id)

        planned_obj = PlannedRoutineDescription.objects.filter(
            routine_description_id=planned_routine_obj.pk
        ).first()

        spares_for_routine = []
        if planned_obj:
            spares_for_routine = PlannedRoutineSpareList.objects.filter(
                planned_routine_description=planned_obj.id, is_deleted=False
            )

        issue_list = (
            Issue.objects.filter(dart_number=planned_routine_obj.dart_number)
            if planned_routine_obj.dart_number
            else Issue.objects.none()
        )

        fullname = ""
        rankname = ""
        if request.user.is_authenticated:
            if hasattr(request.user, "user_profile") and request.user.user_profile:
                profile = request.user.user_profile
                fullname = f"{getattr(profile, 'firstname', '') or ''} {getattr(profile, 'lastname', '') or ''}".strip()
                rank_val = getattr(profile, "rank", "")
                rankname = getattr(rank_val, "name", str(rank_val)) if rank_val else ""
            else:
                fullname = getattr(request.user, "get_full_name", lambda: "")() or str(
                    request.user
                )
                rankname = getattr(request.user, "rank", "") or ""

        max_hours = getattr(planned_routine_obj.equipment_name, "rhsi", None)

        # Extract flattened form-display fields
        old_dart_number = planned_routine_obj.dart_number

        equipment_class = ""
        equipment_serial_no = ""
        nomenclature = ""
        location_on_board = ""
        maintop_routine_number = ""

        if planned_routine_obj.equipment_name:
            eq = planned_routine_obj.equipment_name
            equipment_class = eq.extra or ""
            nomenclature = eq.name or ""
            maintop_routine_number = planned_routine_obj.maintop_no or ""

            if eq.sfd_equipment:
                sfd_eq = eq.sfd_equipment
                equipment_serial_no = sfd_eq.equipment_serial_no or ""
                location_on_board = sfd_eq.location_on_board or ""

        due_date = planned_routine_obj.due_date
        if isinstance(due_date, datetime):
            due_date = due_date.date()

        last_completion_date = None
        if planned_routine_obj.add_routine_details:
            last_completion = (
                planned_routine_obj.add_routine_details.last_routine_completion_date
            )
            if last_completion:
                if isinstance(last_completion, datetime):
                    last_completion_date = last_completion.date()
                else:
                    last_completion_date = last_completion

        maintop_remarks = "DYD"  # Matches the default template value from old HTML

        data = {
            "old_dart_number": old_dart_number,
            "equipment_class": equipment_class,
            "equipment_serial_no": equipment_serial_no,
            "nomenclature": nomenclature,
            "due_date": due_date,
            "last_completion_date": last_completion_date,
            "maintop_remarks": maintop_remarks,
            "location_on_board": location_on_board,
            "maintop_routine_number": maintop_routine_number,
            "routine_description": planned_routine_obj.routine_description,
            "minutes_range": list(range(5, 61, 10)),
            "hours_range": list(range(1, 13)),
            "planned_routine": planned_routine_obj,
            "planned_obj": planned_obj,
            "spares_for_routine": spares_for_routine,
            "issue_list": issue_list,
            "rankname": rankname,
            "fullname": fullname,
            "rank_obj": rank_obj,
            "max_hours": max_hours,
        }

        serializer = InitiateCloseRoutineResponseSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        tags=["EMS"],
        request=CompleteRoutineRequestSerializer,
        responses={200: OpenApiTypes.OBJECT},
    )
    def post(self, request, pk, *args, **kwargs):
        routine = get_object_or_404(RoutineDescription, id=pk)
        routine_details = get_object_or_404(
            AddRoutineDetails, id=routine.add_routine_details.pk
        )

        serializer = CompleteRoutineRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dep = getattr(request.user.user_profile, "department", None)
        elif request.user.is_authenticated and getattr(
            request.user, "department", None
        ):
            dep = request.user.department
        else:
            dep = Department.objects.first()
        if not dep:
            dep = Department.objects.first()

        dep_name = getattr(dep, "name", None)
        dep_id = getattr(dep, "id", None)
        frequency = routine_details.routine_category
        due_date = None
        due_at_rh = None
        complete_date = None

        complete_date_str = validated_data.get("date_of_completion")
        if complete_date_str:
            if isinstance(complete_date_str, (date, datetime)):
                complete_date = complete_date_str
            else:
                try:
                    complete_date = datetime.strptime(complete_date_str, "%Y-%m-%d")
                except ValueError:
                    try:
                        complete_date = parse_datetime(complete_date_str)
                    except Exception:
                        complete_date = None

        if frequency == "CALENDAR BASED":
            if complete_date:
                if isinstance(complete_date, date) and not isinstance(
                    complete_date, datetime
                ):
                    complete_date = datetime.combine(complete_date, datetime.min.time())
                due_date = complete_date + relativedelta(
                    months=routine_details.frequency_in_months
                )
                last_completion = routine_details.last_routine_completion_date
                if isinstance(last_completion, datetime):
                    last_completion = last_completion.date()

                comp_date_val = (
                    complete_date.date()
                    if isinstance(complete_date, datetime)
                    else complete_date
                )

                if (
                    routine_details.last_routine_completion_date is None
                    or last_completion < comp_date_val
                ):
                    routine_details.last_routine_completion_date = complete_date
                    routine_details.save(update_fields=["last_routine_completion_date"])

        elif frequency == "RUNNING HOUR BASED":
            running_hrs = validated_data.get("running_hour") or 0.0
            if running_hrs:
                due_at_rh = float(running_hrs) + float(
                    routine_details.frequency_in_hours
                )
                last_rh = routine_details.last_routine_completion_atrunning_hrs or 0.0
                if float(last_rh) <= float(running_hrs):
                    routine_details.last_routine_completion_atrunning_hrs = running_hrs
                    routine_details.save(
                        update_fields=["last_routine_completion_atrunning_hrs"]
                    )

        # Safe max
        last_entry = (
            RoutineDescription.objects.filter(
                department_f_key=dep_id, dart_sr_no__regex=r"^\d+$"
            )
            .annotate(dart_sr_no_int=Cast("dart_sr_no", IntegerField()))
            .aggregate(max_sr=Max("dart_sr_no_int"))["max_sr"]
            or 0
        )

        dart_number, dart_sr_no = generate_routine_dart_number(
            dep_name, "Routine", last_entry
        )

        rank_id = validated_data.get("rank_routine")
        rank_other = validated_data.get("rank_other")
        rank_obj = MRanklist.objects.filter(rankid=rank_id).first() if rank_id else None
        custom_rank = rank_other if str(rank_id) in ["43", "44", "45", "46"] else None

        completed_routine = CompletedRoutine.objects.create(
            routine=routine,
            date_of_completion=complete_date.date() if complete_date else None,
            hours=validated_data.get("hours"),
            minutes=validated_data.get("minutes"),
            carried_by=validated_data.get("carried_by"),
            p_no=validated_data.get("p_no"),
            rank=rank_obj,
            other_rank=custom_rank,
            total_manpower=validated_data.get("total_manpower"),
            running_hour=validated_data.get("running_hour"),
            due_running_hour=validated_data.get("due_running_hour"),
            repair_remark=validated_data.get("remarks"),
            completion_details=validated_data.get("completion_details"),
            trial_team=validated_data.get("trial_team") == "on"
            or validated_data.get("trial_team") == "Yes",
            rec_for_deletion=validated_data.get("rec_for_deletion") == "on",
            old_dart_number=validated_data.get("old_dart_number"),
            new_dart_number=dart_number,
        )

        # Update old routine description
        routine.is_close = True
        routine.save()

        # Create new due routine description
        new_routine = RoutineDescription.objects.create(
            equipment_name=routine.equipment_name,
            routine_name=routine.routine_name,
            add_routine_details=routine.add_routine_details,
            maintop_no=routine.maintop_no,
            routine_no=routine.routine_no,
            routine_description=routine.routine_description,
            dart_number=dart_number,
            by_whom=routine.by_whom,
            is_fuss=routine.is_fuss,
            department_f_key=routine.department_f_key,
            dart_sr_no=dart_sr_no,
            previous_routine=routine,
            last_routine_completion_date=(
                complete_date.date() if complete_date else None
            ),
            last_routine_completion_atrunning_hrs=validated_data.get("running_hour"),
            due_date=due_date.date() if isinstance(due_date, datetime) else due_date,
            previous_completed_date=complete_date,
            due_at_rh=due_at_rh,
            previous_completed_at_rh=validated_data.get("running_hour"),
        )
        new_routine.universal_id_t_dart = f"303-M-{new_routine.id}"
        new_routine.save(update_fields=["universal_id_t_dart"])

        # Update Planned status if exists
        planned = PlannedRoutineDescription.objects.filter(
            routine_description_id=routine.id
        ).first()
        if planned:
            planned.is_deleted = True
            planned.save()
            PlannedRoutineSpareList.objects.filter(
                planned_routine_description_id=planned.id
            ).update(is_deleted=True)

        # Bulk Spares
        spares = validated_data.get("spares", [])
        CompletedRoutineSpare.objects.bulk_create(
            [
                CompletedRoutineSpare(completed_routine=completed_routine, spare_name=s)
                for s in spares
            ]
        )

        return Response(
            {"status": "success", "message": "Routine closed successfully!"},
            status=status.HTTP_201_CREATED,
        )


# ==================== FUSS VIEWS ====================


class MulRaiseFussAPIView(APIView):
    @extend_schema(
        tags=["EMS"],
        responses={200: MulRaiseFussResponseSerializer},
        parameters=[
            OpenApiParameter(
                name="add_routine_id",
                type=int,
                description="The ID of the master AddRoutineDetails record",
                required=False,
            ),
            OpenApiParameter(
                name="selected_ids",
                type=str,
                description="Fallback: Comma-separated or list of Routine IDs for GET initialization",
                required=False,
            ),
        ],
    )
    def get(self, request, *args, **kwargs):
        # GET request: load initialization data
        add_routine_id = request.query_params.get("add_routine_id")
        selected_ids = request.query_params.getlist("selected_ids")
        if len(selected_ids) == 1 and "," in selected_ids[0]:
            selected_ids = selected_ids[0].split(",")

        if add_routine_id:
            routines = RoutineDescription.objects.filter(
                add_routine_details_id=add_routine_id
            )
        elif selected_ids:
            first_r = RoutineDescription.objects.filter(id__in=selected_ids).first()
            if first_r and first_r.add_routine_details:
                routines = RoutineDescription.objects.filter(
                    add_routine_details=first_r.add_routine_details
                )
            else:
                routines = RoutineDescription.objects.filter(id__in=selected_ids)
        else:
            first_ar = AddRoutineDetails.objects.first()
            if first_ar:
                routines = RoutineDescription.objects.filter(
                    add_routine_details=first_ar
                )
            else:
                routines = RoutineDescription.objects.all()[:5]

        # Update selected_ids to contain all resolved sibling sub-routine IDs
        selected_ids = [str(r.id) for r in routines]

        routines_data = []
        for r in routines:
            # Resolve details for this routine
            ship_name = ""
            if r.add_routine_details and r.add_routine_details.ship:
                ship_name = getattr(
                    r.add_routine_details.ship, "name", str(r.add_routine_details.ship)
                )

            dept_name = ""
            if r.equipment_name:
                if r.equipment_name.sub_department:
                    dept_name = r.equipment_name.sub_department.name
                elif r.equipment_name.section:
                    dept_name = r.equipment_name.section.name

            eq_name = r.equipment_name.name if r.equipment_name else ""

            serial_no = ""
            location_on_board = ""
            location_code = ""
            if r.equipment_name and r.equipment_name.sfd_equipment:
                sfd = r.equipment_name.sfd_equipment
                serial_no = getattr(sfd, "equipment_serial_no", "")
                location_on_board = getattr(sfd, "location_on_board", "")
                location_code = getattr(sfd, "location_code", "")

            eq_code = ""
            if r.add_routine_details:
                eq_code = r.add_routine_details.equipment_code or ""

            routines_data.append(
                {
                    "id": r.id,
                    "routine_no": r.routine_no,
                    "routine_description": r.routine_description,
                    "maintop_no": r.maintop_no,
                    "frequency": (
                        r.add_routine_details.frequency if r.add_routine_details else ""
                    ),
                    "category": (
                        r.add_routine_details.routine_category
                        if r.add_routine_details
                        else ""
                    ),
                    "equipment_name": eq_name,
                    "dart_no": r.dart_number or "",
                    # Flat form population fields inside the routine object
                    "ship": ship_name,
                    "sub_department": dept_name,
                    "serial_no": serial_no,
                    "location_on_board": location_on_board,
                    "location_code": location_code,
                    "equipment_sr_no": eq_code,
                }
            )

        deferments = MDeferment.objects.filter(active=True)
        inabilities = MInability.objects.filter(active=True)
        reasons = MReason.objects.filter(active=True)

        if (
            request.user.is_authenticated
            and hasattr(request.user, "user_profile")
            and request.user.user_profile
        ):
            department_id = getattr(request.user.user_profile, "department_id", None)
        elif request.user.is_authenticated and getattr(
            request.user, "department_id", None
        ):
            department_id = request.user.department_id
        else:
            dept = Department.objects.first()
            department_id = dept.id if dept else None

        from obs.models import EquipmentClass, SpareClass

        spare_classes = (
            SpareClass.objects.filter(department_id=department_id)
            if department_id
            else SpareClass.objects.all()
        )
        equipment_classes = EquipmentClass.objects.filter(spare_class__in=spare_classes)
        spare_obj = Spares.objects.filter(
            equipment_class__in=equipment_classes
        ).exclude(
            Q(authority__name="MO ITEM")
            | Q(authority__name="B & D")
            | Q(authority__name="MO ALLOWANCE")
        )
        spares_list = [
            {
                "pattern_number": s.pattern_number,
                "description": s.description or "",
                "name": s.description or "",
            }
            for s in spare_obj
        ]

        establishment_obj = MEstablishment.objects.filter(
            universal_id_m_establishmentcategory=1
        )
        organizations_obj = MMaterialOrganizations.objects.all()

        res_data = {
            "routines": routines_data,
            "selected_ids": selected_ids,
            "deferments": [
                {
                    "id": d.deferment_id,
                    "description": d.description,
                    "code": d.deferment_code,
                }
                for d in deferments
            ],
            "inabilities": inabilities,
            "reasons": [
                {
                    "id": r.reason_id,
                    "description": r.description,
                    "code": r.reason_code,
                    "universal_id_m_reason": r.universal_id_m_reason,
                }
                for r in reasons
            ],
            "spare_obj": spares_list,
            "establishment_obj": [
                {"id": e.est_id, "name": e.est_name} for e in establishment_obj
            ],
            "organizations_obj": [
                {"id": o.mo_id, "name": o.mo_name} for o in organizations_obj
            ],
        }

        serializer = MulRaiseFussResponseSerializer(res_data)
        return Response(serializer.data)

    @extend_schema(
        tags=["EMS"],
        request=MulRaiseFussRequestSerializer,
        responses={201: OpenApiTypes.OBJECT},
    )
    def post(self, request, *args, **kwargs):
        # Coerce empty strings to None for date and integer fields to prevent DRF validation failures
        req_data = (
            request.data.copy() if hasattr(request.data, "copy") else dict(request.data)
        )

        date_fields = [
            "fuss_date",
            "last_undertaken",
            "due_date",
            "schedule_date",
            "last_smp_completed",
            "last_amp_completed",
            "amp_smp_required_wef",
        ]
        for field in date_fields:
            if field in req_data and req_data[field] == "":
                req_data[field] = None

        int_fields = [
            "recomm_deferment",
            "reason",
            "inability",
            "mos_wed",
            "yard",
            "add_routine_id",
        ]
        for field in int_fields:
            if field in req_data and req_data[field] == "":
                req_data[field] = None

        serializer = MulRaiseFussRequestSerializer(data=req_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        selected_ids = data.get("selected_ids", [])
        add_routine_id = data.get("add_routine_id")

        if not selected_ids and add_routine_id:
            # Resolve all sub-routines linked to this master routine scheduler
            selected_ids = list(
                RoutineDescription.objects.filter(
                    add_routine_details_id=add_routine_id
                ).values_list("id", flat=True)
            )

        spares_data = data.get("spares_data", {})
        spares_json_str = data.get("spares_json")
        if not spares_data and spares_json_str:
            import json

            try:
                spares_data = json.loads(spares_json_str)
            except Exception:
                spares_data = {}

        ship = data.get("ship")
        department = data.get("department")
        serial_no = data.get("serial_no")

        fuss_date = data.get("fuss_date")
        last_undertaken = data.get("last_undertaken")
        due_date = data.get("due_date")
        schedule_date = data.get("schedule_date")

        equipment = data.get("equipment")
        location_on_board = data.get("location_on_board")
        equipment_sr_no = data.get("equipment_sr_no")
        location_code = data.get("location_code")

        maintop_no = data.get("maintop_no")
        frequency = data.get("frequency")
        amendment_no = data.get("amendment_no")
        new_equipment = data.get("new_equipment", False)

        deferment_id = data.get("recomm_deferment")
        inability_id = data.get("inability")
        reason_id = data.get("reason")

        mos_wed = data.get("mos_wed")
        yard = data.get("yard")

        last_smp_completed = data.get("last_smp_completed")
        last_smp_duration = data.get("last_smp_duration")

        last_amp_completed = data.get("last_amp_completed")
        last_amp_duration = data.get("last_amp_duration")

        amp_smp_required_wef = data.get("amp_smp_required_wef")
        amp_smp_duration = data.get("amp_smp_duration")

        remarks = data.get("remarks")
        demand_details = data.get("demand_details")

        deferment_obj = (
            MDeferment.objects.filter(deferment_id=deferment_id).first()
            if deferment_id
            else None
        )
        inability_obj = (
            MInability.objects.filter(inability_id=inability_id).first()
            if inability_id
            else None
        )
        reason_obj = (
            MReason.objects.filter(reason_id=reason_id).first() if reason_id else None
        )

        mo_wed_obj = (
            MMaterialOrganizations.objects.filter(mo_id=mos_wed).first()
            if mos_wed
            else None
        )
        establishment_obj = (
            MEstablishment.objects.filter(est_id=yard).first() if yard else None
        )

        dept_obj = None
        if request.user.is_authenticated and hasattr(request.user, "user_profile"):
            dept_obj = request.user.CustomUser_profile.department

        created_fusses = []
        for routine_id in selected_ids:
            routine = RoutineDescription.objects.filter(id=routine_id).first()
            if not routine:
                continue

            routine.is_fuss = True
            routine.save()

            fuss = FussRaiseDetails.objects.create(
                ship=ship,
                department=department,
                department_f_key=dept_obj,
                serial_no=serial_no,
                routine_description_id=routine,
                fuss_date=fuss_date,
                last_undertaken=last_undertaken,
                due_date=due_date,
                schedule_date=schedule_date,
                equipment=equipment,
                location_on_board=location_on_board,
                equipment_sr_no=equipment_sr_no,
                location_code=location_code,
                maintop_no=maintop_no,
                frequency=frequency,
                amendment_no=amendment_no,
                new_equipment=new_equipment,
                recomm_deferment=deferment_obj,
                inability=inability_obj,
                reason=reason_obj,
                mo_wed_f_key=mo_wed_obj,
                establishment_f_key=establishment_obj,
                mos_wed=mos_wed,
                yard=yard,
                last_smp_completed=last_smp_completed,
                last_smp_duration=last_smp_duration,
                last_amp_completed=last_amp_completed,
                last_amp_duration=last_amp_duration,
                amp_smp_required_wef=amp_smp_required_wef,
                amp_smp_duration=amp_smp_duration,
                remarks=remarks,
                demand_details=demand_details,
                sapres_required=len(spares_data) > 0,
            )
            created_fusses.append(fuss.id)

            for pattern, sdata in spares_data.items():
                if pattern:
                    FussSpare.objects.create(
                        fuss_f_ky=fuss,
                        pattern=pattern,
                        description=sdata.get("description", ""),
                        quantity=int(sdata.get("qty", 1)),
                    )

        return Response(
            {
                "status": "success",
                "message": "FUSS details saved successfully!",
                "data": {
                    "created_fuss_ids": created_fusses,
                },
            },
            status=status.HTTP_201_CREATED,
        )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="routine",
            type=str,
            description="Routine name to filter by",
            required=False,
        )
    ],
)
@api_view(["GET"])
def fuss_raised_details(request):
    routine = request.query_params.get("routine")
    fuss_list = FussRaiseDetails.objects.filter(isclosed_fuss=False).select_related(
        "routine_description_id",
        "routine_description_id__routine_name",
        "routine_description_id__equipment_name",
    )
    if routine:
        fuss_list = fuss_list.filter(routine_description_id__routine_name__name=routine)

    result = []
    for f in fuss_list:
        r = f.routine_description_id
        result.append(
            {
                "id": f.id,
                "ship": f.ship,
                "department": f.department,
                "serial_no": f.serial_no,
                "fuss_date": f.fuss_date.strftime("%Y-%m-%d") if f.fuss_date else "",
                "due_date": f.due_date.strftime("%Y-%m-%d") if f.due_date else "",
                "equipment": f.equipment,
                "maintop_no": f.maintop_no,
                "frequency": f.frequency,
                "routine_id": r.id if r else None,
                "routine_no": r.routine_no if r else "",
                "routine_name": r.routine_name.name if r and r.routine_name else "",
            }
        )
    return Response({"fuss_list": result})


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="year",
            type=str,
            description="Filter by year (e.g. 2026)",
            required=False,
        ),
        OpenApiParameter(
            name="month",
            type=str,
            description="Filter by month (e.g. 06)",
            required=False,
        ),
    ],
)
@api_view(["GET"])
def fuss_raised_routines(request):
    from collections import defaultdict

    today = date.today()
    fuss_raised_list = (
        FussRaiseDetails.objects.select_related(
            "routine_description_id",
            "routine_description_id__routine_name",
            "routine_description_id__equipment_name",
        )
        .order_by("-created_at")
        .filter(isclosed_fuss=False)
    )

    unique_departments = sorted(
        list(set(filter(None, fuss_raised_list.values_list("department", flat=True))))
    )
    unique_equipments = sorted(
        list(set(filter(None, fuss_raised_list.values_list("equipment", flat=True))))
    )
    unique_routine_types = sorted(
        list(
            set(
                filter(
                    None,
                    fuss_raised_list.values_list(
                        "routine_description_id__routine_name__name", flat=True
                    ),
                )
            )
        )
    )

    dept_equipment_map = defaultdict(list)
    for fuss in fuss_raised_list:
        if fuss.department and fuss.equipment:
            if fuss.equipment not in dept_equipment_map[fuss.department]:
                dept_equipment_map[fuss.department].append(fuss.equipment)
    dept_equipment_map = {
        dept: sorted(equips) for dept, equips in dept_equipment_map.items()
    }

    selected_year = request.query_params.get("year", "0000")
    selected_month = request.query_params.get("month", "00")

    result = []
    seen = set()
    for e in fuss_raised_list:
        routine = e.routine_description_id
        if not routine:
            continue

        unique_key = (
            e.department,
            e.equipment,
            routine.routine_name.name if routine.routine_name else "",
        )
        if unique_key in seen:
            continue
        seen.add(unique_key)

        next_due_date = None
        if e.last_undertaken and e.frequency:
            try:
                freq_months = int(float(e.frequency))
                next_due_date = e.last_undertaken + relativedelta(months=freq_months)
            except (ValueError, TypeError):
                next_due_date = None

        should_append = True
        if next_due_date:
            due_month = next_due_date.strftime("%m")
            due_year = next_due_date.strftime("%Y")
            if selected_month != "00" and selected_year != "0000":
                should_append = (
                    selected_month == due_month and selected_year == due_year
                )
            elif selected_month != "00" and selected_year == "0000":
                should_append = selected_month == due_month
            elif selected_month == "00" and selected_year != "0000":
                should_append = selected_year == due_year

        if should_append:
            total_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name, equipment_name=routine.equipment_name
            ).count()

            dyd_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=routine.equipment_name,
                by_whom="DYD",
            ).count()

            ss_routines = total_routines - dyd_routines

            routines_due = 0
            if next_due_date and next_due_date <= today:
                routines_due = 1

            result.append(
                {
                    "id": e.id,
                    "department": e.department,
                    "equipment": e.equipment,
                    "routine_name": (
                        routine.routine_name.name if routine.routine_name else ""
                    ),
                    "status": (
                        "Active" if e.due_date and e.due_date >= today else "Overdue"
                    ),
                    "maintop_no": e.maintop_no or "NA",
                    "routines_by_dyd": dyd_routines,
                    "routines_by_ss": ss_routines,
                    "total_routines": total_routines,
                    "routines_due": routines_due,
                    "due_date": (
                        next_due_date.strftime("%d %b, %Y") if next_due_date else "NA"
                    ),
                }
            )

    return Response(
        {
            "fuss_list": result,
            "departments": unique_departments,
            "equipments": unique_equipments,
            "routine_types": unique_routine_types,
            "dept_equipment_map": dept_equipment_map,
        }
    )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
)
class FussTriggerListAPIView(APIView):
    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated and hasattr(request, "user"):
            department = request.user.department
        else:
            department = Department.objects.first()

        today = date.today()
        current_time = make_aware(datetime.now())

        routine_qs = (
            AddRoutineDetails.objects.with_dashboard_relations().for_department(
                department
            )
        )

        result = []

        for routine in routine_qs:
            sub_routines = RoutineDescription.objects.for_add_routine(routine)

            total_sub = sub_routines.count()

            dyd_count = sub_routines.filter(by_whom="DYD").count()

            ss_count = total_sub - dyd_count

            if ss_count == 0:
                continue

            is_overdue = False
            due_date = None
            due_at_rh = None

            ss_sub_routines = sub_routines.exclude(by_whom="DYD")

            # Calendar based / Alternate periodic check
            if routine.last_routine_completion_date and routine.routine_category in [
                "CALENDAR BASED",
                "ALTERNATE PERIODIC",
            ]:
                due_date = routine.last_routine_completion_date + relativedelta(
                    months=routine.frequency_in_months
                )

                if hasattr(due_date, "date"):
                    due_date = due_date.date()

                if due_date < today:
                    is_overdue = True

            # Running hour based check
            dynamic_rhsi = routine.equipment_name.rhsi

            if (
                routine.last_routine_completion_atrunning_hrs is not None
                and routine.routine_category
                in [
                    "RUNNING HOUR BASED",
                    "ALTERNATE PERIODIC",
                ]
            ):
                due_at_rh = (
                    float(routine.last_routine_completion_atrunning_hrs)
                    + routine.frequency_in_hours
                )

                equipment = routine.equipment_name

                if equipment.state == "ACTIVE" and equipment.start_timedate:
                    start_time = equipment.start_timedate

                    if is_naive(start_time):
                        start_time = make_aware(start_time)

                    dynamic_rhsi += (current_time - start_time).total_seconds() / 3600

                if dynamic_rhsi is not None and dynamic_rhsi > due_at_rh:
                    is_overdue = True

            # Sub-routine due date check
            if not is_overdue:
                for sub_routine in ss_sub_routines:
                    if sub_routine.due_date:
                        sr_due = sub_routine.due_date

                        if hasattr(sr_due, "date"):
                            sr_due = sr_due.date()

                        if sr_due < today:
                            is_overdue = True
                            due_date = sr_due
                            break

            if is_overdue:
                first_sr = sub_routines.first()

                sr_ids = list(
                    sub_routines.values_list(
                        "id",
                        flat=True,
                    )
                )

                result.append(
                    {
                        "pk": (first_sr.id if first_sr else None),
                        "add_routine_pk": routine.pk,
                        "sr_ids": sr_ids,
                        "sub_dept": (
                            routine.equipment_name.sub_department.name
                            if routine.equipment_name.sub_department
                            else (
                                routine.equipment_name.section.name
                                if routine.equipment_name.section
                                else "NA"
                            )
                        ),
                        "equipment_nomenclature": (
                            routine.equipment_name.nomenclature
                            or routine.equipment_name.name
                        ),
                        "routine_name": (
                            routine.routine_name.name if routine.routine_name else ""
                        ),
                        "status": "Overdue",
                        "prev_completion_date": (
                            routine.last_routine_completion_date.strftime("%d %b %Y")
                            if routine.last_routine_completion_date
                            else "NA"
                        ),
                        "next_due_date": (
                            due_date.strftime("%d %b %Y") if due_date else "NA"
                        ),
                        "total_rh_updated_upto": (
                            routine.equipment_name.rhsi_updated_until.strftime(
                                "%d %b %Y"
                            )
                            if routine.equipment_name.rhsi_updated_until
                            else "NA"
                        ),
                        "rhsi": (
                            round(
                                routine.equipment_name.rhsi,
                                2,
                            )
                            if routine.equipment_name.rhsi is not None
                            else "NA"
                        ),
                        "routine_completed_at_rh": (
                            routine.last_routine_completion_atrunning_hrs
                            if routine.last_routine_completion_atrunning_hrs is not None
                            else "NA"
                        ),
                        "routine_due_at_rh": (
                            round(due_at_rh, 2) if due_at_rh is not None else "NA"
                        ),
                        "maintops_no": (first_sr.maintop_no if first_sr else "NA"),
                        "total_sub_subroutines": total_sub,
                        "ss_routines": ss_count,
                        "dyd_routines": dyd_count,
                    }
                )

        unique_departments = sorted(list({item["sub_dept"] for item in result}))

        unique_equipments = sorted(
            list({item["equipment_nomenclature"] for item in result})
        )

        unique_routines = sorted(list({item["routine_name"] for item in result}))

        return Response(
            {
                "result": result,
                "equipment_sections": unique_departments,
                "equipments_names": unique_equipments,
                "routine_names": unique_routines,
            }
        )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
)
class AberTriggerListAPIView(APIView):
    def get(self, request, *args, **kwargs):
        department = (
            getattr(request.user, "department", None)
            if hasattr(request, "user") and request.user.is_authenticated
            else None
        )
        if not department and hasattr(request.user, "user_profile"):
            department = getattr(request.user.user_profile, "department", None)
        if not department:
            department = Department.objects.first()

        today = date.today()

        equip_qs = ShipEquipment.objects.filter(
            section_f_key__department=department,
            installation_date__isnull=False,
        ).select_related(
            "equipment",
            "section_f_key",
        )

        result = []

        for equipment in equip_qs:
            diff = relativedelta(
                today,
                equipment.installation_date,
            )

            years = diff.years

            status = "Red" if years >= 7 else "Green"

            result.append(
                {
                    "pk": equipment.pk,
                    "insma_code": (
                        equipment.equipment.equipment_code
                        if equipment.equipment
                        else "NA"
                    ),
                    "nomenclature": (equipment.nomenclature or "NA"),
                    "compartment": (equipment.compartment or "NA"),
                    "installation_date": (
                        equipment.installation_date.strftime("%d %b %Y")
                    ),
                    "years_since": years,
                    "status": status,
                }
            )

        unique_nomenclatures = sorted(list({item["nomenclature"] for item in result}))

        unique_compartments = sorted(list({item["compartment"] for item in result}))

        return Response(
            {
                "result": result,
                "nomenclatures": unique_nomenclatures,
                "compartments": unique_compartments,
            }
        )


# ==================== ROUTINE HISTORY VIEWS ====================


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def routine_history(request):
    department = (
        getattr(request.user, "department", None)
        if hasattr(request, "user") and request.user.is_authenticated
        else None
    )
    if not department and hasattr(request.user, "user_profile"):
        department = getattr(request.user.user_profile, "department", None)
    if not department:
        department = Department.objects.first()

    today = date.today()
    current_time = make_aware(datetime.now())

    routine_qs = AddRoutineDetails.objects.with_dashboard_relations().for_department(
        department
    )

    result = []
    for r in routine_qs:
        sub_routines = RoutineDescription.objects.for_add_routine(r)
        total_sub = sub_routines.count()
        dyd_count = sub_routines.filter(by_whom="DYD").count()
        ss_count = total_sub - dyd_count

        completions = CompletedRoutine.objects.filter(
            routine__in=sub_routines
        ).ordered_by_completion()
        latest_completion = completions.first()

        prev_date_str = (
            latest_completion.date_of_completion.strftime("%d %b %Y")
            if latest_completion and latest_completion.date_of_completion
            else "NA"
        )
        completed_at_rh = (
            latest_completion.running_hour
            if latest_completion and latest_completion.running_hour
            else "NA"
        )

        is_overdue = False
        due_date = None
        due_at_rh = None

        if r.last_routine_completion_date and r.routine_category in [
            "CALENDAR BASED",
            "ALTERNATE PERIODIC",
        ]:
            due_date = r.last_routine_completion_date + relativedelta(
                months=r.frequency_in_months
            )
            if hasattr(due_date, "date"):
                due_date = due_date.date()
            if due_date < today:
                is_overdue = True

        dynamic_rhsi = r.equipment_name.rhsi
        if (
            r.last_routine_completion_atrunning_hrs is not None
            and r.routine_category in ["RUNNING HOUR BASED", "ALTERNATE PERIODIC"]
        ):
            due_at_rh = (
                float(r.last_routine_completion_atrunning_hrs) + r.frequency_in_hours
            )
            eq = r.equipment_name
            if eq.state == "ACTIVE" and eq.start_timedate:
                st_time = eq.start_timedate
                if is_naive(st_time):
                    st_time = make_aware(st_time)
                dynamic_rhsi += (current_time - st_time).total_seconds() / 3600

            if dynamic_rhsi is not None and dynamic_rhsi > due_at_rh:
                is_overdue = True

        status = "Overdue" if is_overdue else "Active"

        result.append(
            {
                "pk": r.pk,
                "sub_dept": (
                    r.equipment_name.sub_department.name
                    if r.equipment_name.sub_department
                    else (
                        r.equipment_name.section.name
                        if r.equipment_name.section
                        else "NA"
                    )
                ),
                "equipment_nomenclature": r.equipment_name.nomenclature
                or r.equipment_name.name,
                "routine_name": r.routine_name.name if r.routine_name else "",
                "status": status,
                "prev_completion_date": prev_date_str,
                "next_due_date": due_date.strftime("%d %b %Y") if due_date else "NA",
                "total_rh_updated_upto": (
                    r.equipment_name.rhsi_updated_until.strftime("%d %b %Y")
                    if r.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "rhsi": (
                    round(r.equipment_name.rhsi, 2)
                    if r.equipment_name.rhsi is not None
                    else "NA"
                ),
                "routine_completed_at_rh": (
                    completed_at_rh if r.routine_category != "CALENDAR BASED" else "NA"
                ),
                "routine_due_at_rh": (
                    round(due_at_rh, 2)
                    if due_at_rh is not None and r.routine_category != "CALENDAR BASED"
                    else "NA"
                ),
                "maintops_no": r.maintop_no or "NA",
                "ss_routines": ss_count,
                "dyd_routines": dyd_count,
                "total_sub_subroutines": total_sub,
            }
        )

    unique_departments = sorted(list(set([r["sub_dept"] for r in result])))
    unique_equipments = sorted(list(set([r["equipment_nomenclature"] for r in result])))
    unique_routines = sorted(list(set([r["routine_name"] for r in result])))

    return Response(
        {
            "result": result,
            "departments": unique_departments,
            "equipments": unique_equipments,
            "routine_types": unique_routines,
        }
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def equipment_running_history(request):
    department = getattr(request.user, "department", None)
    if not department and hasattr(request.user, "user_profile"):
        department = getattr(request.user.user_profile, "department", None)
    if not department:
        department = Department.objects.first()

    history_qs = (
        PostEquipmentStateChangeHistorySave.objects.filter(
            equipment_name__section__department=department
        )
        .select_related("equipment_name", "equipment_name__section")
        .defer("equipment_name__universal_id_t_equipment_ship_detail")
        .order_by("-entry_creation_date")
    )

    result = []
    for h in history_qs:
        created_by_str = "NA"
        start_time = h.start_time.astimezone(local_tz) if h.start_time else None
        stop_time = h.stop_time.astimezone(local_tz) if h.stop_time else None
        entry_date = (
            h.entry_creation_date.astimezone(local_tz)
            if h.entry_creation_date
            else None
        )

        result.append(
            {
                "hours_run": (
                    round(h.diff_in_hours, 1) if h.diff_in_hours is not None else 0
                ),
                "month": h.month_name or "NA",
                "equipment_nomenclature": h.equipment_name.nomenclature
                or h.equipment_name.name,
                "start_location": h.started_at_location or "NA",
                "start_time": (
                    start_time.strftime("%d %b %Y - %H:%M") if start_time else "NA"
                ),
                "stop_time": (
                    stop_time.strftime("%d %b %Y - %H:%M") if stop_time else "NA"
                ),
                "entry_created_on": (
                    entry_date.strftime("%d %b %Y - %H:%M") if entry_date else "NA"
                ),
                "created_by": created_by_str,
            }
        )

    unique_equipments = sorted(list(set([r["equipment_nomenclature"] for r in result])))
    unique_months = sorted(list(set([r["month"] for r in result])))

    return Response(
        {
            "result": result,
            "equipments": unique_equipments,
            "months": unique_months,
        }
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def slip_history(request):
    history_slip = []
    for entry in (
        PostCalculateLPC.objects.select_related("gt_name")
        .defer("gt_name__universal_id_t_equipment_ship_detail")
        .order_by("-calculation_time")
    ):
        calculation_time = (
            entry.calculation_time.astimezone(local_tz)
            if entry.calculation_time
            else None
        )
        history_slip.append(
            {
                "HPC": entry.at_hpc_rpm,
                "gt_name": entry.gt_name.name if entry.gt_name else "NA",
                "LPC": entry.recorded_lpc,
                "AirPr": entry.recorded_air_pr_after_hpc,
                "ExtTemp": entry.recorded_ext_temp,
                "date": calculation_time.date() if calculation_time else "NA",
                "lpc_slip": entry.calculated_lpc_slip,
                "air_slip": entry.calculated_air_slip,
                "ext_slip": entry.calculated_ext_slip,
            }
        )

    history_slip_gtg = []
    for entry in (
        PostCalculateGTG.objects.select_related("gt_name")
        .defer("gt_name__universal_id_t_equipment_ship_detail")
        .order_by("-calculation_time")
    ):
        calculation_time = (
            entry.calculation_time.astimezone(local_tz)
            if entry.calculation_time
            else None
        )
        history_slip_gtg.append(
            {
                "gt_name": entry.gt_name.name if entry.gt_name else "NA",
                "el_load": entry.recorded_el_load,
                "ext_temp": entry.recorded_ext_temp_gtg,
                "amb_temp": entry.recorded_amb_temp,
                "gtg_slip": entry.calculated_gtg_slip_ext,
                "date": calculation_time.date() if calculation_time else "NA",
            }
        )

    return Response(
        {
            "activate": "slip",
            "history_slip": history_slip,
            "history_slip_gtg": history_slip_gtg,
        }
    )


def get_maintenance_date_range(period):
    today = date.today()
    if period == "weekly":
        start = today - timedelta(days=today.weekday())
        return start, start + timedelta(days=5), "Weekly Maintenance Plan"
    if period == "fortnightly":
        start = today - timedelta(days=today.weekday())
        return start, start + timedelta(days=13), "Fortnightly Maintenance Plan"
    if period == "monthly":
        start = today.replace(day=1)
        next_month = (today.replace(day=28) + timedelta(days=4)).replace(day=1)
        return start, next_month - timedelta(days=1), "Monthly Maintenance Plan"
    if period == "six_monthly":
        start = today.replace(day=1)
        return (
            start,
            start + relativedelta(months=6) - timedelta(days=1),
            "Six Monthly Maintenance Plan",
        )
    if period == "annual":
        return (
            today.replace(month=1, day=1),
            today.replace(month=12, day=31),
            "Annual Maintenance Plan",
        )
    return today, today, "Maintenance Plan"


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="sub_dept",
            type=int,
            description="Optional Sub Department ID filter",
            required=False,
        )
    ],
)
@api_view(["GET"])
def maintenance_plan(request, period="weekly"):
    valid_periods = {"weekly", "fortnightly", "monthly", "six_monthly", "annual"}
    if period not in valid_periods:
        return Response(
            {"detail": "Invalid maintenance plan period."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    start_date, end_date, title = get_maintenance_date_range(period)
    sub_dept_id = request.query_params.get("sub_dept")

    sub_depts = SubDepartment.objects.all().order_by("name")
    base_qs = (
        RoutineDescription.objects.filter(due_date__range=[start_date, end_date])
        .select_related(
            "equipment_name",
            "routine_name",
            "equipment_name__sub_department",
            "add_routine_details",
            "add_routine_details__frequency_f_key",
        )
        .order_by("due_date", "equipment_name__name")
    )

    if sub_dept_id:
        base_qs = base_qs.filter(equipment_name__sub_department_id=sub_dept_id)

    def serialize_routine(routine):
        add_details = routine.add_routine_details
        frequency = add_details.frequency_f_key if add_details else None
        prefix = frequency.frequency_prefix if frequency else ""
        equipment = routine.equipment_name
        sub_dept = equipment.sub_department if equipment else None

        return {
            "id": f"rd_{routine.id}",
            "dart_no": routine.dart_number or "NA",
            "equipment": equipment.name if equipment else "NA",
            "nomenclature": (
                equipment.nomenclature or equipment.name if equipment else "NA"
            ),
            "sub_routine": routine.routine_name.name if routine.routine_name else "NA",
            "routine_no": routine.routine_no or "NA",
            "routine_no_combined": (
                f"{prefix}{routine.routine_no}" if routine.routine_no else "NA"
            ),
            "description": routine.routine_description,
            "due_date": routine.due_date,
            "by_whom": routine.by_whom or "NA",
            "sub_dept": sub_dept.name if sub_dept else "NA",
            "category": add_details.routine_category if add_details else "NA",
        }

    all_pending = [serialize_routine(r) for r in base_qs.filter(is_close=False)]
    all_completed = [serialize_routine(r) for r in base_qs.filter(is_close=True)]

    selected_sub_dept_name = "All Sub-Departments"
    if sub_dept_id:
        selected_sub_dept = SubDepartment.objects.filter(id=sub_dept_id).first()
        if selected_sub_dept:
            selected_sub_dept_name = selected_sub_dept.name

    return Response(
        {
            "title": title,
            "start_date": start_date,
            "end_date": end_date,
            "sub_depts": list(sub_depts.values("id", "name")),
            "selected_sub_dept": int(sub_dept_id) if sub_dept_id else None,
            "selected_sub_dept_name": selected_sub_dept_name,
            "all_pending": all_pending,
            "all_completed": all_completed,
            "total_pending": len(all_pending),
            "total_completed": len(all_completed),
            "period": period,
        }
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def routine_history_timeline_data(request, pk):
    r = get_object_or_404(AddRoutineDetails, pk=pk)
    sub_routines = RoutineDescription.objects.filter(add_routine_details=r)
    completions = (
        CompletedRoutine.objects.filter(routine__in=sub_routines)
        .select_related("routine")
        .order_by("-date_of_completion")
    )

    is_calendar = r.routine_category in ["CALENDAR BASED"]

    next_due_date_str = "NA"
    if is_calendar and r.last_routine_completion_date and r.frequency_in_months:
        nd = r.last_routine_completion_date + relativedelta(
            months=r.frequency_in_months
        )
        if hasattr(nd, "date"):
            nd = nd.date()
        next_due_date_str = nd.strftime("%d %b %Y")

    due_at_rh = "NA"
    if (
        not is_calendar
        and r.last_routine_completion_atrunning_hrs is not None
        and r.frequency_in_hours
    ):
        due_at_rh = round(
            float(r.last_routine_completion_atrunning_hrs) + r.frequency_in_hours, 2
        )

    eq = r.equipment_name
    rhsi_val = round(eq.rhsi, 2) if eq.rhsi is not None else "NA"
    rhsi_updated_upto_str = (
        eq.rhsi_updated_until.strftime("%d %b %Y") if eq.rhsi_updated_until else "NA"
    )

    data = []
    for cr in completions:
        prev_date_str = (
            cr.date_of_completion.strftime("%d %b %Y")
            if cr.date_of_completion
            else "NA"
        )
        completed_at_rh = cr.running_hour if not is_calendar else "NA"

        data.append(
            {
                "dart_no": cr.new_dart_number or "NA",
                "routine_no": cr.routine.routine_no or "NA",
                "routine_description": cr.routine.routine_description or "NA",
                "next_due_date": next_due_date_str if is_calendar else "NA",
                "prev_completion_date": prev_date_str,
                "rhsi": rhsi_val if not is_calendar else "NA",
                "rhsi_updated_upto": rhsi_updated_upto_str if not is_calendar else "NA",
                "routine_completed_at_rh": completed_at_rh,
                "routine_due_at_rh": due_at_rh if not is_calendar else "NA",
                "id": cr.id,
                "routine_pk": cr.routine.pk,
            }
        )

    routine_nos = sorted(list(set([d["routine_no"] for d in data])))

    return Response({"data": data, "routine_nos": routine_nos})


# ==================== SEARCH / PLAN VIEWS ====================


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
)
class SearchResultRunningHoursBasedLessThanOneKAPIView(APIView):
    def get(self, request, *args, **kwargs):
        department = request.user.department

        # Step 1: Get all running hour based routines
        routine_qs = AddRoutineDetails.objects.filter(
            routine_category__in=["RUNNING HOUR BASED"],
            equipment_name__section__department=department,
        ).select_related(
            "equipment_name",
            "routine_name",
            "equipment_name__section",
        )

        # Step 2: Filter routines where remaining running hours <= 1000
        routines = []

        for routine in routine_qs:
            last_rhrs = routine.last_routine_completion_atrunning_hrs
            rhsi = routine.equipment_name.rhsi

            if last_rhrs is not None and rhsi is not None:
                next_due_rhrs = last_rhrs + routine.frequency_in_hours

                running_hrs_available = round(
                    next_due_rhrs - rhsi,
                    1,
                )

                if running_hrs_available <= 1000:
                    routine._next_due_rhrs = next_due_rhrs
                    routine._running_hrs_available = running_hrs_available
                    routines.append(routine)

        # Step 3: Return message if no routines matched
        if not routines:
            return Response(
                {
                    "title": "EMS | Search Result",
                    "message": ("No matching result found."),
                    "count": 0,
                    "result": [],
                }
            )

        # Step 4: Prepare response
        result = []

        for routine in routines:
            eq = routine.equipment_name

            routine_desc = RoutineDescription.objects.filter(
                add_routine_details=routine.id
            ).first()

            total_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=eq,
            ).count()

            dyd_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=eq,
                by_whom="DYD",
            ).count()

            ss_routines = total_routines - dyd_routines

            result.append(
                {
                    "pk": routine.pk,
                    "routine_name": routine.routine_name.name,
                    "section": eq.section.name,
                    "equipment_name": eq.name,
                    "maintop_no": (routine_desc.maintop_no if routine_desc else "NA"),
                    "last_routine_running_hrs": (
                        routine.last_routine_completion_atrunning_hrs
                        if routine.last_routine_completion_atrunning_hrs is not None
                        else "NA"
                    ),
                    "next_due_running_hrs": round(
                        routine._next_due_rhrs,
                        1,
                    ),
                    "total_running_hrs": (
                        round(eq.rhsi, 1) if eq.rhsi is not None else "NA"
                    ),
                    "running_hrs_updated_tilldate": (
                        f"{eq.rhsi_updated_until.day} {eq.rhsi_updated_until.strftime('%b, %Y')}"
                        if eq.rhsi_updated_until
                        else "NA"
                    ),
                    "running_hrs_available": routine._running_hrs_available,
                    "total_routines": total_routines,
                    "ss_routines": ss_routines,
                    "dyd_routines": dyd_routines,
                    "last_routine_date": (
                        f"{routine.last_routine_completion_date.day} "
                        f"{routine.last_routine_completion_date.strftime('%b, %Y')}"
                        if routine.last_routine_completion_date
                        else "NA"
                    ),
                    "remarks": routine.remarks,
                }
            )

        return Response(
            {
                "title": "EMS | Search Result",
                "count": len(result),
                "result": result,
            }
        )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
)
class SearchResultCalenderBasedLessThanSixMonthsAPIView(APIView):
    def get(self, request, *args, **kwargs):
        local_tz = ZoneInfo("Asia/Kolkata")

        current_datetime = datetime.now(timezone.utc).astimezone(local_tz)
        date_time_after_six_months = current_datetime + relativedelta(months=6)

        department = (
            getattr(request.user, "department", None)
            if hasattr(request, "user") and request.user.is_authenticated
            else None
        )
        if not department:
            department = Department.objects.first()

        # Get Calendar Based routines
        routine_frequency1 = AddRoutineDetails.objects.filter(
            routine_category="CALENDAR BASED",
            equipment_name__section__department=department,
        ).values_list("id", flat=True)

        # Get Alternate Periodic routines
        routine_frequency2 = AddRoutineDetails.objects.filter(
            routine_category="ALTERNATE PERIODIC",
            equipment_name__section__department=department,
        ).values_list("id", flat=True)

        routine_frequency = list(routine_frequency1) + list(routine_frequency2)

        routines_all = AddRoutineDetails.objects.filter(
            id__in=routine_frequency
        ).select_related(
            "routine_name",
            "equipment_name",
            "equipment_name__section",
            "equipment_name__sub_department",
        )

        routines = []

        # Filter routines due within six months
        for routine in routines_all:
            next_due_date = None

            if routine.last_routine_completion_date and routine.routine_category in [
                "CALENDAR BASED",
                "ALTERNATE PERIODIC",
            ]:
                next_due_date = routine.last_routine_completion_date + timedelta(
                    days=routine.frequency_in_months * 30
                )

                if next_due_date <= date_time_after_six_months:
                    routine._next_due_date = next_due_date
                    routines.append(routine)

        # No matching routines
        if not routines:
            return Response(
                {
                    "title": "EMS | Search Result",
                    "message": ("No matching result found."),
                    "count": 0,
                    "result": [],
                }
            )

        result = []

        for routine in routines:
            total_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=routine.equipment_name,
            ).count()

            dyd_routines = RoutineDescription.objects.filter(
                routine_name=routine.routine_name,
                equipment_name=routine.equipment_name,
                by_whom="DYD",
            ).count()

            ss_routines = total_routines - dyd_routines

            routine_other_details = RoutineDescription.objects.filter(
                add_routine_details=routine.id
            ).first()

            last_routine_completion_dt = routine.last_routine_completion_date
            last_routine_completion_date_formatted = (
                f"{last_routine_completion_dt.day} "
                + f"{last_routine_completion_dt.strftime('%b, %Y')}"
            )

            _next_due_date_dt = routine._next_due_date
            _next_due_date_formatted = (
                f"{_next_due_date_dt.day} {_next_due_date_dt.strftime('%b, %Y')}"
            )

            result.append(
                {
                    "pk": routine.pk,
                    "sub_department": (
                        routine.equipment_name.sub_department.name
                        if routine.equipment_name
                        and routine.equipment_name.sub_department
                        else "NA"
                    ),
                    "equipment_nomenclature": (
                        routine.nomenclature
                        if routine.nomenclature and hasattr(routine, "nomenclature")
                        else "NA"
                    ),
                    "routine_name": (routine.routine_name.name),
                    "section": (routine.equipment_name.section.name),
                    "equipment_name": (routine.equipment_name.name),
                    "routine_status": "Active",
                    "total_ss_routines": ss_routines,
                    "total_dyd_routines": dyd_routines,
                    "last_routine_completion_date": last_routine_completion_date_formatted,
                    "next_due_date": _next_due_date_formatted,
                    "maintop_no": (
                        routine_other_details.maintop_no
                        if routine_other_details
                        else "NA"
                    ),
                    "total_sub_routines": total_routines,
                }
            )

        return Response(
            {
                "title": "EMS | Search Result",
                "count": len(result),
                "result": result,
            }
        )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def SearchDetailView_plan(request, pk=None):
    if not pk:
        return Response({"result": []})

    routine_detail = get_object_or_404(AddRoutineDetails, pk=pk)
    next_due_date = None

    category = (
        routine_detail.routine_category.upper().strip()
        if routine_detail.routine_category
        else ""
    )

    if category in ["CALENDAR BASED", "ALTERNATE PERIODIC"]:
        last_date = routine_detail.last_routine_completion_date
        freq_months = routine_detail.frequency_in_months or 0
        if last_date and freq_months:
            next_due_date = last_date + relativedelta(months=freq_months)
            if isinstance(next_due_date, datetime):
                next_due_date = next_due_date.date()

    row_style = ""

    routine_obj = PlannedRoutineDescription.objects.filter(
        routine_description_id__add_routine_details__id=pk, is_deleted=False
    ).values_list("routine_description_id__id", flat=True)

    routine_descriptions = RoutineDescription.objects.filter(
        id__in=routine_obj,
        equipment_name=routine_detail.equipment_name,
        routine_name=routine_detail.routine_name,
    ).order_by("routine_no")

    result = []
    for e in routine_descriptions:
        data1 = PlannedRoutineDescription.objects.filter(
            routine_description_id__id=e.id, is_deleted=False
        ).first()
        result.append(
            {
                "pk": e.pk,
                "routine_name": e.routine_name.name if e.routine_name else "",
                "equipment_name": e.equipment_name.name if e.equipment_name else "",
                "maintop_no": e.maintop_no,
                "dart_number": e.dart_number,
                "routine_description": e.routine_description,
                "routine_no": e.routine_no,
                "planned_commencement_date": (
                    data1.planned_commencement_date.strftime("%Y-%m-%d")
                    if data1 and data1.planned_commencement_date
                    else "NA"
                ),
                "rhsi": (
                    e.equipment_name.rhsi if e.equipment_name.rhsi is not None else "NA"
                ),
                "rhsi_updated_until": (
                    e.equipment_name.rhsi_updated_until.strftime("%Y-%m-%d")
                    if e.equipment_name.rhsi_updated_until
                    else "NA"
                ),
                "previous_routine_completed_date": (
                    e.previous_completed_date.strftime("%Y-%m-%d")
                    if e.previous_completed_date
                    else ""
                ),
                "due_date": e.due_date.strftime("%Y-%m-%d") if e.due_date else "",
                "due_at_rh": e.due_at_rh if e.due_at_rh is not None else "NA",
                "previous_completed_at_rh": (
                    e.previous_completed_at_rh
                    if e.previous_completed_at_rh is not None
                    else "NA"
                ),
                "action_by": e.by_whom,
                "spare_used": False,
                "row_style": row_style,
                "spare_req": "YES" if data1 and data1.spares_required else "NO",
                "category_data": routine_detail.routine_category,
            }
        )

    return Response({"result": result})


# ==================== FUSS COMPLETE & CLOSE VIEWS ====================


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def close_raise_fuss(request, pk):
    planned_routine_obj = get_object_or_404(RoutineDescription, pk=pk)
    planned_obj = PlannedRoutineDescription.objects.filter(
        routine_description_id=planned_routine_obj.pk
    ).first()

    spares_for_routine = []
    if planned_obj:
        spares_qs = PlannedRoutineSpareList.objects.filter(
            planned_routine_description=planned_obj.id
        )
        spares_for_routine = [
            {
                "pattern_number": s.pattern_number,
                "quantity_required": s.quantity_required,
            }
            for s in spares_qs
        ]

    issues_qs = Issue.objects.filter(dart_number=planned_routine_obj.dart_number)
    issue_list = [
        {
            "id": issue.id,
            "voucher_no": issue.voucher_no,
            "date": issue.date.strftime("%Y-%m-%d") if issue.date else "",
        }
        for issue in issues_qs
    ]

    fullname = ""
    rankname = ""
    rank_obj_list = []
    if request.user.is_authenticated and hasattr(request.user, "user_profile"):
        user_profile = request.user.CustomUser_profile
        fullname = (
            f"{user_profile.firstname or ''} {user_profile.lastname or ''}".strip()
        )
        if user_profile.rank:
            rankname = user_profile.rank.name

        rank_qs = (
            MRanklist.objects.filter(
                universal_id_m_department=user_profile.department.universal_id_m_department
            )
            if user_profile.department
            else MRanklist.objects.all()
        )
        rank_obj_list = [{"id": r.rankid, "name": r.name} for r in rank_qs]

    return Response(
        {
            "planned_routine": {
                "pk": planned_routine_obj.pk,
                "maintop_no": planned_routine_obj.maintop_no,
                "dart_number": planned_routine_obj.dart_number,
                "routine_no": planned_routine_obj.routine_no,
                "routine_description": planned_routine_obj.routine_description,
            },
            "planned_obj": (
                {
                    "pk": planned_obj.pk if planned_obj else None,
                    "planned_commencement_date": (
                        planned_obj.planned_commencement_date.strftime("%Y-%m-%d")
                        if planned_obj and planned_obj.planned_commencement_date
                        else None
                    ),
                }
                if planned_obj
                else None
            ),
            "spares_for_routine": spares_for_routine,
            "issue_list": issue_list,
            "rankname": rankname,
            "fullname": fullname,
            "rank_obj": rank_obj_list,
        }
    )


@extend_schema(
    tags=["EMS"],
    request=CompleteFussRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def complete_fuss(request, pk):
    routine = get_object_or_404(RoutineDescription, id=pk)
    serializer = CompleteFussRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    dep_name = (
        getattr(request.user.CustomUser_profile.department, "name", None)
        if (request.user.is_authenticated and hasattr(request.user, "user_profile"))
        else None
    )
    first_letter = dep_name[0].upper() if dep_name else "X"
    prefix = f"M-{first_letter}-"

    import random

    random_number = f"{random.randint(0, 99999):05d}"
    dart_number = f"{prefix}{random_number}"

    rank_id = data.get("rank")
    rank_other = data.get("other_rank")
    rank_obj = MRanklist.objects.filter(rankid=rank_id).first() if rank_id else None

    custom_rank = rank_other if str(rank_id) in ["43", "44", "45", "46"] else None

    completed_routine = CompletedRoutine.objects.create(
        routine=routine,
        date_of_completion=data.get("complete_date") or None,
        hours=None,
        minutes=None,
        carried_by=data.get("carried_by"),
        p_no=data.get("p_no"),
        rank=rank_obj,
        other_rank=custom_rank,
        total_manpower=data.get("total_manpower") or None,
        running_hour=data.get("running_hour"),
        due_running_hour=data.get("due_running_hour"),
        repair_remark=data.get("remarks"),
        completion_details=data.get("completion_details"),
        trial_team=bool(data.get("trial_team")),
        rec_for_deletion=False,
        old_dart_number=None,
        new_dart_number=dart_number,
        not_applicable=False,
        rec_deletion=False,
        isfuss_close=True,
    )

    FussRaiseDetails.objects.filter(routine_description_id=routine).update(
        isclosed_fuss=True
    )

    RoutineDescription.objects.filter(id=pk).update(dart_number=dart_number)
    routine_des = RoutineDescription.objects.filter(id=pk).first()

    if routine_des and routine_des.add_routine_details:
        routine_del = AddRoutineDetails.objects.filter(
            id=routine_des.add_routine_details.id
        ).first()
        last_completion_date = data.get("complete_date")

        if last_completion_date and routine_del:
            routine_des_update = RoutineDescription.objects.filter(id=pk)
            routine_des_update.update(
                last_routine_completion_date=last_completion_date,
                last_routine_completion_atrunning_hrs=data.get("running_hour") or None,
            )

        planned_routine = PlannedRoutineDescription.objects.filter(
            routine_description_id=routine_des.id
        ).first()

        if planned_routine:
            planned_routine.is_deleted = True
            planned_routine.save()
            PlannedRoutineSpareList.objects.filter(
                planned_routine_description_id=planned_routine.id
            ).update(is_deleted=True)

    spare_list = data.get("spares", [])
    for spare in spare_list:
        CompletedRoutineSpare.objects.create(
            completed_routine=completed_routine, spare_name=spare
        )

    return Response({"success": True, "message": "FUSS closed successfully!"})


# ==================== SPARES / GENERAL VIEWS ====================


@extend_schema(
    tags=["EMS"],
    request=SaveOemSpareRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def save_oem_spare(request):
    serializer = SaveOemSpareRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    from obs.models import Spares

    authority = Authority.objects.filter(name="PIL").first()
    equipment_class = EquipmentClass.objects.filter(name="SPARES PIL").first()

    deno_id = data.get("denomination_id")
    if not deno_id:
        return Response(
            {"success": False, "error": "Please select a denomination"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    spare = Spares.objects.create(
        pattern_number=data.get("pattern_number", "").upper(),
        description=data.get("description", "").upper(),
        authority_id=authority.id if authority else None,
        equipment_class_id=equipment_class.id if equipment_class else None,
        denomination_id=deno_id,
        critical=False,
        quantity_authorised=0,
        quantity_available=0,
    )

    return Response(
        {
            "success": True,
            "pattern_number": spare.pattern_number,
            "description": spare.description,
            "denomination_name": spare.denomination.name if spare.denomination else "",
        }
    )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def closed_routines_list(request):
    completed_qs = CompletedRoutine.objects.select_related(
        "routine", "routine__equipment_name", "routine__routine_name"
    ).order_by("-created_at")

    data = []
    for cr in completed_qs:
        routine = cr.routine
        if not routine:
            continue
        data.append(
            {
                "maintop_no": routine.maintop_no,
                "dart_no": routine.dart_number,
                "equipment_name": (
                    routine.equipment_name.name if routine.equipment_name else ""
                ),
                "status": "CLOSED",
                "routine_name": (
                    routine.routine_name.name if routine.routine_name else ""
                ),
                "routine_no": routine.routine_no,
                "spare_requested": (
                    routine.spare_requested
                    if hasattr(routine, "spare_requested")
                    else False
                ),
                "routine_description": routine.routine_description,
                "action_by": routine.by_whom,
                "routine_id": routine.id,
            }
        )

    return Response({"data": data})


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def history_close_routine(request):
    return closed_routines_list(request._request)


# ==================== ROUTINE INITIALIZATION VIEWS ====================


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="sub_department_id",
            type=str,
            description="Sub Department ID or 'all'",
            required=False,
        )
    ],
)
@api_view(["GET"])
def get_routine_initialization_data(request):
    sub_dept_id = request.query_params.get("sub_department_id", "all")

    dept = None
    if request.user.is_authenticated and hasattr(request.user, "user_profile"):
        dept = getattr(request.user.CustomUser_profile, "department", None)

    siblings_rh = (
        RoutineDescription.objects.filter(
            is_close=False, previous_completed_date__isnull=True
        )
        .filter(Q(previous_completed_at_rh__isnull=True))
        .order_by("routine_no")
        .values_list("add_routine_details_id", flat=True)
    )

    siblings_calendar = (
        RoutineDescription.objects.filter(
            is_close=False, previous_completed_date__isnull=True
        )
        .filter(
            Q(previous_completed_at_rh__exact="")
            | Q(previous_completed_at_rh__isnull=True)
        )
        .order_by("routine_no")
        .values_list("add_routine_details_id", flat=True)
    )

    routines = AddRoutineDetails.objects.filter(
        Q(id__in=siblings_rh, routine_category="RUNNING HOUR BASED")
        | Q(id__in=siblings_calendar, routine_category="CALENDAR BASED"),
    )
    if dept:
        routines = routines.filter(equipment_name__sub_department__department_name=dept)

    if sub_dept_id != "all":
        routines = routines.filter(equipment_name__sub_department_id=sub_dept_id)

    routines = routines.select_related(
        "equipment_name__sub_department", "equipment_name", "routine_name"
    ).exclude(Q(frequency__endswith="R"))

    data = []
    for r in routines:
        equipment_nomenclature = (
            r.equipment_name.nomenclature if r.equipment_name else None
        )
        if not equipment_nomenclature or equipment_nomenclature.strip() == "":
            equipment_nomenclature = (
                r.equipment_name.name
                if r.equipment_name and r.equipment_name.name
                else "-"
            )

        data.append(
            {
                "pk": r.pk,
                "sub_department": (
                    r.equipment_name.sub_department.name
                    if r.equipment_name and r.equipment_name.sub_department
                    else "-"
                ),
                "equipment_nomenclature": equipment_nomenclature,
                "routine_type": r.routine_name.name if r.routine_name else "-",
                "completion_date": "",
                "undertaken_rh": "",
            }
        )

    return Response({"data": data})


@extend_schema(
    tags=["EMS"],
    request=SaveRoutineInitBulkRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def save_routine_init_bulk(request):
    try:
        serializer = SaveRoutineInitBulkRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        pks = data.get("pks", [])
        completion_date = data.get("completion_date")
        undertaken_rh = data.get("undertaken_rh")

        if not pks:
            return Response(
                {"status": "error", "message": "No records selected"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        parsed_date = pd.to_datetime(completion_date) if completion_date else None

        RoutineDescription.objects.filter(pk__in=pks).update(
            previous_completed_date=parsed_date,
            previous_completed_at_rh=undertaken_rh if undertaken_rh else None,
        )

        return Response(
            {
                "status": "success",
                "message": f"Successfully updated {len(pks)} records.",
            }
        )
    except Exception as e:
        return Response(
            {"status": "error", "message": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
class UniqueMaintopAPIView(APIView):
    def get(self, request, *args, **kwargs):
        objlist = ShipEquipment.objects.values_list("equipment_id", flat=True)

        routine_map = defaultdict(int)
        routine_qs = MaintopDetail.objects.values("maintop_no").annotate(
            cnt=Count("frequency_f_key", distinct=True)
        )

        for row in routine_qs:
            m_no = row["maintop_no"]
            if m_no:
                routine_map[m_no] = row["cnt"]

        max_amendment_map = dict(
            MaintopDetail.objects.values("maintop_no")
            .annotate(max_amendment=Max("amendment_no"))
            .values_list("maintop_no", "max_amendment")
        )

        sub_counts = defaultdict(int)
        sub_qs = MaintopDetail.objects.values(
            "maintop_no",
            "amendment_no",
        ).annotate(cnt=Count("routine_id"))

        for row in sub_qs:
            m_no = row["maintop_no"]
            if m_no and row["amendment_no"] == max_amendment_map.get(m_no):
                sub_counts[m_no] = row["cnt"]

        dyd_map = defaultdict(int)
        dyd_qs = (
            MaintopDetail.objects.filter(by_whom__icontains="DYD")
            .values(
                "maintop_no",
                "amendment_no",
            )
            .annotate(cnt=Count("routine_id"))
        )

        for row in dyd_qs:
            m_no = row["maintop_no"]
            if m_no and row["amendment_no"] == max_amendment_map.get(m_no):
                dyd_map[m_no] = row["cnt"]

        obj = (
            Equipment.objects.filter(id__in=objlist)
            .annotate(
                pk=F("id"),
                equipment_name=F("equipment_class"),
                maintop_no_val=F("maintop_number"),
                eq_count=Count(
                    "ship_allocations",
                    distinct=True,
                ),
            )
            .values(
                "pk",
                "equipment_name",
                "equipment_code",
                "maintop_no_val",
                "eq_count",
            )
        )

        result = []

        equipment_names = set()
        maintop_numbers = set()

        for o in obj:
            m_no = o["maintop_no_val"]

            if o["equipment_name"]:
                equipment_names.add(o["equipment_name"])

            if m_no:
                maintop_numbers.add(m_no)

            result.append(
                {
                    "pk": o["pk"],
                    "equipment_name": o["equipment_name"],
                    "equipment_code": o["equipment_code"],
                    "maintop_no": m_no,
                    "eq_count": o["eq_count"],
                    "routine_count": routine_map.get(m_no, 0),
                    "sub_routine_count": sub_counts.get(m_no, 0),
                    "dyd_routine_count": dyd_map.get(m_no, 0),
                }
            )

        if not result:
            for ar in AddRoutineDetails.objects.select_related("equipment_name").all():
                m_no = ar.maintop_no
                eq_name = (
                    ar.equipment_name.name
                    if ar.equipment_name
                    else (ar.nomenclature or "")
                )
                if eq_name:
                    equipment_names.add(eq_name)
                if m_no:
                    maintop_numbers.add(m_no)

                result.append(
                    {
                        "pk": ar.pk,
                        "equipment_name": eq_name,
                        "equipment_code": ar.equipment_code or "",
                        "maintop_no": m_no or "",
                        "eq_count": 1,
                        "routine_count": routine_map.get(m_no, 1),
                        "sub_routine_count": sub_counts.get(m_no, 1),
                        "dyd_routine_count": dyd_map.get(m_no, 0),
                    }
                )

        maintop_qs = MaintopDetail.objects.filter(maintop_no__in=maintop_numbers)
        sub_departments = SubDepartment.objects.all().values("id", "name")
        three_months_ago = dj_timezone.now() - timedelta(days=90)

        filter_options = {
            "sub_departments": sorted(
                [
                    {
                        "id": sd["id"],
                        "name": sd["name"],
                    }
                    for sd in sub_departments
                ],
                key=lambda x: x["name"],
            ),
            "equipment_names": sorted(equipment_names),
            "routine_categories": ["All", "RUNNING HOUR BASED", "CALENDAR BASED"],
            "routine_names": list(
                maintop_qs.filter(
                    amendment_date__isnull=False,
                    amendment_date__gte=three_months_ago,
                )
                .values(
                    "routine_id",
                    "routine_brief_description",
                )
                .distinct()
                .order_by("routine_brief_description")
            ),
        }

        return Response(
            {
                "result": result,
                "filter_options": filter_options,
            }
        )


# ==================== DL1 DRAFT VIEWS ====================


@extend_schema(
    tags=["EMS"],
    request=GenerateDl1RequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
class GenerateDL1APIView(APIView):
    """
    GET:
        List routines marked as DL drafts.

    POST:
        Mark selected routines as DL drafts.
    """

    def get(self, request, *args, **kwargs):
        search = request.query_params.get("search", "").strip()

        queryset = (
            RoutineDescription.objects.filter(is_dl_draft=True)
            .select_related(
                "equipment_name",
                "routine_name",
            )
            .prefetch_related("ra_dl_entries")
        )

        if search:
            queryset = queryset.filter(
                Q(equipment_name__name__icontains=search)
                | Q(routine_name__name__icontains=search)
            )

        paginator = CustomPagination()
        page = paginator.paginate_queryset(queryset, request)

        dl_drafts = RoutineDescriptionSerializer(
            page,
            many=True,
        ).data

        refit_list = RefitMaintenancePeriodSerializer(
            RefitMaintenancePeriod.objects.all(),
            many=True,
        ).data

        return paginator.get_paginated_response(
            {
                "dl_drafts": dl_drafts,
                "refit_list": refit_list,
            }
        )

    def post(self, request, *args, **kwargs):
        data = request.data.copy()

        # Support pk_list, pk_list[], selected_ids, routine_ids, id_list, comma-separated string, or array
        if "pk_list[]" in data and "pk_list" not in data:
            data["pk_list"] = data.getlist("pk_list[]")

        raw_pk = (
            data.get("pk_list")
            or data.get("selected_ids")
            or data.get("routine_ids")
            or data.get("id_list")
        )

        parsed_list = []
        if isinstance(raw_pk, str):
            if raw_pk.strip().lower() == "all":
                parsed_list = list(
                    RoutineDescription.objects.values_list("id", flat=True)
                )
            else:
                parsed_list = [
                    int(x.strip()) for x in raw_pk.split(",") if x.strip().isdigit()
                ]
        elif isinstance(raw_pk, list):
            for item in raw_pk:
                if isinstance(item, str) and "," in item:
                    parsed_list.extend(
                        [int(x.strip()) for x in item.split(",") if x.strip().isdigit()]
                    )
                elif str(item).isdigit():
                    parsed_list.append(int(item))

        if not parsed_list and hasattr(data, "getlist"):
            values = data.getlist("pk_list") or data.getlist("selected_ids")
            for item in values:
                if isinstance(item, str) and "," in item:
                    parsed_list.extend(
                        [int(x.strip()) for x in item.split(",") if x.strip().isdigit()]
                    )
                elif str(item).isdigit():
                    parsed_list.append(int(item))

        if not parsed_list:
            # Fallback to all draft routines if no IDs passed
            parsed_list = list(
                RoutineDescription.objects.filter(is_dl_draft=True).values_list(
                    "id", flat=True
                )
            )

        data["pk_list"] = parsed_list

        serializer = GenerateDL1RequestSerializer(data=data)
        serializer.is_valid(raise_exception=True)

        id_list = serializer.validated_data["pk_list"]

        updated_count = RoutineDescription.objects.filter(id__in=id_list).update(
            is_dl_draft=True
        )

        return Response(
            {
                "status": "success",
                "message": (
                    f"DL Draft generated for {updated_count} routines successfully."
                ),
                "updated_count": updated_count,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(
    tags=["EMS"],
    request=PlanRoutineMultiSaveRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
class PlanRoutineMultiSaveAPIView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = PlanRoutineMultiSaveRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        raw_ids = (
            data.get("selected_ids")
            or request.data.get("routine_ids")
            or request.data.get("pk_list")
        )
        planned_date = data.get("planned_commencement_date")
        if planned_date == "":
            planned_date = None

        id_list = []
        if isinstance(raw_ids, str):
            id_list = [
                int(x.strip()) for x in raw_ids.split(",") if x.strip().isdigit()
            ]
        elif isinstance(raw_ids, list):
            for item in raw_ids:
                if isinstance(item, str) and "," in item:
                    id_list.extend(
                        [int(x.strip()) for x in item.split(",") if x.strip().isdigit()]
                    )
                elif str(item).isdigit():
                    id_list.append(int(item))

        planned_count = 0
        for pk in id_list:
            get_routine_obj = RoutineDescription.objects.filter(pk=pk).first()
            if get_routine_obj:
                PlannedRoutineDescription.objects.create(
                    routine_description_id=get_routine_obj,
                    spares_required=False,
                    planned_commencement_date=planned_date,
                )
                planned_count += 1

        return Response(
            {
                "status": "success",
                "message": f"Successfully planned {planned_count} routines.",
                "planned_count": planned_count,
                "planned_ids": id_list,
            },
            status=status.HTTP_200_OK,
        )


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def ems_generated_reports_list(request):
    report_rows = (
        RADLMaster.objects.select_related("refit_type_f_key")
        .annotate(
            total_dl_rows=Count(
                "radlroutinedescription",
                filter=Q(radlroutinedescription__dl_type="DL-I")
                & ~Q(radlroutinedescription__status="DELETED"),
                distinct=True,
            )
        )
        .filter(total_dl_rows__gt=0)
        .order_by("-id")
    )

    data = []
    for r in report_rows:
        data.append(
            {
                "id": r.id,
                "ra_dl_name": r.ra_dl_name or "-",
                "dockyard_name": r.dockyard_name or "-",
                "refit_type_name": r.refit_type_name or "-",
                "total_dl_rows": r.total_dl_rows,
            }
        )
    return Response({"report_rows": data})


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def ems_report_inner_rows(request, id):
    queryset = RADLRoutineDescription.objects.filter(radl_master__id=id)
    data = []
    for obj in queryset:
        data.append(
            {
                "id": obj.id,
                "dl_no": obj.dl_no,
                "ra_dl_name": obj.dl_key or "-",
                "dl_type": obj.dl_type,
                "status": obj.status,
                "eq_name": (
                    obj.routine_description.equipment_name.name
                    if obj.routine_description.equipment_name
                    else "-"
                ),
                "description": obj.routine_description.routine_description or "-",
                "routine_name": (
                    obj.routine_description.routine_name.name
                    if obj.routine_description.routine_name
                    else "-"
                ),
            }
        )
    return Response({"status": True, "data": data})


@extend_schema(
    tags=["EMS"],
    request=SaveDlDraftRowsRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def save_dl_draft_rows(request):
    try:
        serializer = SaveDlDraftRowsRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        rows_data = data.get("rows", [])
        yard = data.get("yard")
        refit_type_id = data.get("refit_type")

        master = None
        if yard or refit_type_id:
            refit_obj = None
            if refit_type_id:
                refit_obj = RefitMaintenancePeriod.objects.filter(
                    id=refit_type_id
                ).first()

            master = RADLMaster.objects.create(
                dockyard_name=yard,
                refit_type_f_key=refit_obj,
                refit_type_name=refit_obj.name if refit_obj else None,
            )

        for row in rows_data:
            routine_id = row.get("id")
            dl_no = row.get("dl_number")
            additional_remarks = row.get("additional_remarks")
            ss_remarks = row.get("remarks")

            routine = RoutineDescription.objects.get(id=routine_id)

            RADLRoutineDescription.objects.update_or_create(
                routine_description=routine,
                defaults={
                    "radl_master": master,
                    "dl_no": dl_no,
                    "additional_remarks": additional_remarks,
                    "remarks": ss_remarks,
                    "dl_type": "DL-I",
                    "status": "DRAFT",
                    "is_active": True,
                },
            )
        return Response({"status": "success"})
    except Exception as e:
        return Response(
            {"status": "error", "message": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


@extend_schema(
    tags=["EMS"],
    request=DeleteDlDraftRowRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def delete_dl_draft_row(request):
    try:
        serializer = DeleteDlDraftRowRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        routine_id = data.get("id")
        RoutineDescription.objects.filter(id=routine_id).update(is_dl_draft=False)
        RADLRoutineDescription.objects.filter(
            routine_description_id=routine_id
        ).delete()
        return Response({"status": "success"})
    except Exception as e:
        return Response(
            {"status": "error", "message": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


# ==================== FUEL SOUNDING VIEWS ====================
@extend_schema(
    tags=["EMS"],
    request=GetManualNamesRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="tank_type", type=str, description="Tank type", required=True
        )
    ],
)
@api_view(["GET", "POST"])
def get_manual_names(request):
    data = request.data if request.method == "POST" else request.query_params
    tank_type = data.get("tank_type")
    if not tank_type:
        return Response(
            {"success": False, "error": "tank_type required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    manual_names = (
        MeasurementFuelsondingFinal.objects.filter(tank_type=tank_type)
        .values_list("tank_name", flat=True)
        .distinct()
    )

    manual_names = [name for name in manual_names if name]
    return Response({"success": True, "manual_names": list(manual_names)})


@extend_schema(
    tags=["EMS"],
    request=LookupFuelSoundingRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["POST"])
def lookup_fuel_sounding(request):
    try:
        serializer = LookupFuelSoundingRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        mm_measurement = data.get("mm_measurement")
        if mm_measurement is None:
            mm_measurement = data.get("mm_value")
        if mm_measurement is None:
            return Response(
                {"success": False, "error": "mm_measurement or mm_value required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        mm_value = float(mm_measurement)
        tank_type = data.get("tank_type")
        manual_name = data.get("manual_name")

        volume, weight, interpolated, base_mm, remainder = calculate_volume_weight(
            mm_value, tank_type, manual_name
        )

        return Response(
            {
                "success": True,
                "volume": volume,
                "weight": weight,
                "interpolated": interpolated,
                "base_mm": base_mm,
                "remainder": remainder,
            }
        )
    except Exception as e:
        return Response(
            {"success": False, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )


# ─────────────────────────────────────────────────────────────
# CMMS FUSS & MAINTOP Integration Views
# ─────────────────────────────────────────────────────────────
@extend_schema(tags=["EMS"])
class FussSyncPayloadView(APIView):
    """Return an empty FUSS sync payload."""

    def get(self, request):
        serializer = FussSyncPayloadResponseSerializer(data={})
        serializer.is_valid()
        return Response(serializer.data)

    def post(self, request):
        """Validate a FUSS deferment payload and acknowledge receipt."""
        serializer = FussRaiseRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serial_no = f"FUSS-{date.today().year}-{serializer.validated_data.get('routine_description_id')}"
        resp_serializer = GenericSuccessResponseSerializer(
            data={
                "success": True,
                "message": "Deferment payload validated successfully.",
                "data": {"serial_no": serial_no},
            }
        )
        resp_serializer.is_valid()
        return Response(resp_serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=["EMS"])
class FussMastersView(APIView):
    """Return empty FUSS master lookup tables."""

    def get(self, request):
        serializer = FussMastersResponseSerializer(data={})
        serializer.is_valid()
        return Response(serializer.data)


@extend_schema(tags=["EMS"])
class MaintopSyncView(APIView):
    """Validate MAINTOP headers and details."""

    def post(self, request):
        serializer = MaintopSyncRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        val = serializer.validated_data
        from ems.tasks import validate_maintop_sync_task

        dispatch = dispatch_task(validate_maintop_sync_task, request.data)
        if dispatch["queued"]:
            return Response(
                {
                    "status": True,
                    "headers_processed": len(val.get("T_maintopheader", [])),
                    "details_processed": len(val.get("T_maintopdetail", [])),
                    "task_id": dispatch["task"].id,
                }
            )
        return sync_task_response(
            dispatch["result"],
            "Background task service unavailable. MAINTOP sync validation completed synchronously.",
        )


@extend_schema(tags=["EMS"])
class MaintopJICView(APIView):
    """Validate JIC data."""

    def post(self, request):
        serializer = MaintopJICRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        val = serializer.validated_data
        return Response(
            {
                "status": True,
                "jics_processed": len(val.get("T_maintopJIC", [])),
                "spares_processed": len(val.get("T_JICspares", [])),
                "tools_processed": len(val.get("T_JICtools", [])),
                "attachments_processed": len(val.get("T_JICattachments", [])),
            }
        )


@extend_schema(tags=["EMS"])
class MaintopDistributionView(APIView):
    """Validate distribution data."""

    def post(self, request):
        serializer = MaintopDistributionRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        val = serializer.validated_data
        return Response(
            {
                "status": True,
                "addresses_processed": len(val.get("M_address", [])),
                "distributions_processed": len(val.get("M_distribution_address", [])),
                "defaults_processed": len(val.get("T_MaintoplibraryDisDef", [])),
            }
        )


@extend_schema(tags=["EMS"])
class MaintopTaskStatusView(APIView):
    """Retrieve Celery task status for MAINTOP tasks."""

    def get(self, request, task_id):
        res = AsyncResult(task_id, app=celery_app)
        response_data = {
            "task_id": task_id,
            "status": res.status,
            "result": res.result if res.ready() else None,
        }
        return Response(response_data)


class _DelegatedFunctionAPIView(APIView):
    view_func = None

    def _call_view(self, request, *args, **kwargs):
        if type(self).view_func is None:
            raise NotImplementedError("view_func must be set on the subclass.")
        raw_request = getattr(request, "_request", request)
        return type(self).view_func(raw_request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        return self._call_view(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        return self._call_view(request, *args, **kwargs)

    def put(self, request, *args, **kwargs):
        return self._call_view(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self._call_view(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        return self._call_view(request, *args, **kwargs)


class GetSectionNameAPIView(_DelegatedFunctionAPIView):
    view_func = GetSectionNameView


class GetEquipmentNameAPIView(_DelegatedFunctionAPIView):
    view_func = GetEquipmentNameView


class GetSrarEquipmentNameAPIView(_DelegatedFunctionAPIView):
    view_func = GetsrarEquipmentNameView


class GetEquipmentNameWithoutRhsiNullRowsAPIView(_DelegatedFunctionAPIView):
    view_func = GetEquipmentNameWithoutRHSINullRowsView


class GetRoutineNameAPIView(_DelegatedFunctionAPIView):
    view_func = GetRoutineNameView


class SectionCreateAPIView(_DelegatedFunctionAPIView):
    view_func = SectionCreateView


class EquipmentCreateAPIView(_DelegatedFunctionAPIView):
    view_func = EquipmentCreateView


class TotalRunningHoursCreateAPIView(_DelegatedFunctionAPIView):
    view_func = TotalRunningHoursCreateView


class MonthlyRunningHoursSaveAPIView(_DelegatedFunctionAPIView):
    view_func = MonthlyRunningHoursSaveView


class GetEquipmentHistoryJsonAPIView(_DelegatedFunctionAPIView):
    view_func = GetEquipmentHistoryJSON


class UpdateEquipmentStateAPIView(_DelegatedFunctionAPIView):
    view_func = UpdateEquipmentStateView


class PlanRoutineAPIView(_DelegatedFunctionAPIView):
    view_func = Planroutine


class PlanRoutineSaveAPIView(APIView):
    @extend_schema(
        tags=["EMS"],
        request=PlanRoutineSaveRequestSerializer,
        responses={200: OpenApiTypes.OBJECT},
    )
    def post(self, request, pk=None, *args, **kwargs):
        raw_request = getattr(request, "_request", request)
        return plan_routine_save(raw_request, pk=pk)


class DeletePlannedRoutineAPIView(_DelegatedFunctionAPIView):
    view_func = DeletePlanned_routine


class SearchMergedAPIView(_DelegatedFunctionAPIView):
    view_func = SearchMergedView


class SearchDetailAPIView(_DelegatedFunctionAPIView):
    view_func = SearchDetailView


class PlannedRoutinesMasterFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = PlannedRoutinesMasterView


class GetRoutineNamesAPIView(_DelegatedFunctionAPIView):
    view_func = get_routine_names


class GetRoutineNameSingleAPIView(_DelegatedFunctionAPIView):
    view_func = get_routine_name


class GetEquipmentBySectionAPIView(_DelegatedFunctionAPIView):
    view_func = get_equipment_by_section


class FussRaisedDetailsAPIView(_DelegatedFunctionAPIView):
    view_func = fuss_raised_details


class FussRaisedRoutinesAPIView(_DelegatedFunctionAPIView):
    view_func = fuss_raised_routines


class RoutineHistoryAPIView(_DelegatedFunctionAPIView):
    view_func = routine_history


class EquipmentRunningHistoryAPIView(_DelegatedFunctionAPIView):
    view_func = equipment_running_history


class SlipHistoryAPIView(_DelegatedFunctionAPIView):
    view_func = slip_history


class MaintenancePlanAPIView(_DelegatedFunctionAPIView):
    view_func = maintenance_plan


class RoutineHistoryTimelineDataAPIView(_DelegatedFunctionAPIView):
    view_func = routine_history_timeline_data


class SearchDetailPlanAPIView(_DelegatedFunctionAPIView):
    view_func = SearchDetailView_plan


class CloseRaiseFussAPIView(_DelegatedFunctionAPIView):
    view_func = close_raise_fuss


class CompleteFussFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = complete_fuss


class SaveOemSpareFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = save_oem_spare


class ClosedRoutinesListAPIView(_DelegatedFunctionAPIView):
    view_func = closed_routines_list


class HistoryCloseRoutineAPIView(_DelegatedFunctionAPIView):
    view_func = history_close_routine


class GetRoutineInitializationDataFunctionAPIView(UniqueMaintopAPIView):
    pass


@extend_schema(
    tags=["EMS"],
    request=SaveRoutineInitializationRowRequestSerializer,
    responses={200: OpenApiTypes.OBJECT},
    examples=[
        OpenApiExample(
            "Routine Initialization Save Request Example",
            summary="Save a single routine initialization row",
            value={
                "routine_id": 1,
                "completion_date": "12 Apr 2026 12:36 PM",
                "undertaken_rh": "1250",
            },
            request_only=True,
        )
    ],
)
class SaveRoutineInitializationRowFunctionAPIView(APIView):
    serializer_class = SaveRoutineInitializationRowRequestSerializer

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            routine_id = data.get("routine_id") or data.get("pk") or data.get("id")
            if not routine_id:
                return Response(
                    {"status": "error", "message": "routine_id is required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            completion_date = data.get("completion_date") or data.get(
                "previous_completed_date"
            )
            undertaken_rh = data.get("undertaken_rh") or data.get(
                "previous_completed_at_rh"
            )

            try:
                routine = RoutineDescription.objects.get(pk=routine_id)
            except RoutineDescription.DoesNotExist:
                return Response(
                    {"status": "error", "message": "Routine description not found."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            if completion_date:
                try:
                    routine.previous_completed_date = pd.to_datetime(completion_date)
                except Exception:
                    routine.previous_completed_date = None
            else:
                routine.previous_completed_date = None

            routine.previous_completed_at_rh = (
                str(undertaken_rh)
                if undertaken_rh is not None and str(undertaken_rh).strip() != ""
                else None
            )
            routine.save()

            return Response(
                {"status": "success", "message": "Record updated successfully"}
            )
        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            description="AddRoutineDetails or Equipment primary key",
            required=False,
        ),
        OpenApiParameter(
            name="id",
            type=int,
            description="AddRoutineDetails or Equipment primary key",
            required=False,
        ),
    ],
)
class GetRoutineInitDetailsFunctionAPIView(APIView):
    def get(self, request, id=None, pk=None, *args, **kwargs):
        target_id = (
            id or pk or request.query_params.get("id") or request.query_params.get("pk")
        )
        if not target_id:
            return Response({"data": []}, status=status.HTTP_400_BAD_REQUEST)

        try:
            source_routine = AddRoutineDetails.objects.filter(pk=target_id).first()

            if not source_routine:
                source_routine = AddRoutineDetails.objects.filter(
                    equipment_name_id=target_id
                ).first()

            if not source_routine:
                sfd_eq = Equipment.objects.filter(pk=target_id).first()
                if sfd_eq:
                    source_routine = AddRoutineDetails.objects.filter(
                        Q(maintop_no=sfd_eq.maintop_number)
                        | Q(equipment_code=sfd_eq.equipment_code)
                        | Q(equipment_name__name__iexact=sfd_eq.equipment_class)
                    ).first()

            if not source_routine:
                source_routine = AddRoutineDetails.objects.filter(
                    Q(maintop_no=str(target_id)) | Q(routine_no=str(target_id))
                ).first()

            if source_routine:
                siblings = (
                    RoutineDescription.objects.filter(
                        Q(add_routine_details=source_routine)
                        | Q(maintop_no=source_routine.maintop_no),
                        is_close=False,
                        previous_completed_date__isnull=True,
                    )
                    .filter(
                        Q(previous_completed_at_rh__isnull=True)
                        | Q(previous_completed_at_rh__exact="")
                    )
                    .order_by("routine_no")
                )
            else:
                siblings = (
                    RoutineDescription.objects.filter(
                        Q(pk=target_id)
                        | Q(equipment_name_id=target_id)
                        | Q(maintop_no=str(target_id)),
                        is_close=False,
                        previous_completed_date__isnull=True,
                    )
                    .filter(
                        Q(previous_completed_at_rh__isnull=True)
                        | Q(previous_completed_at_rh__exact="")
                    )
                    .order_by("routine_no")
                )

            data = []
            for r in siblings:
                data.append(
                    {
                        "pk": r.pk,
                        "maintop_no": r.maintop_no or "-",
                        "category": (
                            source_routine.routine_category if source_routine else "-"
                        ),
                        "dart_number": r.dart_number or "-",
                        "routine_no": r.routine_no or "-",
                        "frequency": (
                            r.add_routine_details.frequency
                            if r.add_routine_details
                            else "-"
                        ),
                        "description": r.routine_description or "-",
                        "lst_completed_date": (
                            r.previous_completed_date.strftime("%Y-%m-%d")
                            if r.previous_completed_date
                            else "-"
                        ),
                        "lst_completed_rh": r.previous_completed_at_rh or "-",
                    }
                )

            if not data and source_routine:
                existing_rd = RoutineDescription.objects.filter(
                    Q(add_routine_details=source_routine)
                    | Q(maintop_no=source_routine.maintop_no)
                ).first()
                if (
                    not existing_rd
                    and source_routine.equipment_name
                    and source_routine.routine_name
                ):
                    existing_rd = RoutineDescription.objects.create(
                        add_routine_details=source_routine,
                        equipment_name=source_routine.equipment_name,
                        routine_name=source_routine.routine_name,
                        maintop_no=source_routine.maintop_no
                        or f"MT-{source_routine.pk}",
                        dart_number=getattr(source_routine, "dart_number", None)
                        or f"DART-{source_routine.routine_no or source_routine.pk}",
                        routine_no=source_routine.routine_no
                        or f"ROUTINE-{source_routine.pk}",
                        routine_description=getattr(source_routine, "remarks", None)
                        or f"Routine initialization for {source_routine.routine_name.name}",
                        by_whom=getattr(source_routine, "by_whom", None)
                        or "SHIP STAFF",
                        is_close=False,
                        previous_completed_date=None,
                        previous_completed_at_rh=None,
                    )
                if existing_rd:
                    data.append(
                        {
                            "pk": existing_rd.pk,
                            "maintop_no": existing_rd.maintop_no or "-",
                            "category": source_routine.routine_category or "-",
                            "dart_number": existing_rd.dart_number or "-",
                            "routine_no": existing_rd.routine_no or "-",
                            "frequency": source_routine.frequency or "-",
                            "description": existing_rd.routine_description or "-",
                            "lst_completed_date": (
                                existing_rd.previous_completed_date.strftime("%Y-%m-%d")
                                if existing_rd.previous_completed_date
                                else "-"
                            ),
                            "lst_completed_rh": existing_rd.previous_completed_at_rh
                            or "-",
                        }
                    )

            eq_name = ""
            routine_type = ""
            if source_routine:
                eq_name = (
                    source_routine.equipment_name.name
                    if source_routine.equipment_name
                    else (source_routine.nomenclature or "")
                )
                routine_type = (
                    source_routine.routine_name.name
                    if source_routine.routine_name
                    else (source_routine.routine_category or "")
                )
            elif siblings.exists():
                first_sib = siblings.first()
                eq_name = (
                    first_sib.equipment_name.name if first_sib.equipment_name else ""
                )
                routine_type = (
                    first_sib.routine_name.name if first_sib.routine_name else ""
                )

            return Response(
                {
                    "status": "success",
                    "data": data,
                    "equipment": eq_name,
                    "routine_type": routine_type,
                }
            )
        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class SaveRoutineInitBulkFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = save_routine_init_bulk


class EmsGeneratedReportsListFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = ems_generated_reports_list


class EmsReportInnerRowsFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = ems_report_inner_rows


class SaveDlDraftRowsFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = save_dl_draft_rows


class DeleteDlDraftRowFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = delete_dl_draft_row


class GetManualNamesFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = get_manual_names


class LookupFuelSoundingFunctionAPIView(_DelegatedFunctionAPIView):
    view_func = lookup_fuel_sounding


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.BINARY})
@api_view(["GET"])
def maintop_pdf(request, pk):
    equipment = Equipment.objects.filter(id=pk).first()
    if not equipment:
        return HttpResponse("Equipment not found", status=404)

    header = MaintopHeader.objects.filter(maintop_no=equipment.maintop_number).first()

    context = {}
    if header:
        latest_amendment = MaintopDetail.objects.filter(
            maintop_no=equipment.maintop_number,
            maintopheader_f_key=header,
        ).aggregate(max_amend=Max("amendment_no"))["max_amend"]

        details = (
            MaintopDetail.objects.filter(
                maintop_no=equipment.maintop_number,
                amendment_no=latest_amendment,
            )
            .annotate(freq_full_name=F("frequency_f_key__description"))
            .order_by("frequency", "routine_no")
        )

        grouped_data = {}
        for detail in details:
            grouped_data.setdefault(detail.frequency or "OTHER", []).append(detail)

        context = {
            "equipment": equipment,
            "header": header,
            "grouped_data": grouped_data,
        }

    html = render_to_string("ems/maintop_pdf.html", context)

    base_path = settings.BASE_DIR
    options = {
        "header-html": f"file://{base_path}/ems/templates/ems/maintop_pdf_header.html",
        "footer-html": f"file://{base_path}/ems/templates/ems/maintop_pdf_footer.html",
        "margin-top": "35mm",
        "header-spacing": "10",
        "margin-bottom": "35mm",
        "footer-spacing": "8",
        "enable-local-file-access": None,
    }
    pdf = pdfkit.from_string(html, False, options=options)
    return HttpResponse(pdf, content_type="application/pdf")


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.BINARY})
@api_view(["GET"])
def fuss_pdf(request, pk):
    fuss = get_object_or_404(FussRaiseDetails, pk=pk)
    context = {"fuss": fuss, "today": dj_timezone.now()}
    html = render_to_string("ems/fuss_pdf.html", context)

    options = {
        "margin-top": "20mm",
        "margin-bottom": "20mm",
        "margin-left": "10mm",
        "margin-right": "10mm",
        "enable-local-file-access": None,
        "encoding": "UTF-8",
    }
    pdf = pdfkit.from_string(html, False, options=options)

    filename = f"Fuss_Initiation_{fuss.pk}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


def _build_less_routine_result(routines, current_time):
    result = []
    for routine in routines:
        equipment = routine.equipment_name
        start_timedate = equipment.start_timedate
        if start_timedate and is_naive(start_timedate):
            start_timedate = make_aware(start_timedate)

        if (
            equipment.state == "ACTIVE"
            and start_timedate
            and equipment.rhsi is not None
        ):
            dynamic_rhsi = round(
                equipment.rhsi + (current_time - start_timedate).total_seconds() / 3600,
                1,
            )
        else:
            dynamic_rhsi = equipment.rhsi

        total_routines = LessAddRoutineDetails.objects.filter(
            routine_name=routine.routine_name,
            equipment_name=equipment,
        ).count()
        dyd_routines = LessAddRoutineDetails.objects.filter(
            routine_name=routine.routine_name,
            equipment_name=equipment,
            by_whom="DYD",
        ).count()
        routine_other_details = LessRoutineDescription.objects.filter(
            add_routine_details=routine.id
        ).first()

        result.append(
            {
                "pk": routine.pk,
                "routine_name": routine.routine_name.name,
                "section": equipment.section.name if equipment.section else "",
                "equipment_name": equipment.name,
                "maintop_no": (
                    routine_other_details.maintop_no if routine_other_details else "NA"
                ),
                "last_routine_date": (
                    routine.last_routine_completion_date.strftime("%d %b, %Y")
                    if routine.last_routine_completion_date
                    else "NA"
                ),
                "date": "NA",
                "last_routine_running_hrs": (
                    routine.last_routine_completion_atrunning_hrs
                    if routine.last_routine_completion_atrunning_hrs is not None
                    else "NA"
                ),
                "next_due_running_hrs": "NA",
                "total_running_hrs": (
                    round(dynamic_rhsi, 2) if dynamic_rhsi is not None else "NA"
                ),
                "running_hrs_updated_tilldate": (
                    equipment.rhsi_updated_until.strftime("%d %b, %Y")
                    if equipment.rhsi_updated_until
                    else "NA"
                ),
                "running_hrs_available": "NA",
                "total_routines": total_routines,
                "dyd_routines": dyd_routines,
                "remarks": routine.remarks,
                "routine_category": routine.routine_category,
            }
        )
    return result


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(name="section", type=str, required=False),
        OpenApiParameter(name="equipment_name", type=str, required=False),
        OpenApiParameter(name="routine_category", type=str, required=False),
    ],
)
@api_view(["GET"])
def less_routine_search(request):
    current_time = make_aware(datetime.now())
    department_id = request.user.department_id

    routines = LessAddRoutineDetails.objects.select_related(
        "equipment_name", "equipment_name__section", "routine_name"
    ).filter(equipment_name__section__department_id=department_id)

    section_id = request.query_params.get("section")
    equipment_name_id = request.query_params.get("equipment_name")
    routine_category = request.query_params.get("routine_category")

    if section_id and section_id != "0":
        routines = routines.filter(equipment_name__section_id=section_id)
    if equipment_name_id and equipment_name_id != "0":
        routines = routines.filter(equipment_name_id=equipment_name_id)
    if routine_category and routine_category != "0":
        routines = routines.filter(routine_category=routine_category)
    else:
        routines = routines.exclude(routine_name__name__icontains="refit").exclude(
            routine_category="ALTERNATE PERIODIC"
        )

    routine_desc_qs = LessRoutineDescription.objects.filter(
        equipment_name=OuterRef("equipment_name"),
        routine_name=OuterRef("routine_name"),
    )
    routines = routines.annotate(has_routine_desc=Exists(routine_desc_qs)).filter(
        has_routine_desc=True
    )
    routines = routines.order_by("equipment_name__name")

    result = _build_less_routine_result(routines, current_time)
    return Response({"result": result})


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(
            name="pk",
            type=int,
            location=OpenApiParameter.PATH,
            required=False,
        )
    ],
)
@api_view(["GET"])
def less_routine_search_detail(request, pk=None):
    if not pk:
        return Response({"result": []})

    routine_detail = get_object_or_404(LessAddRoutineDetails, pk=pk)
    routine_descriptions = LessRoutineDescription.objects.filter(
        equipment_name=routine_detail.equipment_name,
        routine_name=routine_detail.routine_name,
    ).order_by("routine_no")

    result = [
        {
            "pk": description.pk,
            "routine_name": description.routine_name.name,
            "equipment_name": description.equipment_name.name,
            "maintop_no": description.maintop_no,
            "dart_number": description.dart_number,
            "routine_description": description.routine_description,
            "routine_no": description.routine_no,
            "action_by": description.by_whom,
        }
        for description in routine_descriptions
    ]
    return Response({"result": result, "id": routine_detail.pk})


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.BINARY})
@api_view(["GET"])
def export_routine_excel(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Routine Details"

    headers = [
        "ShipName",
        "MaintopNo",
        "Nomenclature",
        "Frequency",
        "RoutineName",
        "RoutineCategory",
        "Sub Depatment",
        "ClassName",
        "EquipmentName",
        "EquipmentCode",
        "RoutineNo",
        "RoutineDescription",
        "ByWhom",
        "frequency in months",
        "frequency in hours",
        "RHSI",
        "RHSIUpdatedUpto",
        "LastRoutineCompletionDate",
        "LastRoutineCompletedAtRH",
    ]
    worksheet.append(headers)

    department_id = request.user.department_id
    queryset = AddRoutineDetails.objects.select_related(
        "ship", "equipment_name__section__department", "routine_name"
    )
    if department_id:
        queryset = queryset.filter(equipment_name__section__department_id=department_id)

    seen_pairs = set()
    for routine in queryset:
        section = routine.equipment_name.section if routine.equipment_name else None
        nomenclature = (
            routine.nomenclature.strip().lower() if routine.nomenclature else None
        )
        if section is None or nomenclature is None:
            continue

        pair_key = (section.id, nomenclature)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)

        routine_desc = RoutineDescription.objects.filter(
            add_routine_details=routine
        ).first()
        worksheet.append(
            [
                "INS KOCHI",
                routine.maintop_no or "",
                routine.nomenclature or "",
                routine.frequency or "",
                routine.routine_name.name if routine.routine_name else "",
                routine.routine_category or "",
                section.name or "",
                routine.class_name or "",
                routine.equipment_name.name if routine.equipment_name else "",
                (
                    routine.equipment_name.equipment_code
                    if routine.equipment_name
                    else ""
                ),
                routine_desc.routine_no if routine_desc else "",
                routine_desc.routine_description if routine_desc else "",
                routine_desc.by_whom if routine_desc else "",
                routine.frequency_in_months or "",
                routine.frequency_in_hours or "",
                routine.rhs_i or "",
                routine.rhs_i_updated_upto or "",
                "",
                "",
            ]
        )

    for i in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    bold_font = Font(bold=True)
    for col_num, cell in enumerate(worksheet[1], start=1):
        cell.font = bold_font
        if col_num > len(headers) - 2:
            cell.fill = red_fill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=RoutineDetails.xlsx"
    workbook.save(response)
    return response


@extend_schema(
    summary="DL1 Access-DB export (Disabled)",
    description=(
        "Pushes drafted DL1 rows into a shipyard-specific legacy MS-Access "
        "(.accdb) file via a JDBC/jpype bridge. Disabled because that "
        "Windows/Java-specific integration is not part of this cross-platform "
        "deployment, mirroring the sfd app's CMMS sync endpoints."
    ),
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["GET", "POST"])
def dl1_accdb_export_disabled(request):
    return Response(
        {
            "status": "disabled",
            "message": (
                "DL1 Access-DB export is currently disabled to ensure "
                "cross-platform compatibility (the JDBC/jpype bridge to the "
                "legacy .accdb file this endpoint depended on is not part of "
                "this deployment)."
            ),
        },
        status=status.HTTP_501_NOT_IMPLEMENTED,
    )


@extend_schema(
    summary="Routine CMMS sync (Disabled)",
    description=(
        "Fetches routine/missing-routine data from the legacy CMMS via the "
        "swmmapi bridge. Disabled because that pyodbc-based bridge is not "
        "part of this deployment, mirroring the sfd app's CMMS sync endpoints."
    ),
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
)
@api_view(["GET", "POST"])
def routine_cmms_sync_disabled(request):
    return Response(
        {
            "status": "disabled",
            "message": (
                "Routine CMMS synchronisation is currently disabled to ensure "
                "cross-platform compatibility (the pyodbc-based swmmapi bridge "
                "this endpoint depended on is not part of this deployment)."
            ),
        },
        status=status.HTTP_501_NOT_IMPLEMENTED,
    )


def _linear_fit_coeffs(x_data, y_data):
    """Least-squares fit for y = a*x + b (numpy.polyfit avoids adding scipy)."""
    a, b = np.polyfit(x_data, y_data, 1)
    return a, b


def _quadratic_fit_coeffs(x_data, y_data):
    """Least-squares fit for y = a*x + b*x^2 + c, matching the reference's
    (non-standard-order) quadratic_fit signature.
    """
    p_x2, p_x1, p_x0 = np.polyfit(x_data, y_data, 2)
    return p_x1, p_x2, p_x0


def _return_coeff_lpc(gt_name_id):
    data_set_lpc = DataPointsGTLPC.objects.filter(gt_name_id=gt_name_id)
    if data_set_lpc.count() < 3:
        return None

    x_data = np.array([value.hpc_rpm for value in data_set_lpc])
    y_data = np.array([value.lpc_rpm for value in data_set_lpc])
    amb_temp = data_set_lpc.first().amb_temp
    a, b = _linear_fit_coeffs(x_data, y_data)
    return a, b, amb_temp


def _return_coeff_extair(gt_name_id):
    data_set_air = DataPointsAirprHPC.objects.filter(gt_name_id=gt_name_id)
    data_set_ext = DataPointsExtTemp.objects.filter(gt_name_id=gt_name_id)
    if data_set_air.count() < 3 or data_set_ext.count() < 3:
        return None

    x_air = np.array([value.hpc_rpm for value in data_set_air])
    y_air = np.array([value.air_pr_hpc for value in data_set_air])
    x_ext = np.array([value.hpc_rpm for value in data_set_ext])
    y_ext = np.array([value.ext_temp for value in data_set_ext])
    amb_pressure = data_set_air.first().amb_pressure

    a_air, b_air, c_air = _quadratic_fit_coeffs(x_air, y_air)
    a_ext, b_ext, c_ext = _quadratic_fit_coeffs(x_ext, y_ext)
    return a_air, b_air, c_air, a_ext, b_ext, c_ext, amb_pressure


@extend_schema(tags=["EMS"], request=SaveOemDataPointRequestSerializer)
@api_view(["POST"])
def save_oem_data_point(request):
    serializer = SaveOemDataPointRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    gt_name = get_object_or_404(EquipmentName, pk=data["gt_name"])
    dataset = data["dataset"]

    if dataset == "lpc":
        existing = DataPointsGTLPC.objects.filter(
            gt_name=gt_name, hpc_rpm=data.get("hpc_rpm")
        ).first()
        if existing:
            return Response(
                {"message": "Entry already exists."}, status=status.HTTP_400_BAD_REQUEST
            )
        DataPointsGTLPC.objects.create(
            gt_name=gt_name,
            hpc_rpm=data.get("hpc_rpm"),
            lpc_rpm=data.get("lpc_rpm"),
            amb_temp=data.get("amb_temp"),
        )
    elif dataset == "air":
        existing = DataPointsAirprHPC.objects.filter(
            gt_name=gt_name, hpc_rpm=data.get("hpc_rpm")
        ).first()
        if existing:
            return Response(
                {"message": "Entry already exists."}, status=status.HTTP_400_BAD_REQUEST
            )
        DataPointsAirprHPC.objects.create(
            gt_name=gt_name,
            hpc_rpm=data.get("hpc_rpm"),
            air_pr_hpc=data.get("air_pr_hpc"),
            amb_pressure=data.get("amb_pressure"),
        )
    elif dataset == "ext":
        existing = DataPointsExtTemp.objects.filter(
            gt_name=gt_name, hpc_rpm=data.get("hpc_rpm")
        ).first()
        if existing:
            return Response(
                {"message": "Entry already exists."}, status=status.HTTP_400_BAD_REQUEST
            )
        DataPointsExtTemp.objects.create(
            gt_name=gt_name,
            hpc_rpm=data.get("hpc_rpm"),
            ext_temp=data.get("ext_temp"),
            amb_temp=data.get("amb_temp"),
        )
    else:
        existing = DataPointsExtTempGTG.objects.filter(
            gt_name=gt_name, el_load=data.get("el_load")
        ).first()
        if existing:
            return Response(
                {"message": "Entry already exists."}, status=status.HTTP_400_BAD_REQUEST
            )
        DataPointsExtTempGTG.objects.create(
            gt_name=gt_name,
            el_load=data.get("el_load"),
            ext_temp=data.get("ext_temp"),
            amb_temp=data.get("amb_temp"),
        )

    return Response(
        {"message": "New data point created."}, status=status.HTTP_201_CREATED
    )


@extend_schema(tags=["EMS"], request=SaveSlipLimitRequestSerializer)
@api_view(["POST"])
def save_slip_limit(request):
    serializer = SaveSlipLimitRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    gt_name = get_object_or_404(EquipmentName, pk=data["gt_name"])
    existing = SlipLimit.objects.filter(gt_name=gt_name).first()
    if existing:
        existing.delta_n_lpc = data.get("delta_n_lpc")
        existing.delta_t_ext = data.get("delta_t_ext")
        existing.delta_p_air = data.get("delta_p_air")
        existing.save(update_fields=["delta_n_lpc", "delta_t_ext", "delta_p_air"])
        return Response({"message": "Slip Limits for selected equipment updated."})

    SlipLimit.objects.create(
        gt_name=gt_name,
        delta_n_lpc=data.get("delta_n_lpc"),
        delta_t_ext=data.get("delta_t_ext"),
        delta_p_air=data.get("delta_p_air"),
    )
    return Response(
        {"message": "New data point created."}, status=status.HTTP_201_CREATED
    )


@extend_schema(tags=["EMS"], request=CalculateGtSlipRequestSerializer)
@api_view(["POST"])
def calculate_gt_slip(request):
    serializer = CalculateGtSlipRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    gt_name = get_object_or_404(EquipmentName, pk=data["gt_name"])
    recorded_lpc = data["recorded_lpc"]
    recorded_air_pr_after_hpc = data["recorded_air_pr_after_hpc"]
    recorded_amb_pr_gtinlet = data["recorded_amb_pr_gtinlet"]
    recorded_ext_temp = data["recorded_ext_temp"]
    at_hpc_rpm = data["at_hpc_rpm"]
    current_amb_temp = data["current_amb_temp"]

    data_lpc = _return_coeff_lpc(gt_name.pk)
    data_airext = _return_coeff_extair(gt_name.pk)
    if data_lpc is None or data_airext is None:
        return Response(
            {
                "message": (
                    "Enter at least three OEM Graph data points of HPC air "
                    "pressure, exhaust temperature, and LPC rpm before "
                    "attempting to calculate slip."
                )
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    graph_temp, graph_pressure = data_lpc[2], data_airext[6]

    alpha = math.sqrt((graph_temp + 273) / (273 + current_amb_temp))
    actual_lpc_at_std_temp = recorded_lpc * alpha
    at_hpc_rpm_at_std_temp = at_hpc_rpm * alpha

    beta = graph_pressure / recorded_amb_pr_gtinlet
    actual_air_pr_at_std_temp = recorded_air_pr_after_hpc * beta

    lpc_at_std_temp_graph = data_lpc[0] * at_hpc_rpm_at_std_temp + data_lpc[1]
    air_pr_at_std_temp_graph = (
        data_airext[0] * at_hpc_rpm_at_std_temp
        + data_airext[1] * (at_hpc_rpm_at_std_temp**2)
        + data_airext[2]
    )
    ext_temp_at_std_temp_graph = (
        data_airext[3] * at_hpc_rpm_at_std_temp
        + data_airext[4] * (at_hpc_rpm_at_std_temp**2)
        + data_airext[5]
    )

    alpha_square = (graph_temp + 273) / (273 + current_amb_temp)
    actual_ext_temp_at_std_temp = (recorded_ext_temp + 273) * alpha_square - 273

    slip_lpc = lpc_at_std_temp_graph - actual_lpc_at_std_temp
    slip_air = air_pr_at_std_temp_graph - actual_air_pr_at_std_temp
    slip_ext = actual_ext_temp_at_std_temp - ext_temp_at_std_temp_graph

    slip_limit = SlipLimit.objects.filter(gt_name=gt_name).first()
    if not slip_limit:
        return Response(
            {"message": "Enter Slip Limits before calculating slip."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    gtname = gt_name.name
    return Response(
        {
            "message": f"Slip of {gtname} calculated successfully",
            "msg1": f"Slip of {gtname} calculated at {at_hpc_rpm} HPC RPM :-",
            "msg2": f"    Slip by LPC RPM is {slip_lpc}",
            "msg3": f"    Slip by Air pressure after HPC is {slip_air}",
            "msg4": f"    Slip by GT Ext Temp is {slip_ext}",
            "slip_lpc": slip_lpc,
            "slip_air": slip_air,
            "slip_ext": slip_ext,
            "slip_limit_lpc": slip_limit.delta_n_lpc,
            "slip_limit_air": slip_limit.delta_t_ext,
            "slip_limit_ext": slip_limit.delta_p_air,
        }
    )


@extend_schema(tags=["EMS"], request=CalculateGtSlipRequestSerializer)
@api_view(["POST"])
def save_gt_slip(request):
    """Same as calculate_gt_slip but also persists the result, matching the
    reference's (previously disabled) SaveSlipView.
    """
    calc_response = calculate_gt_slip(request._request)
    if calc_response.status_code != status.HTTP_200_OK:
        return calc_response

    data = CalculateGtSlipRequestSerializer(data=request.data)
    data.is_valid(raise_exception=True)
    validated = data.validated_data

    PostCalculateLPC.objects.create(
        gt_name_id=validated["gt_name"],
        recorded_lpc=validated["recorded_lpc"],
        recorded_air_pr_after_hpc=validated["recorded_air_pr_after_hpc"],
        recorded_amb_pr_gtinlet=validated["recorded_amb_pr_gtinlet"],
        recorded_ext_temp=validated["recorded_ext_temp"],
        at_hpc_rpm=validated["at_hpc_rpm"],
        current_amb_temp=validated["current_amb_temp"],
        calculated_lpc_slip=calc_response.data["slip_lpc"],
        calculated_air_slip=calc_response.data["slip_air"],
        calculated_ext_slip=calc_response.data["slip_ext"],
    )
    return calc_response


@extend_schema(
    tags=["EMS"],
    responses={200: OpenApiTypes.OBJECT},
    parameters=[
        OpenApiParameter(name="gt_name", type=int, required=True),
        OpenApiParameter(name="dataset", type=str, required=True),
    ],
)
@api_view(["GET"])
def oem_graph_data(request):
    gt_name_id = request.query_params.get("gt_name")
    dataset = request.query_params.get("dataset")
    if not gt_name_id or not dataset:
        raise ValidationError(
            {"detail": "gt_name and dataset query params are required."}
        )

    model_map = {
        "lpc": (DataPointsGTLPC, "hpc_rpm", "lpc_rpm"),
        "air": (DataPointsAirprHPC, "hpc_rpm", "air_pr_hpc"),
        "ext": (DataPointsExtTemp, "hpc_rpm", "ext_temp"),
        "ext_gtg": (DataPointsExtTempGTG, "el_load", "ext_temp"),
    }
    if dataset not in model_map:
        raise ValidationError({"dataset": f"Must be one of {list(model_map)}."})

    model, x_field, y_field = model_map[dataset]
    queryset = model.objects.filter(gt_name_id=gt_name_id)
    points = [
        {"x": getattr(entry, x_field), "y": getattr(entry, y_field)}
        for entry in queryset
    ]

    fit = None
    if len(points) >= 3:
        x_data = np.array([p["x"] for p in points])
        y_data = np.array([p["y"] for p in points])
        if dataset == "lpc":
            a, b = _linear_fit_coeffs(x_data, y_data)
            fit = {"type": "linear", "a": a, "b": b}
        else:
            a, b, c = _quadratic_fit_coeffs(x_data, y_data)
            fit = {"type": "quadratic", "a": a, "b": b, "c": c}

    return Response({"points": points, "fit": fit})


@extend_schema(tags=["EMS"], responses={200: OpenApiTypes.OBJECT})
@api_view(["GET"])
def combined_routine_search(request):
    main_result = SearchMergedView(request._request).data
    less_result = less_routine_search(request._request).data.get("result", [])
    return Response({"result": list(main_result) + list(less_result)})
